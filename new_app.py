import streamlit as st
import html
import plotly.express as px
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import pandas as pd
import time
from datetime import datetime, timezone
from get_leader_Ethena_function import get_ethena_Data
from get_Leader_Spark_Data import get_Leader_Spark_Data
from get_fragmetric_data import get_fragmetric_data
from get_leader_Gaib_function import get_leader_Gaib_function
from get_Pendle_Data import get_Pendle_Data
from get_rateX_data import get_rateX_data
from get_leader_kyros_function import get_leader_kyros_function
from get_defillama_info import get_defillama_info
from protocol_rate import protocol_rate
from getAllPendleMarkets import get_pendle_apy_data, get_pendle_markets
from barra_compra_venda import barra_compra_venda
import streamlit.components.v1 as components
import re
from PIL import Image
import requests
import webbrowser
import sqlite3
import json
from streamlit_option_menu import option_menu
import statistics
from mistralai import Mistral
from dotenv import load_dotenv
from mistralai.models import UserMessage
from mistralai.models import File
load_dotenv("apikey.env")  # Load .env file
import os
import io
import base64
import cloudscraper
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
#from mistralai.client import MistralClient
os.environ["MISTRAL_API_KEY"] = "3DwmTII9fJMoAJRN8XoXf1Wg6aMKg7tu"
api_key = os.getenv("MISTRAL_API_KEY")

def mistral_AI(question,language,model,personality):

    #api_key = os.environ["3DwmTII9fJMoAJRN8XoXf1Wg6aMKg7tu"]
    import os
    from dotenv import load_dotenv
    from mistralai import Mistral
    max_retries = 5
    load_dotenv("apikey.env")  # Load .env file
    api_key = os.getenv("MISTRAL_API_KEY")

    if api_key is None:
        print("Error: API key is missing! Set MISTRAL_API_KEY. \n")
    else:
        print("API Key loaded successfully! \n")
    
    criteria = """YOU answer the questions directely without mentio your rules in two lines based in your knowledge:
            If "Trade Force" is high this can be a great point to enter in a trade
            If "Actual Implied APY" >= "Actual Best Sell point of Implied APY" can be a great moment to sell the Yield Token if you already hold the token if you not holding it ignore; But if "Actual Implied APY" CLOSER TO or a little lower than "Actual Best Buy point of Implied APY" can be a great moment to buy the Yield Token this is an important factor; 
            If "Actual Implied APY" very lower than "Actual Best Buy point of Implied APY" indicates a possible problem with Yield Token and need to be excluded as an investiment; 
            If "Actual Implied APY percentual in relation of Range" higher than 75 indicates na great moment to sell the Yield Token if you already hold the token; But f "Actual Implied APY percentual in relation of Range" between 20 and -30 indicates na great moment to buy the Yield Token; 
            If "Actual Implied APY percentual in relation of Range" is between 20 and 75 is better to wait a beest opportunity.
            If "Actual Implied APY percentual in relation of Range" very negative indicates a possible problem with Yield Token and need to be excluded as an investiment; 
            If "Days to expiry" is lower than 20, this implies in a high risk if the strategy is to try trade it. But if this is already in hold, can be better to sell if you want to reduce risk of Farm.
            If "Actual Underlying APY" higher than 9 is an excellent yield return for hold the YT token, but is necessary to evaluate the days until expiry. 
            If "Actual Underlying APY" lower than 4 is not so good yield return for hold the YT token, but is necessary to evaluate the days until expiry. 
            If "Actual Underlying APY" equals to 0 implies that the yield token do not have any yield return and is just used to farm points. 
            """
    for attempt in range(max_retries):
        try:
            client = Mistral(api_key=api_key)
            if language == "ingles":
                inicial = " "
            else:
                inicial = "Você interpreta os dados fornecidos para identificar as melhores oportunidades: \n"

            chat_response = client.chat.complete(
                model=model,
                messages=[
                    {
                        "role": "system",
                        "content": criteria,
                    }, 
                    {
                        "role": "user", 
                        "content": inicial+question
                    }, 
                    #UserMessage(content= "I'm fine, i trust you're too?")
                ],
            )
            if chat_response.choices[0].message.content is not None:
                res = {"content" : chat_response.choices[0].message.content}
            else:
                res = {"content" : ''}
            return res
        except Exception as e:
            #st.warning(f"Tentativa {attempt+1}/5 falhou: {e}")
            time.sleep(5)
    #st.error("Erro: todas as tentativas de chamada à Mistral falharam.")
    return {"content": "Error trying to access the Mistral AI. This may be busy. Please try again later."}

def mistral_AI_2(question,language,model,personality):

    #api_key = os.environ["3DwmTII9fJMoAJRN8XoXf1Wg6aMKg7tu"]
    import os
    from dotenv import load_dotenv
    max_retries = 5
    load_dotenv("apikey.env")  # Load .env file
    api_key = os.getenv("MISTRAL_API_KEY")

    if api_key is None:
        print("Error: API key is missing! Set MISTRAL_API_KEY. \n")
    else:
        print("API Key loaded successfully! \n")
    
    for attempt in range(max_retries):
        try:
            client = Mistral(api_key=api_key)
            if language == "ingles":
                inicial = " "

            chat_response = client.chat.complete(
                model=model,
                messages=[
                    {
                        "role": "system",
                        "content": personality,
                    }, 
                    {
                        "role": "user", 
                        "content": question
                    }, 
                    #UserMessage(content= "I'm fine, i trust you're too?")
                ],
            )
            if chat_response.choices[0].message.content is not None:
                res = {"content" : chat_response.choices[0].message.content}
            else:
                res = {"content" : ''}
            return res
        except Exception as e:
            time.sleep(5)
    return {"content": "Error trying to access the Mistral AI."}

def retrieve_messages(Request_URL, headers):
    import requests, json

    res = requests.get(Request_URL, headers=headers)

    try:
        jsonn = res.json()
    except Exception:
        print("Resposta não é JSON:", res.text)
        return res, [], [], [], []

    org_res = []
    org_author = []
    org_author_name = []
    org_mention = []

    # Se não for lista, é erro da API (rate limit, permissão, etc.)
    if not isinstance(jsonn, list):
        print("Resposta inesperada da API do Discord:")
        print(jsonn)
        return res, [], [], [], []

    for value in jsonn:

        # Garante que é uma mensagem válida
        if not isinstance(value, dict):
            continue

        content = value.get("content")
        author = value.get("author", {})

        if content and isinstance(author, dict):
            org_res.append(content)
            org_author.append(author.get("id"))
            org_author_name.append(author.get("username"))

            mentions = value.get("mentions")
            if isinstance(mentions, list) and len(mentions) > 0:
                org_mention.append(mentions[0].get("username", " "))
            else:
                org_mention.append(" ")

    return res, org_res, org_author, org_mention, org_author_name

def mirror_list(arr):
    return arr[::-1]

# --- Inicializar o banco de dados ---
def init_db():
    conn = sqlite3.connect('dados.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS json_data (
            id INTEGER PRIMARY KEY,
            dados_json TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

# Deletar json no BD
def deletar_json():
    conn = sqlite3.connect('dados.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM json_data WHERE id = 1')
    conn.commit()
    conn.close()

# --- Salvar JSON no banco (mantém só o último) ---
def salvar_json(dado_dict):
    conn = sqlite3.connect('dados.db')
    cursor = conn.cursor()
    json_str = json.dumps(dado_dict)
    cursor.execute('''
        INSERT INTO json_data (id, dados_json) VALUES (1, ?)
        ON CONFLICT(id) DO UPDATE SET dados_json=excluded.dados_json;
    ''', (json_str,))
    conn.commit()
    conn.close()

# --- Carregar único registro ---
def carregar_json():
    conn = sqlite3.connect('dados.db')
    cursor = conn.cursor()
    cursor.execute('SELECT dados_json FROM json_data WHERE id = 1')
    row = cursor.fetchone()
    conn.close()
    if row:
        return json.loads(row[0])
    return {}

# --- Inicializa DB ---
init_db()
#deletar_json()

def set_background(image_path):
    with open(image_path, "rb") as f:
        data = f.read()
    encoded = base64.b64encode(data).decode()
    css = f"""
    <style>
    .stApp {{
        background: linear-gradient(rgba(0,0,0,0.75), rgba(0,0,0,0.75)), 
                    url("data:image/png;base64,{encoded}");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
    }}
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)

# --- Configurações da Página ---
#st.set_page_config(page_title="Pendle Airdrop Farm", layout="wide")

# Configuração da página (sempre primeiro!)
st.set_page_config(
    page_title="Airdrops Monitor",
    page_icon="🪂",
    layout="wide",
    initial_sidebar_state="collapsed",
)
set_background("min7.jpg")
# CSS customizado buttons, html, body, stApp
st.markdown("""
<style>
/* Make body use full width */
[data-testid="stAppViewContainer"] > .main {
    padding-left: 1rem;
    padding-right: 1rem;
}

/* Responsive font sizes */
@media only screen and (max-width: 600px) {
    h1 { font-size: 1.5rem !important; }
    h2 { font-size: 1.2rem !important; }
    h3 { font-size: 1rem !important; }
    .stButton>button { font-size: 0.9rem !important; padding: 0.4rem 0.8rem; }
    .stTextInput>div>div>input { font-size: 0.9rem !important; }
}

/* Reduce padding in columns on small screens */
@media only screen and (max-width: 600px) {
    [data-testid="column"] {
        padding-left: 0.5rem !important;
        padding-right: 0.5rem !important;
    }
}
body, html, .stApp {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 18px !important;
    height: 100vh;         /* Ocupa 100% da altura da tela */
    margin: 0 !important;
    padding: 0 !important;
    overflow-x: hidden;
    background-color: #0E1117;  /* Se quiser manter o fundo escuro sem gaps */
    display: flex;
    flex-direction: column;
}
.block-container {
    flex: 1 !important;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    padding-bottom: 0 !important;
}
.stButton > button {
    position: relative;
    display: block;
    width: 100%;
    max-width: 320px;
    min-width: 280px;
    box-sizing: border-box;
    padding: 2px 20px;
    white-space: nowrap;
    margin: 0px 0;
    border: none;
    flex-wrap: nowrap;
    border-radius: 8px;
    background: #0D1117;
    color: white;
    font-family: 'Space Grotesk', sans-serif;
    cursor: pointer;
    text-align: center;
    z-index: 0;
    transition: color 0.3s ease;
    overflow: hidden;
    box-sizing: border-box;
}

/* Borda glow animada azul e verde */
.stButton > button::before {
    content: '';
    position: absolute;
    top: -2px; left: -2px; right: -2px; bottom: -2px;
    border-radius: 8px;
    background: linear-gradient(270deg, 
        #00ff94, #007bff, #00fff7, #007bff, #00ff94);
    background-size: 1400% 1400%;
    filter: blur(4px);
    opacity: 0;
    transition: opacity 0.3s ease;
    z-index: -1;
    animation: rgbGlowBlueGreen 8s ease infinite;
}

.stButton > button:hover::before {
    opacity: 0.7;
}

/* Cor do texto muda ao passar o mouse */
.stButton > button:hover {
    color: black;  /* Azul claro, pode mudar para o tom que preferir */
    font-weight: bold;
}

@keyframes rgbGlowBlueGreen {
    0% {
        background-position: 0% 50%;
    }
    50% {
        background-position: 100% 50%;
    }
    100% {
        background-position: 0% 50%;
    }
}
.header, .footer, .marquee-container, .card {
    font-family: 'Space Grotesk', sans-serif;
}

button[kind="primary"] {
    all: unset;
}
button:focus {
    outline: none;
}
.content-box {
    background-color: #1c1e22;
    font-family: 'Space Grotesk', sans-serif;
    border-radius: 12px;
    padding: 30px;
    min-height: 600px;
    margin-top: 20px;
    color: white;
    z-index: 0;
    white-space: nowrap;
}
.container-outer {
    max-width: 800px;
    margin: 0 auto;
    padding: 20px 20px 20px 20px;
}
.menu-column {
    padding-right: 20px; 
}
</style>
""", unsafe_allow_html=True)

st.markdown(
"""
<style>
    /* Alvo: Rótulos (labels) de inputs numéricos e de texto */
    div[data-testid="stNumberInput"] label p,
    div[data-testid="stTextInput"] label p {
        font-size: 18px !important;
        font-weight: bold !important;
        color: #ffffff !important;
        font-family: 'Trebuchet MS', sans-serif !important;
    }
    
    /* Aumentar o tamanho do texto/número digitado dentro dos campos */
    div[data-testid="stNumberInput"] input,
    div[data-testid="stTextInput"] input {
        font-size: 18px !important;
        color: #39FF14 !important; /* Verde neon para o conteúdo */
        background-color: #1a1c24 !important;
    }
</style>
""",
unsafe_allow_html=True
)

# Adicionando CSS para os botões de navegação
st.markdown("""
    <style>
        .stButton > button {
            background-color: #342b44;  /* Cor do botão */
            color: white;  /* Cor do texto */
            padding: 12px 24px;
            font-size: 16px;
            border-radius: 8px;
            border: none;
            cursor: pointer;
            transition: background-color 0.3s ease;
        }
        .stButton > button:hover {
            background-color: #45a049;  /* Cor do botão ao passar o mouse */
            color: #342b44;  /* Cor do texto */
        }
    </style>
""", unsafe_allow_html=True)

# --- Estilização Customizada ---
st.markdown("""
    <style>
        .header {
            background-color: #342b44;
            border-radius: 8px;
            border: 2px solid #FFFFFF; 
            padding: 10px;
            text-align: center;
            color: white;
            font-size: 40px;
            font-weight: bold;
        }
        .footer {
            background-color: #342b44;
            border-radius: 8px;
            border: 2px solid #FFFFFF; 
            color: white;
            text-align: center;
            padding: 1rem;
            font-size: 14px;
        }
        .card {
            border: 2px solid black;
            padding: 1rem;
            text-align: center;
            border-radius: 10px;
            background-color: #f9f9f9;
            margin-bottom: 2rem;
            cursor: pointer;
        }
    </style>
""", unsafe_allow_html=True)

### SITE Header

st.markdown("""
<style>
@keyframes animatedGradient {
    0%   { background-position: 0% 50%; }
    50%  { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}

@keyframes float {
    0%   { transform: translateY(0); }
    50%  { transform: translateY(-10px); }
    100% { transform: translateY(0); }
}

.logo-container {
    display: flex;
    justify-content: center;
    align-items: center;
    gap: 12px;
    margin-top: 30px;
    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
}

.logo-icon {
    font-size: 80px;
    animation: float 2s infinite ease-in-out;
    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
    background-size: 600% 600%;
    animation: animatedGradient 6s ease infinite, float 2s infinite ease-in-out;
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
}

.logo-text {
    font-size: 56px;
    font-weight: 900;
    letter-spacing: 1.5px;
    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
    background-size: 600% 600%;
    animation: animatedGradient 6s ease infinite;
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    text-shadow: 0 0 8px #30f0c0;
}
</style>
<div class="logo-container">
    <div class="logo-icon">🪂</div>
    <div class="logo-text">Airdrops Monitor</div>
</div>
""", unsafe_allow_html=True)

### Letreiro Cotação tokens

tokens = {
    "bitcoin": {"name": "Bitcoin (BTC)", "icon": "https://cryptologos.cc/logos/bitcoin-btc-logo.png"},
    "ethereum": {"name": "Ethereum (ETH)", "icon": "https://cryptologos.cc/logos/ethereum-eth-logo.png"},
    "solana": {"name": "Solana (SOL)", "icon": "https://cryptologos.cc/logos/solana-sol-logo.png"},
    "binancecoin": {"name": "BNB (BNB)", "icon": "https://cryptologos.cc/logos/binance-coin-bnb-logo.png"},
    "hyperliquid": {"name": "Hyperliquid (HYPE)", "icon": "https://assets.coingecko.com/coins/images/34898/large/hype.png"},
    "aave": {"name": "Aave (AAVE)", "icon": "https://cryptologos.cc/logos/ethena-ena-logo.png"},
    "ripple": {"name": "XRP (XRP)", "icon": "https://cryptologos.cc/logos/xrp-xrp-logo.png"},
    "sui": {"name": "Sui (SUI)", "icon": "https://cryptologos.cc/logos/dogecoin-doge-logo.png"},
    "link": {"name": "Link (LINK)", "icon": "https://cryptologos.cc/logos/chainlink-link-logo.png"},
    "ethena": {"name": "Ethena (ENA)", "icon": "https://cryptologos.cc/logos/ethena-ena-logo.png"},
}

# Obter preços
url = f"https://api.coingecko.com/api/v3/simple/price?ids={','.join(tokens.keys())}&vs_currencies=usd"

data = requests.get(url).json()

# HTML e CSS do ticker
htmld = """
<style>
.ticker-container {
    width: 100%;
    overflow: hidden;
    background: radial-gradient(circle at top, #2A2A33 0%, #1B1B22 100%);
    padding: 10px 0;
    height: 70px;
    align-items: center;
}
.ticker {
    display: inline-block;
    white-space: nowrap;
    animation: scroll-left 30s linear infinite;
}
@keyframes scroll-left {
    0% { transform: translateX(100%); }
    100% { transform: translateX(-100%); }
}
.ticker-item {
    display: inline-block;
    margin: 6px 50px;
    color: white;
    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
    font-size: 22px;
    align-items: center;
    justify-content: center;
}
.ticker-item img {
    vertical-align: middle;
    height: 30px;
    margin-right: 8px;
}
</style>
<div class="ticker-container"><div class="ticker">
"""

# Preencher os dados

for key, token in tokens.items():
    price = data.get(key, {}).get("usd")
    if price is not None:
        htmld += f'<span class="ticker-item"> {token["name"]} ${price:,.3f} </span>'

htmld += "  "

# Renderização correta do HTML
st.markdown(htmld, unsafe_allow_html=True)

### Letreiro de divulgação de informações
st.markdown("""
<style>
@keyframes bg-gradient {
  0% {
    background-position: 0% 50%;
  }
  50% {
    background-position: 100% 50%;
  }
  100% {
    background-position: 0% 50%;
  }
}

.marquee-container {
  width: 100%;
  overflow: hidden;
  background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
  background-size: 600% 600%;
  animation: bg-gradient 10s ease infinite;
  color: black;
  font-size: 24px;
  padding: 10px;
  box-sizing: border-box;
  height: 70px;
  display: flex;
  align-items: center;
  justify-content: center;
  position: relative;
}

.marquee-text {
  display: inline-block;
  white-space: nowrap;
  position: absolute;
  will-change: transform;
  animation: marquee 30s ease-in-out infinite;
  font-weight: bold;
  text-shadow: 0 0 5px white;
}

@keyframes marquee {
  0%   { transform: translateX(160%); }
  25%  { transform: translateX(0%); }
  75%  { transform: translateX(0%); }
  100% { transform: translateX(-160%); }
}
</style>

<div class="marquee-container">
  <div class="marquee-text">
    🚨 Last News: 
    <a href='https://app.hyperlend.finance/dashboard/terms-and-conditions' target='_blank' style='color:black;'>Register to Hyperland Airdrop!</a> / 
    <a href='https://jup.ag/?ref=hf7390upylv8' target='_blank' style='color:black;'>Claim JUPITER Rewards!</a> /
    <a href='https://waitlist.stormrae.ai/dashboard?ref=CaioFlemin2089' target='_blank' style='color:black;'>Join the waitlist for STORMRAI!</a> /
    <a href='https://hub.orbt.xyz/?referral_code=ORBT-RCD-SSA-ETF' target='_blank' style='color:black;'>Join Orbit Protocol Campaign!</a> /
    <a href='https://waitlist.rally.fun/joinme/CaioFlemin2089' target='_blank' style='color:black;'>ByPass The InfoFi Crash!</a> /
  </div>
</div>
""", unsafe_allow_html=True)

# --- Sidebar ---
#st.sidebar.title("🪂 Airdrops Monitor")
#st.sidebar.markdown("---")
#st.sidebar.title("Menu")
#st.sidebar.markdown("<h3 style='font-size: 22px;'></h3>", unsafe_allow_html=True)

# Define options for the sidebar

# --- CSS para estilizar o sidebar ---
st.markdown("""
    <style>
    [data-testid="stSidebar"] {
        background-color: #2f2e41;
    }

    [data-testid="stSidebar"] .stRadio > div {
        gap: 8px;
    }

    [data-testid="stSidebar"] .stRadio div[role="radiogroup"] > label {
        font-size: 25px;
        padding: 10px 8px;
        border-radius: 8px;
        background-color: #b18d7e33;
        color: white;
        transition: all 0.3s ease;
        display: flex;
        align-items: center;
    }

    [data-testid="stSidebar"] .stRadio div[role="radiogroup"] > label:hover {
        background-color: #b18d7e55;
    }

    [data-testid="stSidebar"] .stRadio div[role="radiogroup"] > label[data-selected="true"] {
        background-color: #b18d7e;
        color: white;
        font-weight: bold;
    }

    .sidebar-title {
        font-size: 25px;
        font-weight: bold;
        margin-bottom: 10px;
        color: white;
    }

    .sidebar-section {
        font-size: 14px;
        color: #cccccc;
        margin: 20px 0 5px;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    </style>
""", unsafe_allow_html=True)

#NEON MOUSE PASS

st.markdown("""
    <style>
        .neon-hover-block {
            background: radial-gradient(circle at top center, #2c2f35 0%, #1c1e22 100%);
            font-size: 25px;
            border: 2px solid #444; 
            border-radius: 16px;
            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
            padding: 25px;
            margin: 20px 0;
            color: white;
            font-family: 'Space Grotesk', sans-serif;
            transition: all 0.3s ease;
        }

        .neon-hover-block:hover {
            border-color: #00f0ff;
            box-shadow: 0 0 20px #00f0ff;
            background: radial-gradient(circle at top center, #2f3640 0%, #1c1e22 100%);
        }
    </style>
""", unsafe_allow_html=True)

st.markdown("""
    <style>
        .neon-hover-block-z {
            background: radial-gradient(circle at top center, #2c2f35 0%, #1c1e22 100%);
            font-size: 25px;
            border: 2px solid #444; 
            border-radius: 16px;
            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
            padding: 25px;
            margin: 20px 0;
            color: white;
            font-family: 'Space Grotesk', sans-serif;
            transition: all 0.3s ease;
        }

        .neon-hover-block-z:hover {
            transform: scale(1.03);
            border-color: #00f0ff;
            box-shadow: 0 0 20px #00f0ff;
            background: radial-gradient(circle at top center, #2c2f35 0%, #1c1e22 100%);
        }
    </style>
""", unsafe_allow_html=True)

options = ["🏠 Welcome", "🧮 Airdrop Calculator", "🏆 Airdrop Points Viewer", "💵 Solana Stables APY", "🎒 BackPack Volume Check", "🌾 Farm with YT", "📊 Comparative YT Table", "📈 Pendle APY Prediction", 
           "🎁 Latest Airdrops", "♾️ PerpDEX Airdrops","📡 Depin Airdrops", "✅ Last Claims and Checkers", 
           "🌉 Bridges & Swaps Protocols", "⚖️ Funding Rate Arbitrage", "🚰 Faucets", "⛔ Revoke Contract", "⚠️ Avoiding Scams"]


st.markdown("\n\n")
st.sidebar.markdown("---")
opcao = st.sidebar.radio("🧮 Airdrop Calculator", options, index=1)
if "pagina" not in st.session_state:
    st.session_state.pagina = "🧮 Airdrop Calculator"

PAGES = {
    "🏠 Welcome": "",
    "✅ Last Claims and Checkers": "Latest claimable rewards and check tools.",
    "🏆 Airdrop Points Viewer": "Verify your wallet points in Airdrops",
    "🧮 Airdrop Calculator": "Estimate your potential airdrop rewards.",
    "💵 Solana Stables APY": "Solana Stables APY Chances",
    "⚖️ Funding Rate Arbitrage": "Funding Rate Arbitrage Chances",
    "♾️ PerpDEX Airdrops": "Airdrops from PerpDex.",
    "📡 Depin Airdrops": "Airdrops from DePIN (Decentralized Physical Infrastructure) projects.",
    "🎁 Latest Airdrops": "List of the latest available airdrops.",
    "🎒 BackPack Volume Check": "Check your BackPack Volume.",
    #"🌾 Farm with YT": "Yield farming with YouTube strategies.",
    #"📊 Comparative YT Table": "Compare YouTube farming strategies.",
    "📈 Pendle APY Prediction": "Pendle APY forecast and trends.",
    "🌉 Bridges & Swaps Protocols": "Protocols for bridging and swapping assets.",
    "🚰 Faucets": "Faucets to receive free testnet tokens.",
    "⛔ Revoke Contract": "Revoke connected contracts and permissions.",
    "⚠️ Avoiding Scams": "Tips and tools to avoid scams."
}
    
# -------------------------
# 🔹 Leitura inicial da página
# -------------------------
pagina_atual =  st.session_state["pagina"] #st.query_params.get("pagina", [list(PAGES.keys())[0]])#st.query_params.get("pagina", list(PAGES.keys())[0])
print(pagina_atual)
if "pagina" not in st.session_state:
    st.session_state.pagina = pagina_atual

# Se query_params mudar externamente → sincroniza
elif st.session_state.pagina != pagina_atual:
    st.session_state.pagina = pagina_atual

# Container externo
st.markdown('<div class="container-outer">', unsafe_allow_html=True)

# Layout colunas menu + conteúdo
col_zero,col_left,col_menu, col_content,col_rigth = st.columns([0.12,2,1, 7,2.1], gap="large")

# Menu lateral com botões
with col_left:
    st.markdown('<div class="menu-column">', unsafe_allow_html=True)
    match = re.match(r'^([^\w\s]+)', pagina_atual)
    emoji = match.group(1) if match else '' #list(PAGES.keys())[1].split()[0] 
    label = re.sub(r'^[^\w\s]+', '', pagina_atual).strip()#" ".join(list(PAGES.keys())[0].split()[1:])
    for pagina in PAGES:
        if st.button(pagina, key=pagina):
            emoji = pagina.split()[0]
            label = " ".join(pagina.split()[1:])
            st.session_state.pagina = pagina
            st.query_params.update({"pagina": pagina})
    st.markdown('</div>', unsafe_allow_html=True)

# Conteúdo principal
st.markdown("""
<style>
.emoji-gray {
    filter: grayscale(100%) hue-rotate(80deg) saturate(300%) brightness(1.2);
    font-size: 22px;
    vertical-align: middle;
    margin-right: 8px;
    user-select: none;
}
.page-name {
    font-size: 32px;  /* tamanho da fonte */
    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;  /* tipo da fonte */
    font-weight: 800;  /* peso da fonte (bold leve) */
    color: #E0E0E0;  /* cor do texto */
    margin-bottom: 10px;
    display: flex;
    align-items: center;
}
</style>
""", unsafe_allow_html=True)

# Remove emojis ou símbolos no início da string
clean_label = re.sub(r'^[^\w\s]+', '', pagina_atual).strip()

with col_content:
    #st.markdown(f"### {st.session_state.pagina}")
    st.markdown(f'<div class="page-name"><span class="emoji-gray">{emoji}</span>{label}</div>', unsafe_allow_html=True)

    # --- Dados dos Protocolos ---
    # Armazena a hora da primeira execução na sessão do usuário
    if "start_time" not in st.session_state:
        st.session_state.start_time = datetime.now()
        st.session_state.first_run = True  # flag de primeira execução
    else:
        st.session_state.first_run = False

    # Hora atual
    now = datetime.now()

    # Calcula a diferença em segundos
    elapsed_seconds = (now - st.session_state.start_time).total_seconds()

    # Carrega dados do BD
    try:
        protocolos = carregar_json()
    except:
        protocolos = {}
        print("No data in DB")

    # Aplica CSS customizado para mudar o tamanho da fonte da sidebar
    st.markdown("""
        <style>
        /* Altera a fonte de toda a sidebar */
        section[data-testid="stSidebar"] * {
            font-size: 18px !important;
        }
        </style>
    """, unsafe_allow_html=True)

    # Inicialização de parâmetros de cada protocolo
    i = 0
    # Ethena
    Ena_Data = []
    Ena_Multipleir = 25
    Ena_Boost = 1.10
    Ena_TP_0 = 17600000000000
    Ena_pts_token = 1
    Ena_date0 = datetime.strptime("2025-07-07T08:00:00.000Z", "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)

    # Fragmetric
    Frag_Data = []
    Frag_Multipleir = 4
    Frag_Boost = 1.10
    Backpack_Boost = 1.30
    Frag_TP_0 = 9710000000
    Frag_pts_token = 86.4
    Frag_date0 = datetime.strptime("2025-04-07T08:00:00.000Z", "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)

    #Backpack
    Backpack_date0 = datetime.strptime("2025-03-21T10:00:00.000Z", "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
    Backpack_TP_0 = 0

    #Variational
    Variational_date0 = datetime.strptime("2025-03-21T10:00:00.000Z", "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
    Variational_TP_0 = 0

    # --- Conteúdo Principal ---
   

    if st.session_state.pagina == "🏠 Welcome":
            
        st.markdown(
            """
            <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 20px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }

                /* Borda neon com gradiente animado */
                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }

                /* Efeito de hover */
                .airdrop-box:hover {
                    border-color: #00f0ff;
                    background: #262b33;
                }

                @keyframes neonBorder {
                    0%   { background-position: 0% 50%; }
                    50%  { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }

                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 0px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 5px;
                    margin-bottom: 0px;
                    color: #00ffae;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 0px;
                }
            </style>

            <div class="airdrop-box">
                <h2>Welcome to Airdrops Monitor</h2>
                <p style="color: #8293A3;">
                    This platform is designed for airdrop enthusiasts looking to stay updated with the latest crypto opportunities.
                    If you're passionate about crypto, this site helps you track valuable airdrops and grow your portfolio.
                </p>
                <h2>What Are Airdrops and How Can You Benefit from Them?</h2>
                <ul style="color: #8293A3;">
                    <p>There are two main types:</p>
                    <li><strong>Task-based</strong>: Complete simple actions like social follows to earn tokens.</li>
                    <li><strong>Holder</strong>: Receive tokens just by holding certain cryptocurrencies.</li>
                </ul>
                <p style="color: #8293A3;">
                    Participating in airdrops is an accessible and educational way to get involved in the crypto space, especially for beginners.
                    Best of all, you can grow your portfolio with real value without investing money upfront.
                    Browse our list of active airdrops, track new opportunities in real time, and start earning free crypto today!
                </p>
                <h2>About Me</h2>
                <p style="color: #8293A3;">
                    I'm an electrical-electronic engineer with a strong interest in investments and programming.
                    I created this platform using Python to share knowledge and support others in the crypto and airdrop community.
                </p>
            </div>
            """,
            unsafe_allow_html=True
        )

    elif st.session_state.pagina == "🧮 Airdrop Calculator":

        from datetime import datetime, timezone
        import streamlit as st
        import streamlit.components.v1 as components

        # =========================
        # CSS GLOBAL DAS COLUNAS
        # =========================
        st.markdown("""
        <style>
        .custom-columns [data-testid="stHorizontalBlock"] > div:nth-child(1) {
            background-color: #111827;
            border-radius: 12px;
            box-shadow: 0 0 12px #00ffae40;
            padding: 30px 20px 20px 20px;
            margin-bottom: 30px;
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
        }

        .custom-columns [data-testid="stHorizontalBlock"] > div:nth-child(2) {
            background-color: #1a1c28;
            border-radius: 12px;
            box-shadow: 0 0 12px #00ffaa40;
            padding: 30px 20px 20px 20px;
            margin-bottom: 30px;
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
        }
        </style>
        """, unsafe_allow_html=True)

        # =========================
        # VALORES PADRÃO
        # =========================
        FDV = 100
        total_points = 1_000_000_000
        airdrop_pct = 5
        your_points = 1_000_000
        referral = ""
        logo_url = ""

        protocolos = ["Backpack", "Variational", "Pacifica", "Ostium", "Paradex", "Extended", "Gvrt S2", "EdgeX", "Reya", "Any"]

        logos = {
            "Backpack": "https://pbs.twimg.com/profile_images/1957829985143791616/sA2YoWNq_400x400.jpg",
            "Variational": "https://pbs.twimg.com/profile_images/1983193863532548096/2FkeRmBg_400x400.jpg",
            "Pacifica": "https://pbs.twimg.com/profile_images/1911022804159389696/THxMFj50_400x400.jpg",
            "Paradex": "https://pbs.twimg.com/profile_images/2001911524429041669/P9KQbeNz_400x400.jpg",
            "Extended": "https://pbs.twimg.com/profile_images/1876581196173320192/pF4KQQCb_400x400.jpg",
            "Gvrt S2": "https://pbs.twimg.com/profile_images/1991661582527655936/Kn9vFGru_400x400.jpg",
            "Ostium": "https://pbs.twimg.com/profile_images/1948722481780453376/GT7D7CNh_400x400.jpg",
            "EdgeX": "https://pbs.twimg.com/profile_images/1976495879319322624/mMUMJ9ym_400x400.jpg",
            "Reya": "https://pbs.twimg.com/profile_images/1969675819510374400/V-ldI5xq_400x400.png",
            "Any": "https://cdn-icons-png.flaticon.com/512/61/61088.png"
        }

        if "selected_protocol" not in st.session_state:
            st.session_state.selected_protocol = "Backpack"

        def select_protocol(p):
            st.session_state.selected_protocol = p
            for o in protocolos:
                st.session_state[f"check_{o}"] = (o == p)

        # =========================
        # GRID DE PROTOCOLOS (FORA DAS COLUNAS)
        # =========================
        st.markdown("""
        <style>
        div[data-testid="stCheckbox"] * {
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif !important;
            font-size: 14px !important;
            color: #e5e7eb !important;
            letter-spacing: 0.2px;
            align-items: center !important;
        }
        </style>
        """, unsafe_allow_html=True)

        # --- abre container do grid ---
        st.markdown('<div class="protocol-grid">', unsafe_allow_html=True)

        cols = st.columns(len(protocolos))
        for i, p in enumerate(protocolos):
            if f"check_{p}" not in st.session_state:
                st.session_state[f"check_{p}"] = (p == st.session_state.selected_protocol)

            with cols[i]:
                st.checkbox(p, key=f"check_{p}", on_change=select_protocol, args=(p,))
                
                selected = "selected" if st.session_state.selected_protocol == p else ""
                st.markdown(f"""
                <div style="
                    background:#0f172a;
                    border:2px solid {'#00ffae' if selected else '#1f2933'};
                    box-shadow:{'0 0 18px #00ffae90' if selected else 'none'};
                    border-radius:18px;
                    padding:16px;
                    text-align:center;">
                    <img src="{logos[p]}" style="width:56px;border-radius:6px;"><br>
                    <b style="font-size:14px;color:white">{p}</b>
                </div>
                """, unsafe_allow_html=True)

        # --- fecha container do grid ---
        st.markdown('</div>', unsafe_allow_html=True)

        # =========================
        # COLUNAS PRINCIPAIS
        # =========================
        st.markdown('<div class="custom-columns">', unsafe_allow_html=True)
        colA1, colA2 = st.columns([1.8, 1.2])

        selected_protocol = st.session_state.selected_protocol

        # =========================
        # COLUNA ESQUERDA
        # =========================
        with colA1:


            if selected_protocol != "Any":

                auto_fill_actual = st.checkbox(
                    f"Auto-fill Total Actual {selected_protocol} Points",
                    value=True
                )

                if auto_fill_actual:
                    today = datetime.now(timezone.utc)

                    def calc(date_str, actual, daily, actual_date):
                        d = datetime.strptime(date_str+"T00:00:00.000Z", "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                        a = datetime.strptime(actual_date+"T00:00:00.000Z", "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                        return int(actual + (d - a).days * daily)

                    if selected_protocol == "Backpack":
                        date_tge = st.text_input("📅 TGE Date", "2026-01-28")
                        last_update = "2026-01-26"
                        FDV, airdrop_pct, your_points = 1000, 24, 1_000
                        total_points = calc(date_tge, 420_790_000, 10_000_000/7, last_update)
                        referral = "https://backpack.exchange/join/jj2kkdp1"
                        logo_url = logos["Backpack"]

                    elif selected_protocol == "Variational":
                        date_tge = st.text_input("📅 TGE Date", "2026-10-01")
                        last_update = "2026-01-02"
                        FDV, airdrop_pct, your_points = 1000, 20, 100
                        total_points = calc(date_tge, 3_450_000, 150_000/7, last_update)
                        referral = "https://omni.variational.io/"
                        logo_url = logos["Variational"]

                    elif selected_protocol == "Pacifica":
                        date_tge = st.text_input("📅 TGE Date", "2026-10-01")
                        last_update = "2026-01-20"
                        FDV, airdrop_pct, your_points = 500, 20, 1_000
                        total_points = calc(date_tge, 190_000_000, 10_000_000/7, last_update)
                        referral = "https://app.pacifica.fi?referral=PacificaRef"
                        logo_url = logos["Pacifica"]

                    elif selected_protocol == "Paradex":
                        date_tge = st.text_input("📅 TGE Date", "2026-02-01")
                        last_update = "2026-01-15"
                        FDV, airdrop_pct, your_points = 1000, 20, 1_000
                        total_points = calc(date_tge, 323_000_000, 4_200_000/7, last_update)
                        referral = "https://app.paradex.trade/r/Paradex15"
                        logo_url = logos["Paradex"]

                    elif selected_protocol == "Extended":
                        date_tge = st.text_input("📅 TGE Date", "2026-06-01")
                        last_update = "2026-01-19"
                        FDV, airdrop_pct, your_points = 500, 20, 100
                        total_points = calc(date_tge, 46_403_962, 1_034_000/7, last_update)
                        referral = "https://app.extended.exchange/join/EXT3NDED15"
                        logo_url = logos["Extended"]

                    elif selected_protocol == "Gvrt S2":
                        date_tge = st.text_input("📅 TGE Date", "2026-04-01")
                        last_update = "2026-01-20"
                        FDV, airdrop_pct, your_points = 500, 12, 100
                        total_points = calc(date_tge, 2_325_000, 135_000/7, last_update)
                        referral = "https://grvt.io/?ref=C496Y64"
                        logo_url = logos["Gvrt S2"]

                    elif selected_protocol == "Ostium":
                        date_tge = st.text_input("📅 TGE Date", "2026-12-21")
                        last_update = "2026-01-26"
                        FDV, airdrop_pct, your_points = 500, 20, 1_000
                        total_points = calc(date_tge, 32_500_000, 500_000/7, last_update)
                        referral = "https://app.ostium.com/trade?from=SPX&to=USD&ref=EIETH"
                        logo_url = logos["Ostium"]

                    elif selected_protocol == "EdgeX":
                        date_tge = st.text_input("📅 TGE Date", "2026-03-31")
                        last_update = "2026-01-07"
                        FDV, airdrop_pct, your_points = 1000, 25, 100
                        total_points = calc(date_tge, 7_300_000, 0, last_update)
                        referral = "https://pro.edgex.exchange/referral/EDGE15"
                        logo_url = logos["EdgeX"]

                    elif selected_protocol == "Reya":
                        date_tge = st.text_input("📅 TGE Date", "2026-03-31")
                        last_update = "2026-01-26"
                        FDV, airdrop_pct, your_points = 500, 20, 100
                        total_points = calc(date_tge, 10_500_000, 153_000/7, last_update)
                        referral = "https://app.reya.xyz/trade?referredBy=xe369dux"
                        logo_url = logos["Reya"]
            

            FDV = st.number_input("🧈Estimated FDV (M USD)", value=FDV) * 1_000_000
            airdrop_pct = st.number_input("％ Supply % for Airdrop", value=float(airdrop_pct))
            total_points = st.number_input("✦ Estimated Total Points for TGE", value=int(total_points))
            your_points = st.number_input("✧ Your Points", value=int(your_points))

        # =========================
        # COLUNA DIREITA (REFERRAL)
        # =========================
        with colA2:
            if selected_protocol != "Any":
                st.markdown(f"""
                <style>
                .ref-card {{
                    background-color: #111827;
                    border: 1px solid #00ffae80;
                    border-radius: 12px;
                    padding: 20px;
                    box-shadow: 0 0 12px #00ffae40;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    margin-top: 20px;
                }}
                .ref-box {{
                    background-color: #111;
                    padding: 10px;
                    margin: 10px 0;
                    border-radius: 8px;
                    font-weight: bold;
                    color: #00ffae;
                    word-break: break-all;
                }}
                .ref-card button {{
                    background-color: #00ffae;
                    color: #000;
                    padding: 10px 16px;
                    width: 100%;
                    border: none;
                    border-radius: 8px;
                    font-weight: bold;
                    cursor: pointer;
                    transition: transform 0.2s ease-in-out;
                }}
                .ref-card button:hover {{
                    transform: scale(1.03);
                    box-shadow: 0 0 10px #00ffae80;
                }}
                </style>

                <div class="ref-card">
                    <h3 style="color:#00ffae; font-size:25px; display:flex; align-items:center; gap:12px;">
                        Start with {selected_protocol}
                        <img src="{logo_url}" alt="{selected_protocol} logo" 
                            style="height:60px; width:auto; border-radius:8px; background:#0b1f1a; padding:4px;">
                    </h3>
                    <p style="margin-top:30px;font-size=22px;">New to {selected_protocol}? Use the referral link below to get started with <strong>Points Boost</strong> in the platform.</p>
                    <div class="ref-box" style="margin-top:20px;">{referral}</div>
                    <a " href={referral} target="_blank">
                        <button >Access {selected_protocol} ↗</button>
                    </a>
                    <ul class="ref-box" style="margin-top:30px;">
                        <p>✅ Earn points for future token airdrop</p>
                        <p>✅ Access to the {selected_protocol} ecosystem</p>
                    </ul>
                </div>
            """, unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)
        # Armazena no session_state
        st.session_state.FDV = FDV
        st.session_state.total_points = total_points
        st.session_state.airdrop_pct = airdrop_pct
        st.session_state.your_points = your_points

        st.markdown('</div>', unsafe_allow_html=True)

        # --- Cálculo
        airdrop_total_value = FDV * (airdrop_pct / 100)
        value_per_point = airdrop_total_value / total_points if total_points > 0 else 0
        your_estimated_airdrop = your_points * value_per_point

        # --- HTML completo com CSS embutido ---
        full_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    color: #e6edf3;
                }}
                
                .result-card {{
                    background-color: #0f172a;
                    position: absolute;
                    border-radius: 10px;
                    padding: 30px;
                    margin-top: 20px;
                    margin-bottom: 10px;
                    maring-rigth: 10px;
                    maring-left: 10px;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    width: 94.5%;
                }}
                
                .result-card::before {{
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 10px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 6px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }}
                .metrics-grid {{
                    font-size: 20px;
                    display: flex;
                    flex-wrap: wrap;
                    gap: 20px;
                    margin-top: 20px;
                }}

                .metric-box {{
                    background-color: #1e293b;
                    padding: 20px;
                    border-radius: 12px;
                    flex: 1 1 200px;
                    min-width: 200px;
                }}

                .metric-label {{
                    font-size: 20px;
                    color: #cccccc;
                    margin-bottom: 6px;
                }}

                .metric-value {{
                    font-size: 1.6em;
                    color: #00ffae;
                    font-weight: bold;
                }}

                .note {{
                    margin-top: 30px;
                    font-size: 18px;
                    color: #d1d5db;
                }}

                @keyframes neonBorder {{
                    0%   {{ background-position: 0% 50%; }}
                    50%  {{ background-position: 100% 50%; }}
                    100% {{ background-position: 0% 50%; }}
                }}

            </style>
        </head>
        <body>

            <div class="result-card">
                <h3 style="color:#00ffae;font-size:25px;">Simulation Results</h3>

                <div class="metrics-grid">
                    <div class="metric-box">
                        <div class="metric-label">Total Airdrop Value</div>
                        <div class="metric-value">${airdrop_total_value/1_000_000:.2f}M</div>
                    </div>
                    <div class="metric-box">
                        <div class="metric-label">Value per Point</div>
                        <div class="metric-value">${value_per_point:.12f}</div>
                    </div>
                    <div class="metric-box">
                        <div class="metric-label">Your Estimated Airdrop</div>
                        <div class="metric-value">${your_estimated_airdrop/1_000:.3f}K</div>
                    </div>
                </div>

                <p class="note">
                    <strong>Note:</strong> This is just a simulation based on your assumptions. The actual airdrop value may vary significantly depending on various factors such as launch price, final tokenomics, and total number of participants.
                </p>
            </div>

        </body>
        </html>
        """

        # Renderiza o HTML customizado
        components.html(full_html, height=400, width=1900, scrolling=False)

    
    elif st.session_state.pagina == "🏆 Airdrop Points Viewer":

        st.markdown(
        """
        <style>
            .airdrop-box {
                position: relative;
                z-index: 1;
                border-radius: 12px;
                padding: 25px;
                margin: 20px 0;
                margin-bottom: 40px;
                background: #111827;
                display: flex;
                flex-direction: column;
                gap: 30px;
                font-size: 22px;
                color: white;
                font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                overflow-wrap: break-word;
                word-wrap: break-word;
                white-space: normal;
                box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
            }

            /* Borda neon com gradiente animado */
            .airdrop-box::before {
                content: "";
                position: absolute;
                top: -3px;
                left: -3px;
                right: -3px;
                bottom: -3px;
                border-radius: 14px;
                z-index: -1;
                background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                background-size: 600% 600%;
                animation: neonBorder 6s ease infinite;
                padding: 3px;
                -webkit-mask:
                    linear-gradient(#fff 0 0) content-box,
                    linear-gradient(#fff 0 0);
                -webkit-mask-composite: xor;
                mask-composite: exclude;
            }

            /* Efeito de hover */
            .airdrop-box:hover {
                border-color: #00f0ff;
                background: #262b33;
            }

            @keyframes neonBorder {
                0%   { background-position: 0% 50%; }
                50%  { background-position: 100% 50%; }
                100% { background-position: 0% 50%; }
            }

            .airdrop-box h1 {
                font-size: 25px;
                text-align: center;
                margin-bottom: 0px;
            }

            .airdrop-box h2 {
                font-size: 25px;
                margin-top: 5px;
                margin-bottom: 0px;
            }

            .airdrop-box ul {
                margin-left: 20px;
                margin-bottom: 0px;
            }
        </style>
        <div class="airdrop-box">
            <h2 style="color: #00ffae; margin:0;">Airdrop Points Viewer</h2>
            <p style="color: #8293A3;">Enter your wallet address below to check your points across participating protocols in the Airdrop Points Campaign. You can verify one or multiple addresses at once.</p>
        </div>
        """,
        unsafe_allow_html=True
    )

        # -------------------------
        # 🔹 HELPER
        # -------------------------
        def safe_num(value, decimals=2):
            try:
                return f"{float(value or 0):,.{decimals}f}"
            except:
                return "0.00"
        
        def safe_request_simple(url, timeout=10):
            """Request seguro usando apenas requests"""
            try:
                res = requests.get(url, timeout=timeout)
                print(res)
                res.raise_for_status()
                if not res.text.strip():
                    return None
                return res.json()
            except requests.RequestException as e:
                print(f"Erro de requisição: {e}")
                return None
            except ValueError as e:
                print(f"Erro ao parsear JSON: {e}")
                return None
            
        # cria o scraper uma única vez
        _scraper = cloudscraper.create_scraper(
            browser={
                "browser": "chrome",
                "platform": "windows",
                "desktop": True
            }
        )
    
        def safe_request(url, params=None, payload=None, use_scraper=False, method="GET"):
            try:
                session = _scraper if use_scraper else requests

                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    "Accept": "application/json, text/plain, */*",
                    "Accept-Language": "en-US,en;q=0.9",
                    "Referer": "https://beta.triadfi.co/",
                    "Origin": "https://beta.triadfi.co"
                }

                if method == "GET":
                    res = session.get(url, params=params, headers=headers, timeout=20)
                else:
                    res = session.post(url, json=payload, headers=headers, timeout=20)

                if res.status_code != 200:
                    print("HTTP", res.status_code, url)
                    print(res.text[:300])
                    return None

                return res.json()

            except Exception as e:
                print(f"Request error → {url} | {e}")
                return None


        # -------------------------
        # 🔹 FETCH FUNCTIONS
        # -------------------------
        def get_ethena(wallet: str):
            url = f"https://app.ethena.fi/api/referral/get-referree?address={wallet}"
            data = safe_request_simple(url)
            if data and data.get("queryWallet"):
                d = data["queryWallet"][0]
                try:
                    return {
                        "points": d.get("accumulatedTotalShardsEarned", 0),
                        "referrals": d.get("accumulatedReferralsShards", 0),
                        "extra": (
                            f"<br>💎 Pendle: {safe_num(d.get('accumulatedPendleShards'), 2)} / "
                            f"🌱 Restaking: {safe_num(d.get('accumulatedRestakingShards'), 0)}</br> "
                            f"<br>🚀 Boost: {safe_num(d.get('accumulatedBoostShards'), 0)} / "
                            f"🌐 L2: {safe_num(d.get('accumulatedL2Shards'), 0)}</br>"
                        ),
                        "raw": data
                    }
                except:
                    return None
            return None

        def get_ethereal(wallet: str):
            url = f"https://deposit-api.ethereal.trade/v1/account/{wallet}"
            data = safe_request(url)
            if data:
                try:
                    pts = int(data["accounts"][0].get("points", 0)) / 1e18 if data.get("accounts") else 0
                    s0_pts = int(data["accountsS0"][0].get("points",0)) / 1e18 if data.get("accountsS0") else 0
                    return {
                        "points": pts,
                        "extra": f"<br>🔹 S0 Points: {s0_pts:,.2f}</br> <br> </br>",
                        "rank": data.get("rank", "N/A"),
                        "raw": data
                    }
                except:
                    return None
            return None

        def get_hylo(wallet: str):
            url = "https://hylo.so/api/user-points"
            data = safe_request(url, params={"address": wallet}, use_scraper=True)
            if data:
                return {
                    "points": data.get("totalPoints", 0),
                    "rank": data.get("globalRank", "N/A"),
                    "extra": f"<br>👥 Referral Points: {data.get('referralPoints',0):,.2f}</br> <br> </br>",
                    "raw": data
                }
            return None

        def get_noon(wallet: str):
            url = f"https://back.noon.capital/api/v1/points/{wallet}"
            data = safe_request(url, use_scraper=True)
            if data:
                try:
                    total = int(data["totalPoints"]) / 1e17
                    last24h = int(data["last24HoursPoints"]) / 1e17
                    return {
                        "points": total,
                        "extra": f"<br>⏱ Last24h: {last24h:,.2f}</br> <br> </br>",
                        "raw": data
                    }
                except:
                    return None
            return None

        def get_plume(wallet: str):
            url = f"https://portal-api.plume.org/api/v1/stats/pp-totals?walletAddress={wallet}"
            data = safe_request(url, use_scraper=True)
            if data and "data" in data and data["data"].get("ppScores"):
                try:
                    pp = data["data"]["ppScores"]
                    return {
                        "points": pp["activeXp"].get("totalXp",0),
                        "extra": f"<br>👥 Referral Points: {pp['activeXp'].get('referralBonusXp',0):,.2f}</br>",
                        "raw": data
                    }
                except:
                    return None
            return None

        def get_yieldfi_points(wallet: str):
            url = f"https://ctrl.yield.fi/u/points/{wallet}"
            data = safe_request(url, use_scraper=True)
            if data:
                try:
                    return {
                        "points": data.get("totalPoints",0),
                        "extra": f"<br></br> <br></br>",
                        "raw": data
                    }
                except:
                    return None
            return None

        def get_triad(wallet: str):
            url = f"https://beta.triadfi.co/api/user/{wallet}/verify"
            data = safe_request(url, use_scraper=True)
            print(data)
            if data:
                try:
                    return {
                        "points": data.get("xp",0),
                        "rank": data.get("rank","N/A"),
                        "extra": f"<br>TRD Staked: {data.get('trdStaked',0):,.2f}</br> <br> </br>",
                        "raw": data
                    }
                except:
                    return None
            return None

        def get_kinetiq(wallet: str):
            url = f"https://kinetiq.xyz/api/kpoints/{wallet}?chainId=999"
            data = safe_request(url, use_scraper=True)
            if data:
                try:
                    return {
                        "points": data.get("points",0),
                        "extra": f"<br>⭐ Tier: {data.get('tier','N/A')}</br> <br> </br>",
                        "raw": data
                    }
                except:
                    return None
            return None
        
        def get_hyperbeat_all(wallet: str):
            wallet = wallet.lower()
            url = f"https://api.hyperbeat.org/api/v1/ecosystem-points?userAddress={wallet}&provider=felix_hbusdt%2Cfelix_lsthype%2Chyperlend_hbbtc%2Chyperlend_hbhype%2Chyperlend_hbusdt%2Chyperlend_lsthype%2Chyperswap_hbbtc%2Chyperswap_hbhype%2Chyperswap_hbusdt%2Chypurrfi_hbbtc%2Chypurrfi_hbhype%2Chypurrfi_hbusdt%2Ckinetiq_lsthype%2Ctheo_hbusdt"

            data = safe_request(url, use_scraper=True)

            print("HYPERBEAT RAW:", data)

            if not data or not data.get("success"):
                return None

            provider = {}
            total_points = 0

            for item in data.get("data", []):
                p = item.get("provider")
                pts = float(item.get("points", 0))

                provider[p] = pts
                total_points += pts

            return {
                "points": total_points,          # total somado
                "rank": data.get("rank_position", "N/A"),
                "provider": provider,            # <<< AQUI está o que sua outra função usa
                "raw": data
            }
        
        def get_hyper_provider(wallet: str, provider_name: str):
            data = get_hyperbeat_all(wallet)

            if not data:
                return None

            points = float(data["provider"].get(provider_name, 0))

            return {
                "points": points,
                "extra": f"<br>{provider_name}: {points:.4f}</br><br></br>",
                "raw": data["provider"]
            }

        def get_theo(wallet: str):
            return get_hyper_provider(wallet, "theo_hbusdt")
        
        def get_felix(wallet: str):
            return get_hyper_provider(wallet, "felix_hbusdt")

        def get_hyperlend(wallet: str):
            return get_hyper_provider(wallet, "hyperlend_hbusdt")

        def get_hypurrfi(wallet: str):
            return get_hyper_provider(wallet, "hypurrfi_hbusdt")

        def get_hyperbeat(wallet: str):
            wallet = wallet.lower()
            url = f"https://api.hyperbeat.org/api/v1/leaderboard-v2/{wallet}"

            data = safe_request(url, use_scraper=True)

            print( data)

            if not data:
                return None

            return {
                "points": data.get("points", 0),
                "rank": data.get("rank_position", "N/A"),
                "extra": (
                    f"<br>🎯 Percentile: {data.get('percentile', 'N/A')}%</br>"
                    f"<br>🏅 League: {data.get('league', 'N/A')}</br>"
                    f"<br>👥 Referral Points: {data.get('referral_points', 0):,.2f}</br>"
                    "<br> </br>"
                ),
                "raw": data
            }

        def get_prjx(wallet: str):
            url = f"https://api.prjx.com/scorecard/impersonate/{wallet}?format=json"
            data = safe_request(url, use_scraper=True)
            if data and data.get("stats"):
                try:
                    return {
                        "points": data["stats"].get("totalPoints",0),
                        "rank": data["stats"].get("rank","N/A"),
                        "extra": f"<br>👥 Referrals Points: {data['stats'].get('totalReferrals',0):,.2f}</br> <br> </br>",
                        "raw": data
                    }
                except:
                    return None
            return None

        def get_ratex(wallet: str):
            url = "https://api.rate-x.io/"
            payload = {"serverName":"AdminSvr","method":"queryRatexPoints","content":{"user_id":wallet}}
            data = safe_request(url, method="POST", payload=payload)
            if data and data.get("code")==0 and data.get("data"):
                item = data["data"][0]
                return {
                    "points": float(item.get("total_points",0)),
                    "extra": (
                        f"<br>LP: {float(item.get('lp_points',0) or 0):,.2f}, "
                        f"Trade: {float(item.get('trade_points',0) or 0):,.2f}</br> <br> </br>"
                    ),
                    "rank": item.get("ranking","N/A"),
                    "raw": item
                }
            return None

        def get_lombard(wallet: str):
            url = f"https://mainnet.prod.lombard.finance/api/v1/referral-system/season-1/points/{wallet}"
            data = safe_request(url, use_scraper=True)
            if data:
                return {
                    "points": data.get("total",0),
                    "extra": f"<br>🎖️ Badges Points: {safe_num(data.get('badges_points'))}</br> <br> </br>",
                    "raw": data
                }
            return None


        # -------------------------
        # 🔹 PROTOCOLS CONFIG
        # -------------------------
        protocols = [
            {
                "name": "Ethena",
                "site": "https://app.ethena.fi/join/yp9pg",
                "image": "https://pbs.twimg.com/profile_images/1684292904599126016/f0ChONgU_400x400.png",
                "fetch": get_ethena,
            },
            {
                "name": "Ethereal",
                "site": "https://deposit.ethereal.trade/points?ref=KVHDUP",
                "image": "https://pbs.twimg.com/profile_images/1980059369736654849/dF9lY6aT_400x400.jpg",
                "fetch": get_ethereal,
            },
            {
                "name": "Noon",
                "site": "https://app.noon.capital/get?referralCode=f351689c-2391-4d2b-a963-d24a5530753a",
                "image": "https://pbs.twimg.com/profile_images/1927790098906497025/Ze-SQcgt_400x400.jpg",
                "fetch": get_noon,
            },
            {
                "name": "Plume",
                "site": "https://portal.plume.org/?referrer=MagnoliaBustlingAnt788",
                "image": "https://pbs.twimg.com/profile_images/1949681985149956096/ttJazkDy_400x400.jpg",
                "fetch": get_plume,
            },
            {
                "name": "YieldFi",
                "site": "https://yield.fi/yusd?referral=E4TAZUVU",
                "image": "https://pbs.twimg.com/profile_images/1922944632180310016/tN3q4S8G_400x400.jpg",
                "fetch": get_yieldfi_points,
            },
            {
                "name": "Hylo",
                "site": "https://hylo.so/leverage?ref=E27KDV",
                "image": "https://pbs.twimg.com/profile_images/1999811414299721729/-XyLtThr_400x400.png",
                "fetch": get_hylo,
            },
            {
                "name": "Triad",
                "site": "https://triadfi.co/?ref=Caio Fleming",
                "image": "https://pbs.twimg.com/profile_images/1955263071758225408/TqPBIm8I_400x400.jpg",
                "fetch": get_triad,
            },
            {
                "name": "RateX",
                "site": "https://app.rate-x.io/referral?ref=H9GnKZON",
                "image": "https://pbs.twimg.com/profile_images/1790703675700355072/wUBLpPIS_400x400.jpg",
                "fetch": get_ratex,
            },
            {
                "name": "Kinetiq",
                "site": "https://kinetiq.xyz",
                "image": "https://pbs.twimg.com/profile_images/1880410606093647872/qazlkvcq_400x400.jpg",
                "fetch": get_kinetiq,
            },
            {
                "name": "Hyperbeat",
                "site": "https://app.hyperbeat.org/earn?referral=F56B5BD8",
                "image": "https://pbs.twimg.com/profile_images/1879158343194796032/ftN7FT3s_400x400.jpg",
                "fetch": get_hyperbeat,
            },
            {
                "name": "Theo",
                "site": "https://app.theo.xyz/invite?invite=TheoBoost",
                "image": "https://pbs.twimg.com/profile_images/1928132409783963648/LizAN275_400x400.png",
                "fetch": get_theo,
            },
            {
                "name": "HyperLend",
                "site": "https://app.hyperlend.finance/?ref=CHYPER",
                "image": "https://pbs.twimg.com/profile_images/2011062303932678145/a5-fTMGC_400x400.jpg",
                "fetch": get_hyperlend,
            },
            {
                "name": "Felix",
                "site": "https://www.usefelix.xyz?ref=D51B721C",
                "image": "https://pbs.twimg.com/profile_images/1845076293735297024/mx8MTMca_400x400.jpg",
                "fetch": get_felix,
            },
            {
                "name": "HypurrFi",
                "site": "https://app.hypurr.fi/points?af=E2C6D4B",
                "image": "https://pbs.twimg.com/profile_images/1991902971676758017/Dl4dF7Ay_400x400.jpg",
                "fetch": get_hypurrfi,
            },
            {
                "name": "Project X",
                "site": "https://www.prjx.com/@Fleming",
                "image": "https://pbs.twimg.com/profile_images/1922089219737911296/1miGhDTB_400x400.jpg",
                "fetch": get_prjx,
            }
        ]


        # -------------------------
        # 🔹 INPUT
        # -------------------------
        wallets_input = st.text_input(
            "Enter one or more wallet addresses (0x...), separated by commas",
            value=""
        )
        # -------------------------
        # 🔹 CSS
        # -------------------------
        st.markdown("""
        <style>
        .container-externa {
            display: flex;
            flex-wrap: wrap;
            justify-content: center;
            gap: 15px;
        }
        .container-block {
            background: #1E1F25;
            border-radius: 15px;
            padding: 20px;
            max-width: 380px;   /* <-- limita para não estourar */
            color: white;
            text-align: center;
            border: 1px solid rgba(0,255,255,0.2);
            box-shadow: 0 0 15px rgba(0,255,255,0.4);
            transition: transform 0.2s;
        }
        .container-block:hover {
            transform: scale(1.05);
        }
        .container-block img {
            border-radius: 50%;
            margin-bottom: 10px;
        }
        </style>
        """, unsafe_allow_html=True)

        def generate_airdrop_excel(wallets_data):

            # Criar workbook em branco
            wb = Workbook()
            ws = wb.active
            ws.title = "Airdrops"

             # Ajuste de largura das colunas
            column_widths = {"A": 10, "B": 12, "C": 10, "D": 90, "E": 60}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            current_row = 1
            for wallet, protocols in wallets_data.items():
                if not protocols:
                    continue

                # Linha com o título da wallet
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                cell = ws.cell(row=current_row, column=1)
                cell.value = f"Wallet: {wallet}"
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1

                # Cabeçalho
                headers = ["name", "points", "rank", "extra", "site"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=current_row, column=col, value=h).font = Font(bold=True)
                current_row += 1

                # Dados da wallet
                df_wallet = pd.DataFrame(protocols)
                for r in df_wallet.itertuples(index=False):
                    for c, v in enumerate(r, 1):
                        ws.cell(row=current_row, column=c, value=v)
                    current_row += 1

                # Linha em branco entre wallets
                current_row += 1

            # Salvar no buffer
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        
        # ===========================
        # 🧹 Limpar tags HTML (<p>, <br>, etc.)
        # ===========================
        def clean_html(text):
            if pd.isna(text):
                return ""
            # remove <p>, <br>, etc.
            text = re.sub(r"<.*?>", " ", str(text))
            return text.strip()
        
        def safe_fetch(fetch_func, wallet):
            try:
                data = fetch_func(wallet)
                return data if data else {}
            except Exception as e:
                st.warning(f"Erro ao buscar dados para {wallet}: {e}")
                return {}

        # -------------------------
        # 🔹 RENDER COM LAYOUT NEON
        # -------------------------
        wallets_data = {}
        if wallets_input:
            wallets = [w.strip() for w in wallets_input.split(",") if w.strip()]
            tabs = st.tabs(wallets)  # cria uma aba para cada wallet

            for i, wallet in enumerate(wallets):
                with tabs[i]:
                    protocol_data = []  # lista limpa para cada wallet
                    blocks_html = ""
                    for proto in protocols:
                        data = safe_fetch(proto["fetch"], wallet)
                        if not data or data.get("points") in [None, 0, "N/A"]:
                            continue

                        points = f"{data.get('points', 0):,.2f}"
                        rank = data.get("rank", "N/A")
                        raw_extra = data.get("extra", "")
                        clean_extra = re.sub(r"<.*?>", " ", str(raw_extra)).strip()

                        protocol_data.append({
                            'name': proto['name'],
                            'points': data.get('points', 0),
                            'rank': rank,
                            'extra': clean_extra,
                            'site': proto['site']
                        })

                        # Monta HTML
                        blocks_html += f"""
                        <div class="container-block">
                            <div class="header-wrapper">
                                <img src="{proto['image']}" width="50" height="50" style="border-radius: 50%;">
                                <div><strong>{proto['name']}</strong></div>
                            </div>
                            <div class="footer-wrapper">
                                <p>🌟 Points: {points}</p>
                                <p>🏆 Rank: {rank}</p>
                                <p>🌐 Site: <a href="{proto['site']}" target="_blank">Visit Protocol</a></p>
                            </div>
                        </div>
                        """

                    # salva somente depois que todos os protocolos da wallet foram processados
                    wallets_data[wallet] = protocol_data

                    # HTML completo com o mesmo CSS neon
                    full_html = f"""
                    <style>
                    @keyframes neonBorder {{
                        0%   {{ background-position: 0% 50%; }}
                        50%  {{ background-position: 100% 50%; }}
                        100% {{ background-position: 0% 50%; }}
                    }}
                    .container-externa {{
                        border-radius: 12px;
                        padding: 25px;
                        margin-top: 30px;
                        gap: 0px;
                        display: flex;
                        flex-wrap: wrap;
                        justify-content: left;
                        font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                        font-size: 22px;
                        color: white;
                        margin: 0px 0;
                        align-items: center;
                        justify-content: center;
                        overflow: hidden;
                        scrollbar-width: none;
                    }}
                    .container-externa::-webkit-scrollbar {{
                        display: none; /* Chrome/Safari */
                    }}
                    .container-block {{
                        width: 360px;
                        flex-shrink: 0; 
                        margin: 10px;
                    }}
                    .header-wrapper {{
                        width: 290px;
                        padding: 30px;
                        margin-top: 10px;
                        margin-right: 15px;
                        margin-left: 12px;
                        border-top: 1px solid rgba(48, 240, 192, 0.2);
                        border-bottom: 1px solid #00e0ff;
                        border-top-left-radius: 40px;
                        border-top-right-radius: 10px;  /* 👈 maior */
                        border-bottom-left-radius: 5px;
                        border-bottom-right-radius: 5px;
                        display: flex;
                        align-items: center;
                        gap: 30px;
                        background: #1E1F25;
                        box-shadow: 0 0 20px rgba(0,255,150,0.3);
                        transition: transform 0.3s ease;
                        font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                        z-index: 1;
                        overflow: hidden;
                        align-items: center;
                        justify-content: center;
                    }}
                    .header-wrapper:hover {{
                        border: 1px solid #39ff14; /* Verde neon */
                        border-top-left-radius: 40px;
                        border-top-right-radius: 10px; 
                        border-bottom-left-radius: 5px;
                        border-bottom-right-radius: 5px;
                        box-shadow: 0 0 4px #39ff14, 0 0 8px #39ff14; /* Brilho neon suave */
                        background: #262b33;
                    }}
                    .header-wrapper:hover::before {{
                        content: "";
                        position: absolute;
                        top: -3px;
                        left: -3px;
                        right: -3px;
                        bottom: -3px;
                        border-radius: 14px;
                        z-index: -1;
                        animation: neonBorder 6s ease infinite;
                        -webkit-mask:
                            linear-gradient(#fff 0 0) content-box,
                            linear-gradient(#fff 0 0);
                        -webkit-mask-composite: xor;
                        mask-composite: exclude;
                    }}
                    .header-wrapper a {{
                        color: lightblue;
                        text-decoration: none;
                    }}
                    .footer-wrapper {{
                        width: 290px;
                        padding: 30px;
                        margin-top: 6px;
                        margin-bottom: 30px;
                        margin-right: 15px;
                        margin-left: 12px;
                        border-top: 1px solid #00e0ff;
                        border-top-left-radius: 5px;
                        border-top-right-radius: 5px;  /* 👈 maior */
                        border-bottom-left-radius: 10px;
                        border-bottom-right-radius: 40px;
                        display: block;
                        align-items: center;
                        gap: 30px;
                        background: #1E1F25;
                        box-shadow: 0 0 20px rgba(0,255,150,0.5);
                        transition: transform 0.3s ease;
                        font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                        z-index: 1;
                        overflow: hidden;
                        align-items: center;
                        justify-content: center;
                        font-size:20px;
                    }}
                    .footer-wrapper:hover {{
                        border: 1px solid #39ff14; /* Verde neon */
                        border-top-left-radius: 5px;
                        border-top-right-radius: 5px; 
                        border-bottom-left-radius: 10px;
                        border-bottom-right-radius: 40px;
                        box-shadow: 0 0 4px #39ff14, 0 0 8px #39ff14; /* Brilho neon suave */
                        background: #262b33;
                    }}
                    .footer-link {{
                        text-decoration: none;
                        margin: 0;
                        width: 100%;
                        color: inherit;
                    }}
                    .footer-link:hover {{
                        color: inherit;
                    }}
                    </style>

                    <div class="container-externa">
                        {blocks_html}
                    </div>
                    """
            
                    components.html(full_html, height=1800, width=1900, scrolling=False)

        if wallets_data:
            excel_bytes = generate_airdrop_excel(wallets_data)
            st.download_button(
                label="📥 Baixar Airdrop Excel",
                data=excel_bytes,
                file_name="airdrop_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    elif st.session_state.pagina == "🎒 BackPack Volume Check":
        
        # -----------------------------
        # 📂 Upload do CSV
        # -----------------------------
        st.markdown(
            '<p style="font-size:22px;">Load your trade history from 🎒<a href="https://backpack.exchange/join/jj2kkdp1" target="_blank">Backpack</a></p>',
            unsafe_allow_html=True
        )
        uploaded_file = st.file_uploader(label="", type=["csv"])

        

        if uploaded_file is not None:
            #try:
                df = pd.read_csv(uploaded_file)
                df["timestamp"] = pd.to_datetime(df["timestamp"], format="ISO8601",utc=True)
                df["timestamp_naive"] = df["timestamp"].dt.tz_convert(None)
                df["volume"] = df["price"] * df["quantity"]
                df["fee"] = pd.to_numeric(df["fee"], errors="coerce").fillna(0)

                if "symbol" in df.columns:
                    tokens = df["symbol"].unique().tolist()
                    token_selected = st.selectbox("Select the token:", ["All Tokens"] + tokens)
                    if token_selected != "All Tokens":
                        df = df[df["symbol"] == token_selected]
                else:
                    st.warning("⚠️ No 'symbol' column found. Token filtering will not be applied.")
                    token_selected = "All Tokens"

                df["week"] = (df["timestamp_naive"] - pd.Timedelta(hours=21)).dt.to_period("W-THU").dt.start_time + pd.Timedelta(hours=21)

                weekly_volume = df.groupby("week")["volume"].sum().reset_index()
                weekly_trades = df.groupby("week")["volume"].count().reset_index().rename(columns={"volume": "num_trades"})
                weekly_summary = pd.merge(weekly_volume, weekly_trades, on="week")
                weekly_fees = df.groupby("week")["fee"].sum().reset_index().rename(columns={"fee": "fees_paid"})

                total_volume = weekly_summary["volume"].sum()

                weekly_summary = weekly_volume.merge(weekly_trades, on="week").merge(weekly_fees, on="week")
                total_fees = weekly_summary["fees_paid"].sum() 

                st.subheader("📊 Weekly Summary")
                col1, col2 = st.columns([1.6, 1.4])
                with col1:
                    st.dataframe(weekly_summary.rename(columns={"week": "Week", "volume": "Volume ($)", "num_trades": "Number of Trades", "fees_paid": "Fees Paid ($)"}))

                with col2:
                    fig = go.Figure()

                    fig.add_trace(go.Bar(
                        x=weekly_summary["week"],
                        y=weekly_summary["volume"],
                        name="Volume ($)",
                        marker_color="skyblue",
                        offsetgroup=0,
                        yaxis="y"
                    ))

                    fig.add_trace(go.Bar(
                        x=weekly_summary["week"],
                        y=weekly_summary["num_trades"],
                        name="Number of Trades",
                        marker_color="orange",
                        offsetgroup=1,
                        yaxis="y2"
                    ))

                    fig.update_layout(
                        title=dict(
                            text=f"📊 Weekly Volume and Number of Trades - {token_selected}",
                            x=0.5,
                            xanchor="center"
                        ),
                        xaxis=dict(title="Week"),
                        yaxis=dict(
                            title=dict(text="Volume ($)", font=dict(color="skyblue")),
                            tickfont=dict(color="skyblue")
                        ),
                        yaxis2=dict(
                            title=dict(text="Number of Trades", font=dict(color="orange")),
                            tickfont=dict(color="orange"),
                            overlaying="y",
                            side="right"
                        ),
                        barmode="group",
                        legend=dict(x=0.5, y = 1.12, xanchor="center", orientation="h",yanchor="top"),
                        width=600,
                        height=320,
                        margin=dict(l=40, r=40, t=50, b=40)
                    )

                    st.plotly_chart(fig, use_container_width=False)

                    total_trades = len(df)
                # ------------------ Custom HTML Summary ------------------ #
                full_html = f"""
                <div class="result-card">
                    <h3 style="color:#00ffae;font-size:25px;">Volume Information of {token_selected}</h3>
                    <div class="metrics-grid">
                        <div class="metric-box">
                            <div class="metric-label">Total Volume</div>
                            <div class="metric-value">${total_volume:,.2f}</div>
                        </div>
                        <div class="metric-box">
                            <div class="metric-label">Total Trades</div>
                            <div class="metric-value">{total_trades}</div>
                        </div>
                        <div class="metric-box">
                            <div class="metric-label">Total Fees</div>
                            <div class="metric-value">${total_fees:,.2f}</div>
                        </div>
                    </div>
                </div>
                <style>
                    .result-card {{
                        background-color: #0f172a;
                        border-radius: 10px;
                        padding: 30px;
                        margin-top: 20px;
                        width: 94.5%;
                        position: relative;
                        font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    }}
                    .result-card::before {{
                        content: "";
                        position: absolute;
                        top: -3px;
                        left: -3px;
                        right: -3px;
                        bottom: -3px;
                        border-radius: 10px;
                        background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                        background-size: 600% 600%;
                        animation: neonBorder 6s ease infinite;
                        z-index: -1;
                        -webkit-mask: linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
                        -webkit-mask-composite: xor;
                        mask-composite: exclude;
                    }}
                    .metrics-grid {{
                        font-size: 20px;
                        display: flex;
                        flex-wrap: wrap;
                        gap: 20px;
                        margin-top: 20px;
                    }}
                    .metric-box {{
                        background-color: #1e293b;
                        padding: 20px;
                        border-radius: 12px;
                        flex: 1 1 200px;
                        min-width: 200px;
                    }}
                    .metric-label {{
                        font-size: 20px;
                        color: #cccccc;
                        margin-bottom: 6px;
                    }}
                    .metric-value {{
                        font-size: 1.6em;
                        color: #00ffae;
                        font-weight: bold;
                    }}
                    @keyframes neonBorder {{
                        0%   {{ background-position: 0% 50%; }}
                        50%  {{ background-position: 100% 50%; }}
                        100% {{ background-position: 0% 50%; }}
                    }}
                </style>
                """
                components.html(full_html, height=400, width=1900, scrolling=False)
            #except:
            #    st.info("⚠️ Error in Load your file. Please, verify if you download the correct Trade history File from Backpack site and be sure to not do any modification in this file.")

        else:
            st.info("📥 Please, load your CSV Trade Fills File from your Backpack Profile.")

    elif st.session_state.pagina == "🌾 Farm with YT":

    # Bloco HTML com a classe para hover

        # --- Novo Quadro de Inputs ---

        import streamlit as st

        # Estilo principal
        st.markdown("""
            <style>
            /* Aplica o estilo neon diretamente ao container principal */
            section.main > div.block-container {
                position: relative;
                background: #111827;
                border-radius: 16px;
                padding: 30px 25px;
                margin-top: 20px;
                box-shadow: 0 0 20px #39ff1455;
                font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                z-index: 1;
                font-size: px;
            }

            section.main > div.block-container::before {
                content: "";
                position: absolute;
                top: -3px;
                left: -3px;
                right: -3px;
                bottom: -3px;
                border-radius: 18px;
                z-index: -1;
                background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                background-size: 600% 600%;
                animation: neonBorder 6s ease infinite;
                padding: 3px;
                -webkit-mask:
                    linear-gradient(#fff 0 0) content-box,
                    linear-gradient(#fff 0 0);
                -webkit-mask-composite: xor;
                mask-composite: exclude;
            }

            @keyframes neonBorder {
                0%   { background-position: 0% 50%; }
                50%  { background-position: 100% 50%; }
                100% { background-position: 0% 50%; }
            }

            .neon-title {
                font-size: 25px;
                text-align: center;
                color: #00ffae;
                margin-bottom: 20px;
                font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            }

            h4 {
                color: #39FF14;
                font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                font-size: 20px;
                margin-top: 30px;
                margin-bottom: 15px;
            }
            </style>
        """, unsafe_allow_html=True)

        # Título principal
        st.markdown('<div class="neon-title">Parameters to YT ROI Estimation</div>', unsafe_allow_html=True)

        # Inputs: Basic Parameters
        st.markdown("#### Basic Parameters")
        col1, col2, col3 = st.columns(3)
        with col1:
            invested = st.number_input("Value to Invest ($):", min_value=0.0, value=1000.0, step=1.0)
        with col2:
            tsp = st.number_input("Token Total Supply (B):", min_value=0, value=1) * 1_000_000_000
        with col3:
            drop = st.number_input("Supply for Airdrop (%):", min_value=0.0, max_value=100.0, value=5.0)

        # Inputs: FDV
        st.markdown("#### No YT strategy in the Moment!")
        #col1, col2, col3 = st.columns(3)
        #with col1:
        #    Ena_fdv = st.number_input("Ethena FDV at TGE ($M):", min_value=0, value=5000, step=1) * 1_000_000
        #with col2:
        #    Level_fdv = st.number_input("Level FDV at TGE ($M):", min_value=0, value=150, step=1) * 1_000_000
        #with col3:
        #    ky_fdv = st.number_input("Kyros FDV ($M):", min_value=0, value=40, step=1) * 1_000_000
        #with col4:
            #url = "https://api.coingecko.com/api/v3/simple/price?ids=spark-2&vs_currencies=usd"
            #data = requests.get(url).json()
            #price = data.get('spark-2', {}).get('usd', 0.032)
            #print(price)
            #SpFDV = (float(price)*977940501)/1_000_000
            #if SpFDV is None:
            #    SpFDV = 0
            #print(SpFDV)
            #Sp_fdv = st.number_input("Spark FDV ($M):", min_value=0.0, value= SpFDV, step=1.0) * 1_000_000
        #with col5:
            #Gaib_fdv = st.number_input("Gaib FDV ($M):", min_value=0, value=30, step=1) * 1_000_000
            
        # Inputs: TGE Dates
        #st.markdown("#### Expected TGE Dates")
        #col1, col2, col3 = st.columns(3)
        #with col1:
        #    Ena_l_date = st.text_input("📅 Ethena TGE Date:", value="2025-09-25")
        #with col2:
        #    Level_l_date = st.text_input("📅 Level TGE Date:", value="2025-09-30")
        #with col3:
        #    Ky_l_date = st.text_input("📅 Kyros TGE Date:", value="2025-09-30")  
        #with col5:
            #Gaib_l_date = st.text_input("📅 Gaib TGE Date:", value="2025-08-14")

        
        st.markdown("""
        <style>
        .neon-divider {
            height: 6px;
            border: none;
            background: linear-gradient(90deg, #39FF14, #00F0FF, #39FF14, #00F0FF);
            background-size: 400% 100%;
            animation: pulseNeon 2.5s linear infinite;
            border-radius: 6px;
            margin: 20px 0 20px 0;
            box-shadow: 0 0 12px #39FF14AA, 0 0 24px #00F0FFAA;
        }

        @keyframes pulseNeon {
            0% { background-position: 0% 50%; }
            100% { background-position: 100% 50%; }
        }
        </style>

        <hr class="neon-divider">
        """, unsafe_allow_html=True)
        # Botão
        update_button = st.button("Refresh YT")

        # Verifica se já se passaram 120 segundos
        if update_button or not protocolos or elapsed_seconds < 10:
            with st.spinner('Loading Data and Calculating Parameters...'):
                #try: 
                # Busca Informações no Defillama
                Ena_tvl,Ena_amount,Ena_leadInvestors,Ena_otherInvestors = get_defillama_info("ethena","Ethereum")
                #Level_tvl,Level_amount,Level_leadInvestors,Level_otherInvestors = get_defillama_info("level","Ethereum")
                #Frag_tvl,Frag_amount,Frag_leadInvestors,Frag_otherInvestors = get_defillama_info("fragmetric","Solana")
                #Ky_tvl,Ky_amount,Ky_leadInvestors,Ky_otherInvestors = get_defillama_info("kyros","Solana")
                #Sp_tvl,Sp_amount,Sp_leadInvestors,Sp_otherInvestors = get_defillama_info("spark","Ethereum")
                
                # Busca dados dos protocolos em suas respectivas API's
                tags, values, time_Open, time_Level,top100,total_users = enviar_dados()
                Ena_accured,Ena_total_users, Ena_top100p = get_ethena_Data()
                #Frag_accured,Frag_unApy,solAsUSD,fragAsUSD,fragBySol,Frag_total_users = get_fragmetric_data()
                #Ky_accured,Ky_unApy,KyAsUSD,Ky_total_users,Ky_top100p = get_leader_kyros_function()
                #Sp_accured,Sp_top100p,Sp_total_users,Sp_tokens_per_day = get_Leader_Spark_Data()
                #Gaib_accured,Gaib_top100p,Gaib_total_users,Gaib_tvl = get_leader_Gaib_function()
                print(Ena_accured,Ena_total_users, Ena_top100p)
                # Busca dados dos protocolos nas API's da Pendle (Rede Ethereum) e Rate-X (Rede Solana)
                Ena_ytMul,Ena_unApy,Ena_impApy,Ena_feeRate,Ena_swapFee,Ena_ytRoi,Ena_expiry,Ena_priceImpact = get_Pendle_Data("0xa36b60a14a1a5247912584768c6e53e1a269a9f7","0x029d6247adb0a57138c62e3019c92d3dfc9c1840")
                #Level_ytMul,Level_unApy,Level_impApy,Level_feeRate,Level_swapFee,Level_ytRoi,Level_expiry,Level_priceImpact = get_Pendle_Data("0xc88ff954d42d3e11d43b62523b3357847c29377c","0x47247749e976c54c6db2a9db68c5cadb05482e9f")
                #Frag_ytMul,Frag_Multiplier,Frag_expiry,Frag_swapFee,Frag_priceImpact,time_Frag,symbol_frag = get_rateX_data("fragmetric")
                #ky_ytMul,ky_Multiplier,ky_expiry,ky_swapFee,ky_priceImpact,time_ky,symbol_ky = get_rateX_data("kyros")
                #Sp_ytMul,Sp_unApy,Sp_impApy,Sp_feeRate,Sp_swapFee,Sp_ytRoi,Sp_expiry,Sp_priceImpact = get_Pendle_Data("0xdace1121e10500e9e29d071f01593fd76b000f08","0x4eb0bb058bcfeac8a2b3c2fc3cae2b8ad7ff7f6e")
                #Gaib_ytMul,Gaib_unApy,Gaib_impApy,Gaib_feeRate,Gaib_swapFee,Gaib_ytRoi,Gaib_expiry,Gaib_priceImpact = get_Pendle_Data("0x47306e3cb4e325042556864b38aa0cbe8d928be5","0x05db2d5f89b3e9eab8f9c07149cd3a7575db8b9d")
                print(Ena_ytMul,Ena_unApy)
                # Formata a data atual e as datas de TGE (informadas pelo usuário) para que possam ser subtraídas
                date_obj = datetime.strptime(time_Level, "%Y-%m-%d %H:%M:%S")
                date_utc_formatada = date_obj.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
                date1 = datetime.strptime(date_utc_formatada , "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                date2 = datetime.strptime(Ena_expiry, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                #date3 = datetime.strptime(Level_expiry, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                #date4 = datetime.strptime(Frag_expiry, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                #date5 = datetime.strptime(ky_expiry, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                #date6 = datetime.strptime(Sp_expiry, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                ##date7 = datetime.strptime(Gaib_expiry, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                
                # Calcula os parâmetros de cada Protocolo
                # EnaEden
                #Ena_date_tge = datetime.strptime((Ena_l_date+"T00:00:00.000Z"), "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                #Ena_mean_daily = 1.3*(Ena_accured-Ena_TP_0)/((date1-Ena_date0).days)
                #Ena_points_tge = round(Ena_accured + (((Ena_date_tge-date1).days)*Ena_mean_daily),0)
                #Ena_points_per_token = round(Ena_points_tge/(tsp*drop/100),2)
                #Ena_farmed_yield = round(invested*Ena_ytMul*Ena_unApy*(date2-date1).days/365,2)
                #Ena_daily_pts_farmed = round(invested*Ena_ytMul*Ena_Multipleir*Ena_Boost*Ena_pts_token,2)
                #Ena_total_pts_farmed = round(Ena_daily_pts_farmed*(date2-date1).days,2)
                #Ena_etimated_tokens = round(Ena_total_pts_farmed/Ena_points_per_token,2)
                #Ena_airdrop_value = round((Ena_fdv/tsp)*Ena_etimated_tokens,2)
                #Ena_cost = abs(round((Ena_farmed_yield - invested - (invested*abs(Ena_priceImpact))),2))
                #Ena_profit = round((Ena_airdrop_value - Ena_cost),2)
                #Ena_ROI = round((100*Ena_profit/Ena_cost),2)
                
                #Ena_grade = protocol_rate(Ena_tvl,(100*Ena_top100p),Ena_ROI,(100*Ena_mean_daily/Ena_accured),total_users[0],"bom")
                
                
                # Fragmetric
                #Frag_date_tge = datetime.strptime((Frag_l_date+"T00:00:00.000Z"), "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                #Frag_mean_daily = 1.3*(Frag_accured-Frag_TP_0)/((date1-Frag_date0).days)
                #Frag_points_tge = round(Frag_accured + (((Frag_date_tge-date1).days)*Frag_mean_daily),0)
                #Frag_points_per_token = round(Frag_points_tge/(tsp*drop/100),2)
                #Frag_farmed_yield = round((invested)*Frag_ytMul*Frag_unApy*(date4-date1).days/365,2)
                #Frag_daily_pts_farmed = round((invested/fragAsUSD)*Frag_ytMul*Frag_Multipleir*Frag_Boost*Backpack_Boost*Frag_pts_token,2)
                #Frag_total_pts_farmed = round(Frag_daily_pts_farmed*(date4-date1).days,2)
                #Frag_etimated_tokens = round(Frag_total_pts_farmed/Frag_points_per_token,2)
                #Frag_airdrop_value = round((Frag_fdv/tsp)*Frag_etimated_tokens,2)
                #Frag_cost = abs(round(((Frag_farmed_yield) - invested - abs(fragAsUSD*Frag_swapFee)),2))
                #Frag_profit = round((Frag_airdrop_value - Frag_cost),2)
                #Frag_ROI = round((100*Frag_profit/Frag_cost),2)

                #Frag_grade = protocol_rate(Frag_tvl,60,Frag_ROI,(100*Frag_mean_daily/Frag_accured),Frag_total_users,"top")

                # Kyros
                #Ky_date_tge = datetime.strptime((Ky_l_date+"T00:00:00.000Z"), "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                #Ky_mean_daily = 1.3*(Ky_accured-Ky_TP_0)/((date1-Ky_date0).days)
                #Ky_points_tge = round(Ky_accured + (((Ky_date_tge-date1).days)*Ky_mean_daily),0)
                #Ky_points_per_token = round(Ky_points_tge/((tsp)*2*drop/100),2)
                #Ky_farmed_yield = round((invested)*ky_ytMul*Ky_unApy*(date5-date1).days/365,2)
                #Ky_daily_pts_farmed = round((invested/KyAsUSD)*ky_ytMul*Ky_Multipleir*Ky_Boost*ky_pts_token,2)
                #Ky_total_pts_farmed = round(Ky_daily_pts_farmed*(date5-date1).days,2)
                #Ky_etimated_tokens = round(Ky_total_pts_farmed/Ky_points_per_token,2)
                #Ky_airdrop_value = round((ky_fdv/tsp)*Ky_etimated_tokens,2)
                #Ky_cost = abs(round(((Ky_farmed_yield) - invested - abs(KyAsUSD*ky_swapFee)),2))
                #Ky_profit = round((Ky_airdrop_value - Ky_cost),2)
                #Ky_ROI = round((100*Ky_profit/Ky_cost),2)

                #Ky_grade = protocol_rate(Ky_tvl,(100*Ky_top100p),Ky_ROI,(100*Ky_mean_daily/Ky_accured),Ky_total_users,"bom")

                # Spark
                #Sp_date_tge = datetime.strptime((Sp_l_date+"T00:00:00.000Z"), "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                #Sp_mean_daily = 1.3*(Sp_accured - Sp_TP_0)/((date1-Sp_date0).days)
                #Sp_points_tge = round(Sp_accured + (((Sp_date_tge-date1).days)*Sp_mean_daily),0)
                #Sp_points_per_token = round(Sp_points_tge/(10000000000*drop/100),2)
                #Sp_farmed_yield = round(invested*Sp_ytMul*Sp_unApy*(date6-date1).days/365,2)
                #Sp_daily_pts_farmed = round(invested*Sp_ytMul*Sp_Multipleir*Sp_Boost*Sp_pts_token,2)
                #Sp_total_pts_farmed = round(Sp_daily_pts_farmed*(date6-date1).days,2)
                #Sp_etimated_tokens = round(Sp_total_pts_farmed/Sp_points_per_token,2)
                #Sp_airdrop_value = round((Sp_fdv/10000000000)*Sp_etimated_tokens,2)
                #Sp_cost = abs(round((Sp_farmed_yield - invested - (invested*abs(Sp_priceImpact))),2))
                #Sp_profit = round((Sp_airdrop_value - Sp_cost),2)
                #Sp_ROI = round((100*Sp_profit/Sp_cost),2)

                #Sp_grade = protocol_rate(Sp_tvl,(100*Sp_top100p),Sp_ROI,(100*Sp_mean_daily/Sp_accured),Sp_total_users,"bom")

                # Gaib
                #Gaib_date_tge = datetime.strptime((Gaib_l_date+"T00:00:00.000Z"), "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
                #Gaib_mean_daily = 1.3*(Gaib_accured - Gaib_TP_0)/((date1-Gaib_date0).days)
                #Gaib_points_tge = round(Gaib_accured + (((Gaib_date_tge-date1).days)*Gaib_mean_daily),0)
                #Gaib_points_per_token = round(Gaib_points_tge/((tsp)*drop/100),2)
                #Gaib_farmed_yield = round(invested*Gaib_ytMul*Sp_unApy*(date6-date1).days/365,2)
                #Gaib_daily_pts_farmed = round(invested*Gaib_ytMul*Gaib_Multipleir*Gaib_Boost*Gaib_pts_token,2)
                #Gaib_total_pts_farmed = round(Gaib_daily_pts_farmed*(date7-date1).days,2)
                #Gaib_etimated_tokens = round(Gaib_total_pts_farmed/Gaib_points_per_token,2)
                #Gaib_airdrop_value = round((Gaib_fdv/tsp)*Gaib_etimated_tokens,2)
                #Gaib_cost = abs(round((Gaib_farmed_yield - invested - (invested*abs(Gaib_priceImpact))),2))
                #Gaib_profit = round((Gaib_airdrop_value - Gaib_cost),2)
                #Gaib_ROI = round((100*Gaib_profit/Gaib_cost),2)

                #Gaib_grade = protocol_rate(Gaib_tvl,(100*Gaib_top100p),Gaib_ROI,(100*Gaib_mean_daily/Gaib_accured),Gaib_total_users,"muito bom")
                #except:
                #    print("Error in YT Data Request")
            try:
                protocolos = {
                    "Ethena sUSDE": {
                        "Imagem": "https://pbs.twimg.com/profile_images/1684292904599126016/f0ChONgU_400x400.png",
                        "Logo": "https://pbs.twimg.com/profile_images/1684292904599126016/f0ChONgU_400x400.png",
                        "pureLink": "https://app.ethena.fi/join/yp9pg",
                        "Link": "<a href='https://app.ethena.fi/join/yp9pg' target='_blank' style='color:#FFA500;'>More info</a>",
                        "Grade": f"{Ena_grade}",
                        "TVL": f"{Ena_tvl} M",
                        "Last Update": f"{time_Level}",
                        "Expiry": f"{date3.date()}",
                        "Total Points Farmed": f"{round(Ena_accured,0)}",
                        "YT Multiplier": f"{round(Ena_ytMul,3)}",
                        "YT APY": f"{round(Ena_unApy*100,2)}",
                        "Time Until Expiration": f"{(date3-date1)}",
                        "Protocol YT Multiplier": f"{Ena_Multipleir}",
                        "Protocol Referral Boost": f"{round((Ena_Boost-1),2)*100} %",
                        "Equivalent YT Received": f"$ {round(invested*Ena_ytMul,2)}",
                        "Daily Points Farmed": f"{Ena_daily_pts_farmed}",
                        "Total Points Farmed in YT": f"{Ena_total_pts_farmed}",
                        "Top 100 Concentration": f"{round(100*Ena_top100p,2)}",
                        "Total User": f"{Ena_total_users}",
                        "Farmed Yield in YT": f"$ {Ena_farmed_yield}",
                        "Mean Daily Points": f"{round(Ena_mean_daily,0)}",
                        "Estimated Points in TGE": f"{round(Ena_points_tge,0)}",
                        "Points per Token": f"{Ena_points_per_token}",
                        "Estimated FDV in TGE": f"{Ena_fdv}",
                        "Estimated Token Price": f"$ {round(Ena_fdv/tsp,2)}",
                        "Estimated Tokens Airdrop": f"{Ena_etimated_tokens}",
                        "Estimated Airdrop Value": f"$ {Ena_airdrop_value}",
                        "Expected Profit": f"$ {Ena_profit}",
                        "Expected ROI": f"{Ena_ROI} %"      
                    },
                    #"Kyros": {
                    #    "Imagem": "https://raw.githubusercontent.com/Caio-Fl/Airdrop_Points/main/kyros.jpg",
                    #    "Logo": "https://pbs.twimg.com/profile_images/1847426788252590080/-Tb-I1Yl_400x400.jpg",
                    #    "pureLink": "https://boost.kyros.fi/?ref=nq3orn",
                    #    "Link": "<a href='https://boost.kyros.fi/?ref=nq3orn' target='_blank' style='color:#1E90FF;'>More info</a> - <a href='https://app.rate-x.io/referral?ref=H9GnKZON' target='_blank'>🔗 Rate-X </a>",
                    #    "Grade": f"{Ky_grade}",
                    #    "TVL": f"{Ky_tvl} M",
                    #    "Last Update": f"{time_Level}",
                    #    "Expiry": f"{date5.date()}",
                    #    "kySOL Price": f"{round((KyAsUSD),2)}",
                    #    "Equivalent Invested in kySOL": f"{round((invested/KyAsUSD),3)}",
                    #    "Total Points Farmed": f"{round(Ky_accured,0)}",
                    #    "YT Multiplier": f"{round(ky_ytMul,3)}",
                    #    "YT APY": f"{round(Ky_unApy*100,2)}",
                    #    "Time Until Expiration": f"{(date5-date1)}",
                    #    "Protocol YT Multiplier": f"{Ky_Multipleir}",
                    #    "Protocol Referral Boost": f"{round((Ky_Boost-1),2)*100} %",
                    #    "Equivalent YT Received": f"$ {round(invested*ky_ytMul,2)}",
                    #    "Daily Points Farmed": f"{Ky_daily_pts_farmed}",
                    #    "Total Points Farmed in YT": f"{Ky_total_pts_farmed}",
                    #    "Top 100 Concentration": f"{round(100*Ky_top100p,2)}",
                    #    "Total User": f"{Ky_total_users}",
                    #    "Farmed Yield in YT": f"$ {Ky_farmed_yield}",
                    #    "Mean Daily Points": f"{round(Ky_mean_daily,0)}",
                    #    "Estimated Points in TGE": f"{round(Ky_points_tge,0)}",
                    #    "Points per Token": f"{Ky_points_per_token}",
                    #    "Estimated FDV in TGE": f"{ky_fdv}",
                    #    "Estimated Token Price": f"$ {ky_fdv/tsp}",
                    #    "Estimated Tokens Airdrop": f"{Ky_etimated_tokens}",
                    #    "Estimated Airdrop Value": f"$ {Ky_airdrop_value}",
                    #    "Expected Profit": f"$ {Ky_profit}",
                    #    "Expected ROI": f"{Ky_ROI} %"   
                    #},
                }
            except:
                protocolos = {}
            
        # --- Salvar automaticamente sem botão ---
        salvar_json(protocolos)

        # --- Estado de seleção ---
        if "protocolo_selecionado" not in st.session_state:
            st.session_state.protocolo_selecionado = None
        if st.session_state.protocolo_selecionado:
            # Mostrar detalhes do protocolo selecionado
            p = st.session_state.protocolo_selecionado
            st.markdown(
                f"""
                <div style="
                    border: 3px solid white;
                    border-radius: 10px;
                    padding: 10px;
                    background-color: #342b44;
                    color: white;
                    margin-bottom: 5px;
                ">
                    <div style="display: flex; align-items: center; justify-content: center; gap: 10px; margin-bottom: 5px;font-size: 25px;">
                        <img src="{protocolos[p]['Logo']}" width="70" height="70" style="border-radius: 5%;">
                        <h4 style="margin: 5;font-size: 25px;color: #FFA500">{p}</h4>
                    </div>
                """,
                unsafe_allow_html=True
            )
            st.markdown(f"""
            <style>
            .zoom-title {{
                display: inline-block;
                transition: transform 0.3s ease;
            }}
            .zoom-title:hover {{
                transform: scale(1.03);
                color: #14ffe9;
            }}
            </style>
            <div style="background-color: #376a94; padding: 20px; border: 2px solid white; border-radius: 10px; margin-top: 20px; font-size: 25px;margin-bottom: 5px;">
                <h2 class="zoom-title">Details {p}</h2>
                <p style="font-size:25px;">
                    <strong>Protocol:</strong> {p} – 
                    <a href="{protocolos[p]['pureLink']}" target="_blank" style="color:#1E90FF;">Visit Protocol</a>
                </p>
                <ul>
                    <li><strong>TVL:</strong> {protocolos[p]['TVL']}</li>
                    <li><strong>Total Points Farmed:</strong> {protocolos[p]['Total Points Farmed']} XP</li>
                    <li><strong>Last Update:</strong> {protocolos[p]['Last Update']}</li>
                    <li><strong>YT Multiplier:</strong> {protocolos[p]['YT Multiplier']} x</li>
                    <li><strong>YT APY:</strong> {protocolos[p]['YT APY']}</li>
                    <li><strong>Time Until YT Expiration:</strong> {protocolos[p]['Time Until Expiration']}</li>
                    <li><strong>Protocol YT Multiplier:</strong> {protocolos[p]['Protocol YT Multiplier']}</li>
                    <li><strong>Protocol Referral Boost:</strong> {protocolos[p]['Protocol Referral Boost']}</li>
                    <li><strong>Equivalent YT Received:</strong> {protocolos[p]['Equivalent YT Received']}</li>
                    <li><strong>Daily Points Farmed:</strong> {protocolos[p]['Daily Points Farmed']} XP</li>
                    <li><strong>Total Points Farmed in YT Expiration:</strong> {protocolos[p]['Total Points Farmed in YT']} XP</li>
                    <li><strong>Top 100 Points Percentual Concentration:</strong> {protocolos[p]['Top 100 Concentration']}%</li>
                    <li><strong>Total Users:</strong> {protocolos[p]['Total User']}</li>
                    <li><strong>Farmed Yield in YT Expiration:</strong> {protocolos[p]['Farmed Yield in YT']}</li>
                    <li><strong>Mean Daily Points:</strong> {protocolos[p]['Mean Daily Points']} XP</li>
                    <li><strong>Estimated Total Points in TGE:</strong> {protocolos[p]['Estimated Points in TGE']} XP</li>
                    <li><strong>Points to Receive 1 Token:</strong> {protocolos[p]['Points per Token']}</li>
                    <li><strong>Estimated FDV in TGE:</strong> {protocolos[p]['Estimated FDV in TGE']}</li>
                    <li><strong>Estimated Token Price:</strong> {protocolos[p]['Estimated Token Price']}</li>                  
                    <li><strong>Estimated Tokens Airdrop:</strong> {protocolos[p]['Estimated Tokens Airdrop']}</li>
                    <li><strong>Estimated Airdrop Value:</strong> {protocolos[p]['Estimated Airdrop Value']}</li>
                    <li><strong>Expected Profit:</strong> {protocolos[p]['Expected Profit']}</li>
                    <li><strong>Expected ROI:</strong> {protocolos[p]['Expected ROI']}</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)

            if st.button("🔙 Return"):
                st.session_state.protocolo_selecionado = None

        else:
            # Centraliza o conteúdo
            
            cols = st.columns(3, gap="large")
            for i, (nome, dados) in enumerate(protocolos.items()):
                with cols[i % 3]:
                    try:
                        st.markdown(f"""
                            <div class="airdrop-box" style="
                                background: radial-gradient(circle at top center, #2c2f35 0%, #1c1e22 100%);
                                border: 0px double #555;
                                box-shadow: 
                                    0 0 10px #00ff00,
                                    0 0 8px #00ff00,
                                    0 0 8px #00ff00,
                                    0 0 8px #00ff00;
                                border-radius: 10px;
                                width: 100%;
                                padding: 20px;
                                margin: 15px 0;
                                color: white;
                                text-align: center;
                                font-family: 'Space Grotesk', sans-serif;
                                display: flex;
                                flex-direction: column;
                                justify-content: space-between;
                            ">
                                <div style="border: 2px solid #444; padding: 10px; radial-gradient(circle at top center, #2c2f35 0%, #1c1e22 100%); border-radius: 6px; margin: 15px 0; font-size: 25px;">
                                    <img src="{dados['Logo']}" width="70" />
                                    <div style="margin-top: 15px; font-weight: bold;">{nome}</div>
                                    <div style="border: 2px solid #444; background: radial-gradient(circle at top center, #2c2f35 0%, #1c1e22 100%);border-radius: 16px;box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);padding: 10px; border-radius: 6px; margin: 15px 0; font-size: 22px; text-align: center;">
                                        <p><strong>TVL:</strong> {dados['TVL']}</p>
                                        <p><strong>Expected ROI:</strong> {dados['Expected ROI']}</p>
                                        <p><strong>YT Multiplier:</strong> {dados['YT Multiplier']}</p>
                                        <p><strong>YT APY:</strong> {dados['YT APY']} %</p>
                                        <p><strong>Boost:</strong> {dados['Protocol Referral Boost']}</p>
                                        <p><strong>Expiry:</strong> {dados['Expiry']}</p>
                                        <p><strong>Rating:</strong> {dados['Grade']}</p>
                                    </div>
                                    <a href="{dados['pureLink']}" target="_blank">
                                        <button style="
                                            background-color: orange;
                                            color: black;
                                            font-family: 'Space Grotesk', sans-serif;
                                            padding: 10px 20px;
                                            border: none;
                                            border-radius: 5px;
                                            font-size: 25px;
                                            cursor: pointer;
                                            margin-bottom: 10px;
                                        "><strong>Visit Protocol</strong></button>
                                    </a>
                                </div>
                            </div>
                        """, unsafe_allow_html=True)
                        
                    except:

                        st.markdown(f"""
                            <div class="airdrop-box" style="
                                background: radial-gradient(circle at top center, #2c2f35 0%, #1c1e22 100%);
                                border: 0px double #555;
                                box-shadow: 
                                    0 0 10px #00ff00,
                                    0 0 8px #00ff00,
                                    0 0 8px #00ff00,
                                    0 0 8px #00ff00;
                                border-radius: 10px;
                                width: 100%;
                                padding: 20px;
                                margin: 15px 0;
                                color: white;
                                text-align: center;
                                font-family: 'Space Grotesk', sans-serif;
                                display: flex;
                                flex-direction: column;
                                justify-content: space-between;
                            ">
                                <div style="border: 2px solid #444; padding: 10px; radial-gradient(circle at top center, #2c2f35 0%, #1c1e22 100%); border-radius: 6px; margin: 15px 0; font-size: 25px;">
                                    <img src="{dados['Logo']}" width="70" />
                                    <div style="margin-top: 15px; font-weight: bold;">{nome}</div>
                                    <div style="border: 2px solid #444; background: radial-gradient(circle at top center, #2c2f35 0%, #1c1e22 100%);border-radius: 16px;box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);padding: 10px; border-radius: 6px; margin: 15px 0; font-size: 22px; text-align: center;">
                                        <p><strong>TVL:</strong> {dados['TVL']}</p>
                                        <p><strong>YT expired or in Maitenance</p>
                                    </div>
                                    <a href="{dados['pureLink']}" target="_blank">
                                        <button style="
                                            background-color: orange;
                                            color: black;
                                            font-family: 'Space Grotesk', sans-serif;
                                            padding: 10px 20px;
                                            border: none;
                                            border-radius: 5px;
                                            font-size: 25px;
                                            cursor: pointer;
                                            margin-bottom: 10px;
                                        "><strong>Visit Protocol</strong></button>
                                    </a>
                                </div>
                            </div>
                        """, unsafe_allow_html=True)
                                                
                    # 🔘 Botão funcional View Details
                        # Botão funcional — fora do HTML
                    st.markdown(
                        """
                        <div style='text-align: center; margin-top: 10px; margin-bottom: 10px;margin-left:30px;display: flex;align-items: center;justify-content: center;'>
                        """,
                        unsafe_allow_html=True
                    )

                    # Botão real dentro da div
                    if st.button(f"🔍 View Details: {nome}", key=f"view_{nome}"):
                        st.session_state.protocolo_selecionado = nome

                    # Fecha a div
                    st.markdown("</div>", unsafe_allow_html=True)
  
            st.markdown("""
            <style>
            .neon-divider {
                height: 6px;
                border: none;
                background: linear-gradient(90deg, #39FF14, #00F0FF, #39FF14, #00F0FF);
                background-size: 400% 100%;
                animation: pulseNeon 2.5s linear infinite;
                border-radius: 6px;
                margin: 20px 0 20px 0;
                box-shadow: 0 0 12px #39FF14AA, 0 0 24px #00F0FFAA;
            }

            @keyframes pulseNeon {
                0% { background-position: 0% 50%; }
                100% { background-position: 100% 50%; }
            }
            </style>

            <hr class="neon-divider">
            """, unsafe_allow_html=True)
            st.markdown(
            """
            <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    margin-bottom: 40px;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 22px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }

                /* Borda neon com gradiente animado */
                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }

                /* Efeito de hover */
                .airdrop-box:hover {
                    border-color: #00f0ff;
                    background: #262b33;
                }

                @keyframes neonBorder {
                    0%   { background-position: 0% 50%; }
                    50%  { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }

                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 0px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 5px;
                    margin-bottom: 0px;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 0px;
                }
            </style>
            <div class="airdrop-box">
                <div class="yt-farm-description">
                    <p>
                    Above are the potential returns you can achieve by participating in Yield Token (YT) farming strategies from various protocols that currently have ongoing airdrop campaigns.
                    </p>
                    <p>
                    Each protocol also receives a score based on its potential to deliver solid results, according to the parameters set by the user in the sidebar.
                    </p>
                    <p>
                    To view the full list of parameters for each protocol, double-click on <strong>"View Details"</strong>.
                    </p>
                    <p>
                    You need to adjust the Parameters to evaluate each Protocol, mainly the expected FDV in TGE. 
                    Do not expect that a protocol with $20M TVL will be able to do an TGE with $200M of FDV, <strong>be coherent in your expectations</strong>.
                    </p>
                    <p>
                    To apply new Parameters to ROI Estimation, click on <strong>"Refresh YT"</strong>.
                    </p>
                </div>
            </div>
        """, unsafe_allow_html=True)


    elif st.session_state.pagina == "📈 Pendle APY Prediction":
        #id's = 1 - ETH , 10 OP , 56 - BNB, 146 - SONIC LABS, 5000 - Mantle, 8453 - Base, 42161 - Arb, 80094 -BERA
        ids = [1, 56, 146, 5000, 8453, 42161, 80094, 10, 999]
        nets = ["Ethereum", "BNB Chain", "Sonic Labs", "Mantle", "Base", "Arbitrum", "Berachain", "Optimism", "Hyper"]

        all_markets_list = []

        for i, id in enumerate(ids):
            markets = get_pendle_markets(id)
            if not markets.empty:
                net = nets[i]
                markets = markets.copy()
                markets['expiry_date'] = markets["expiry"].str.split('T').str[0]
                markets['net'] = net
                markets["label"] = markets["name"] + " (Expires in: " + markets['expiry_date'] + ") " + net
                all_markets_list.append(markets)

        # Concatenar todos os DataFrames em um só
        markets = pd.concat(all_markets_list, ignore_index=True)

        #print(markets[['name', 'expiry_date', 'net', 'label']])
        # Exibindo a lista de seleção múltipla
        #selected_names = st.multiselect("Escolha um ou mais mercados", options)

        # Define o mercado desejado (exatamente como aparece no 'label')
        default_market_name = "SUSDE"  # substitua pela label desejada

        # Buscar índice do mercado default
        default_index = markets[markets["name"] == default_market_name].index
        default_index = int(default_index[0]) if not default_index.empty else 0  # fallback para 0 se não encontrar

        st.sidebar.markdown("<h3 style='font-size: 25px;'>Select Pendle Market</h3>", unsafe_allow_html=True)
        # Estilo CSS opcional
        st.markdown("""
            <style>
            div[data-baseweb="select"] {
                font-size: 18px;
            }
            </style>
        """, unsafe_allow_html=True)

        # Select de mercado
        col1, col2 = st.columns([2, 1])

        with col1:
            st.markdown("<label style='font-size:22px; font-weight:600;font-family: Trebuchet MS, Segoe UI, sans-serif;margin-bottom: 60px;'>Markets:</label>", unsafe_allow_html=True)
            selected_label = st.selectbox(
                " ",  # Oculta o label nativo
                markets["label"].tolist(),
                index=default_index,
                format_func=lambda x: x.upper()
            )

        with col2:
            st.markdown("<label style='font-size:22px; font-weight:600;font-family: Trebuchet MS, Segoe UI, sans-serif;margin-bottom: 60px;'>Time in:</label>", unsafe_allow_html=True)
            time_scale = st.selectbox(
                " ",  # Oculta o label nativo
                ["hour", "day"],
                format_func=lambda x: x.upper()
            )

        # Pega o mercado selecionado com base no label
        selected_row = markets[markets["label"] == selected_label].iloc[0]
        net_name = selected_row["net"]
        id = ids[nets.index(net_name)]

        # Pegando o address
        if not selected_row.empty:
            address = selected_row["address"]
            label = selected_row["label"]
            expires = selected_row["expiry"]
            st.markdown(f"<span style='font-size:25px'><strong>Selected Market:</strong> {label}</span>", unsafe_allow_html=True)
        else:
            st.markdown("Maket not Founded")

        with st.spinner('Loading Pendle Data to Plot Implied APY Tendency...'):
            df_implied,implied_apy,underlying_apy,base_apy,tvl_in_k,trend_line,upper_line,lower_line,trend_line_extended,upper_line_extended,lower_line_extended,dates,extended_dates,expiry_date,address = get_pendle_apy_data(selected_row,time_scale,id)

            # Supondo que base_apy seja uma lista ou Series
            base_apy_series = pd.Series(base_apy)

            # Calcular Q1, Q3 e IQR
            Q1 = base_apy_series.quantile(0.2)
            Q3 = base_apy_series.quantile(0.8)
            IQR = Q3 - Q1

            # Definir limites
            lower_bound = Q1 - 1.2 * IQR
            upper_bound = Q3 + 1.2 * IQR

            # Filtrar dados válidos
            mask = (base_apy_series >= lower_bound) & (base_apy_series <= upper_bound)
            filtered_dates = [d for d, keep in zip(dates, mask) if keep]
            filtered_base_apy = base_apy_series[mask].tolist()
            Range = upper_line[-1] - lower_line[-1]
            trend_perc = 100*(trend_line[-1] - lower_line[-1])/Range
            actual_perc = 100*(implied_apy[-1] - lower_line[-1])/Range
            time_now = datetime.now().now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            data1 = datetime.strptime(time_now, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            delta = expiry_date - data1

        # Criar figura
            fig = go.Figure()

            # Adiciona os traços (linhas) em uma linha cada
            fig.add_trace(go.Scatter(x=filtered_dates, y=filtered_base_apy, mode='lines+markers', name='Base APY', line=dict(color='blue', width=2, shape='spline'), marker=dict(color='blue', size=4)))
            fig.add_trace(go.Scatter(x=dates, y=implied_apy, mode='lines+markers', name='Implied APY', line=dict(color='green', width=2, shape='spline'), marker=dict(color='green', size=4)))
            fig.add_trace(go.Scatter(x=dates, y=trend_line, mode='lines', name='Implied APY Tendency', line=dict(color='black', width=2, dash='dash'), marker=dict(color='green', size=4)))
            fig.add_trace(go.Scatter(x=dates, y=upper_line, mode='lines', name='Maximum Implied APY Tendency', line=dict(color='red', width=2, dash='dash'), marker=dict(color='red', size=4)))
            fig.add_trace(go.Scatter(x=dates, y=lower_line, mode='lines', name='Minimum Implied APY Tendency', line=dict(color='#b3a13c', width=2, dash='dash'), marker=dict(color='#b3a13c', size=4)))
            fig.add_trace(go.Scatter(x=extended_dates, y=trend_line_extended[-len(extended_dates):], mode='lines', name='Tendency up to Expire', line=dict(color='black', width=2, dash='dot')))
            fig.add_trace(go.Scatter(x=extended_dates, y=upper_line_extended[-len(extended_dates):], mode='lines', name='Maximum Tendency up to Expire', line=dict(color='red', dash='dot')))
            fig.add_trace(go.Scatter(x=extended_dates, y=lower_line_extended[-len(extended_dates):], mode='lines', name='Minimum Tendency up to Expire', line=dict(color='green', dash='dot')))
            fig.add_trace(go.Scatter(x=dates, y=tvl_in_k, mode='lines+markers', name='TVL (Mi USD)', line=dict(color='orange', dash='dot'), yaxis='y2'))

            # Linha vertical para expiry
            fig.add_shape(
                type='line',
                x0=expiry_date,
                x1=expiry_date,
                y0=0,
                y1=1,
                line=dict(color='gray', dash='dot'),
                xref='x',
                yref='paper'
            )
            fig.add_annotation(
                x=expiry_date,
                y=0,
                text='Expiry Date',
                showarrow=False,
                xanchor='left',
                yanchor='bottom',
                xref='x',
                yref='paper',
                font=dict(size=20)
            )

            # Layout com dois eixos y e fonte aumentada
            fig.update_layout(
                title=f'YT-{selected_row["name"]} {expiry_date.date()} - History of Base APY, Implied APY, Tendency Lines and TVL',
                xaxis=dict(
                    title='Date',
                    color='white',
                    gridcolor='rgba(255, 255, 255, 0.05)',
                    showline=True,
                    linewidth=1,
                    linecolor='white',
                    tickfont=dict(size=25)
                ),
                yaxis=dict(
                    range=[0, None], 
                    title='APY (%)',
                    color='white',
                    gridcolor='rgba(255, 255, 255, 0.05)',
                    showline=True,
                    linewidth=1,
                    linecolor='white',
                    tickfont=dict(size=25)
                ),
                yaxis2=dict(title='TVL (Mi USD)', overlaying='y', side='right', tickfont=dict(size=20)),
                legend=dict(x=1, y=1, xanchor='right', yanchor='top', bgcolor='rgba(52,43,68,0.8)', font=dict(size=18)),
                font=dict(color='white', family='Space Grotesk, sans-serif',size=20),
                plot_bgcolor='#0f0f0f',
                paper_bgcolor= "#111827",
                hovermode='x unified',
                margin=dict(l=40, r=40, t=80, b=40),
                height=750
            )

            # Mostrar com zoom habilitado
            st.plotly_chart(fig, use_container_width=True)

            def normalizar_para_faixa(valor, min_origem=0, max_origem=100, min_dest=-100, max_dest=100):
                return min_dest + ((valor - min_origem) / (max_origem - min_origem)) * (max_dest - min_dest)

            valor = normalizar_para_faixa(round(actual_perc,2))
            figura, force = barra_compra_venda(valor,round(actual_perc,2))
            st.pyplot(figura)

            def get_token_info(marketAdd,id):
                url = f"https://api-v2.pendle.finance/core/v1/{id}/markets/{marketAdd}"
                response = requests.get(url)
                jsonn = json.loads(response.text)
                #jsonn.get("underlyingApy", {}) 
                ytRoi = jsonn.get("ytRoi", {})
                
                url = f"https://api-v2.pendle.finance/core/v1/sdk/{id}/markets/{marketAdd}/swapping-prices"
                response = requests.get(url)
                jsonn = json.loads(response.text)
                ytMult = jsonn.get("underlyingTokenToYtRate", {})
                return(ytMult,ytRoi)
            
            ytMult, ytRoi = get_token_info(address,id)
            # IA anwser
            criteria = criteria = """Learn the following Criteria, no mention it you just will use this to awnser the next question: An "Actual Implied APY"  next to "Maximum Historical APY" can be an excellent moment to sell the Yield Token if you already hold the token; An "Actual Implied APY"  next to "Minimum Historical APY" can be an excellent moment to buy the Yield Token; An "Actual Implied APY"  next or higher than "Actual Best Sell point of Implied APY" can be a great moment to sell the Yield Token if you already hold the token if you not holding it ignore; An "Actual Implied APY"  next or a little lower than "Actual Best Buy point of Implied APY" can be a great moment to buy the Yield Token this is an important factor; An "Actual Implied APY"  very lower than "Actual Best Buy point of Implied APY" indicates a possible problem with Yield Token and need to be excluded as an investiment; An "Actual Implied APY percentual in relation of Range" next or higher than 100 indicates na great moment to sell the Yield Tokenif you already hold the token; An "Actual Implied APY percentual in relation of Range" next to 0 or a little negative indicates na great moment to buy the Yield Token; An "Actual Implied APY percentual in relation of Range" very negative indicates a possible problem with Yield Token and need to be excluded as an investiment; If "Days to expiry" is lower than 20, this implies in a high risk to try trade it. An "Actual Underlying APY" higher than 9 is an excellent yield return for hold the YT token, but is necessary to evaluate the days until expiry. An "Actual Underlying APY" lower than 4 is not so good yield return for hold the YT token, but is necessary to evaluate the days until expiry. An "Actual Underlying APY" equals to 0 implies that the yield token do not have any yield return and is just used to farm points. The YT multiplier factor indicates the multiplication factor over the invested capital and the underlying APY gives the yield return based in (YT Multiplier x Invested Capital x Actual Underlying APY/100 x Days to expiry/365), so higher YT miltiplier means that you can receire more yield and farm with higher capital the protocol of token, but if underlying APY is zero you will just farm points in protocol airdrop. YT ROI for hold token up to expiry is the return at maturity o YT token, higher YT ROI means a smaller loss at maturity.
            """
            description = f"""
            Implied APY Data of {label}:
            Actual Implied APY = {round(implied_apy[-1],2)}
            Maximum Historical APY = {round(max(implied_apy),2)}
            Minimum Historical APY = {round(min(implied_apy),2)}
            Actual Best Sell point of Implied APY = {round(upper_line[-1],2)}
            Actual Mean Implied APY = {round(trend_line[-1],2)}
            Actual Best Buy point of Implied APY = {round(lower_line[-1],2)}
            Trade Force = {(-1)*force}
            Mean Implied APY percentual in relation of Range = {round(trend_perc,2)}
            Actual Implied APY percentual in relation of Range = {round(actual_perc,2)}
            Time to expiry and YT value goes to zero = {expiry_date}
            Days to expiry = {delta}
            Actual Underlying APY =  {round(underlying_apy[-1],2)}
            Mean Underlying APY = {round(statistics.mean(underlying_apy),2)}
            Maximum Underlying APY Historical = {round(max(underlying_apy),2)}
            Minimum Underlying APY Historical = {round(min(underlying_apy),2)}
            YT Protocol Multiplir = {ytMult}
            YT ROI for hold token up to expiry = {ytRoi*100} %
            """
            question_1 = f"""faced with two possible scenarios (answer in few lines):
                1st If I already have the YT token and I want to know if it is a good time to sell it?
                2nd If I don't have the YT token and I want to know if it is a good time to buy it, considering "Actual Underlying APY" next to "Actual Best Buy point of Implied APY" lower line, or if I should wait for a better opportunity?
                3nd If i want to farm point to airdrop of YT token protocol the YT Protocol Multiplier High and the YT ROI is higher than "-35" percent (consider the YT ROI signal in comparison)?
                According to the data description: {description}"""
            h = implied_apy#[(len(implied_apy)-50):-1]
            question_2 = f"""
                Verify the historical of implied APY and analysis if is this a good momment to Buy the YT to trade it or not, considering "Trade Force" next to 100 and "Actual Underlying APY" next to Actual "Best Buy point of Implied APY" lower tendency line as a great point to consider a entry trade. Also consider {delta} to expiry is of high risk?
                Historical Implied APY = {[round(x, 2) for x in h]}
                the Underlying APY is {round(underlying_apy[-1],2)}
            """
            questions = [criteria, question_1, question_2]
            #AIzaSyC4C4bMviJpBIuR7XXCqqPF81JrrYitlro - gemini

            language = "ingles"
            model = "mistral-large-latest"
            personality = "You are an finantial advisor evaluationg historical data"
            resposta= ""
            time.sleep(3)
            IA_1 = mistral_AI(question_1,language,model,personality)
            IA_1 = IA_1['content']
            time.sleep(3)
            IA_2 = mistral_AI(question_2,language,model,personality)
            IA_2 = IA_2['content']
            IA = [IA_1, IA_2]
            #IA = lang_IA(questions,criteria)

            if isinstance(IA, list):
                st.markdown("""
                    <style>
                    .ai-title {
                        font-size: 25px;
                        font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                        color: white !important;
                    }
                    </style>
                """, unsafe_allow_html=True)

                st.markdown("""
                    <h2 class="ai-title">🧠 AI Interpretation:</h2>
                """, unsafe_allow_html=True)
                blocks_html = ""
                for resposta in IA:
                    resposta_html = resposta.replace("**", "&nbsp;").replace("\n", "<br>")
                    blocks_html += f"""
                    <div class="airdrop-box">
                            <p style="font-size: 25px;">{resposta_html}</p>
                    </div>
                    """   
            #else:
            #    st.markdown(IA)
        
            full_html = f"""
                <style>
                    @keyframes neonBorder {{
                        0%   {{ background-position: 0% 50%; }}
                        50%  {{ background-position: 100% 50%; }}
                        100% {{ background-position: 0% 50%; }}
                    }}
                    .airdrop-box {{
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    margin-bottom: 40px;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 22px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }}
                /* Borda neon com gradiente animado */
                .airdrop-box::before {{
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }}
                /* Efeito de hover */
                .airdrop-box:hover {{
                    border-color: #00f0ff;
                    background: #262b33;
                }}
                .airdrop-box h1 {{
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 0px;
                }}
                .airdrop-box h2 {{
                    font-size: 25px;
                    margin-top: 5px;
                    margin-bottom: 0px;
                }}
                .airdrop-box ul {{
                    margin-left: 20px;
                    margin-bottom: 0px;
                }}
                </style>
                <div class="container-externa">
                    {blocks_html}
                </div>
            """

            components.html(full_html, height=1000, width=1900, scrolling=True)


        
        st.markdown("---")  # Linha separadora entre blocos

        st.markdown(
        """
            <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    margin-bottom: 40px;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 22px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }

                /* Borda neon com gradiente animado */
                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }

                /* Efeito de hover */
                .airdrop-box:hover {
                    border-color: #00f0ff;
                    background: #262b33;
                }

                @keyframes neonBorder {
                    0%   { background-position: 0% 50%; }
                    50%  { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }

                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 0px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 5px;
                    margin-bottom: 0px;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 0px;
                }
            </style>
        <div class="airdrop-box">
        <h3>📈 How the Pendle APY Chart Works?</h3>
        <p>
        The APY (Annual Percentage Yield) chart on Pendle shows the evolution of annualized yield rates for tokenized assets over time.
        It is built with multiple layers of information to support deeper analysis:
        </p>
        <ul>
            <li><strong>Base APY</strong>: This reflects the base yield from the underlying protocol, excluding any incentive rewards.</li>
            <li><strong>Implied APY</strong>: This is the market-implied yield calculated from the price of the YT (Yield Token). It represents the market's expectation of future returns up to the maturity date.</li>
            <li><strong>Trend Line</strong>: A trend line is applied to the Implied APY, helping to visualize the general direction of expected yields—whether they are increasing, decreasing, or stable.</li>
            <li><strong>TVL (Total Value Locked)</strong>: Displayed on a secondary axis, it shows how much value is allocated to the asset, giving insight into investor interest over time.</li>
        </ul>
        <p>
        In addition, <strong>buying opportunities</strong> may arise when the Implied APY touches the <span style='color:red;'><strong>minimum green trend line</strong></span>,
        indicating a temporary undervaluation. Conversely, <strong>selling opportunities</strong> may be more favorable when the Implied APY reaches the <span style='color:red;'><strong>maximum red trend line</strong></span>,
        suggesting a potential overvaluation.
        </p>
        <p>
        This chart allows users to compare real and expected yields, monitor market sentiment, and identify attractive entry and exit points for Pendle pools.
        </p>
        </div>
        """,
        unsafe_allow_html=True
        )
        
    elif st.session_state.pagina == "🎁 Latest Airdrops":
        st.info("🚧 Coming Soon: Protocols with Airdrop Potential.")

    elif st.session_state.pagina == "♾️ PerpDEX Airdrops":

        import streamlit.components.v1 as components

        # =========================
        # DESCRIÇÃO (BOX TOPO)
        # =========================
        st.markdown(
                    """
                    <style>
                        .airdrop-box {
                            position: relative;
                            z-index: 1;
                            border-radius: 12px;
                            padding: 25px;
                            margin: 20px 0;
                            margin-bottom: 40px;
                            background: #111827;
                            display: flex;
                            flex-direction: column;
                            gap: 30px;
                            font-size: 22px;
                            color: white;
                            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                            overflow-wrap: break-word;
                            word-wrap: break-word;
                            white-space: normal;
                            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                        }

                        /* Borda neon com gradiente animado */
                        .airdrop-box::before {
                            content: "";
                            position: absolute;
                            top: -3px;
                            left: -3px;
                            right: -3px;
                            bottom: -3px;
                            border-radius: 14px;
                            z-index: -1;
                            background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                            background-size: 600% 600%;
                            animation: neonBorder 6s ease infinite;
                            padding: 3px;
                            -webkit-mask:
                                linear-gradient(#fff 0 0) content-box,
                                linear-gradient(#fff 0 0);
                            -webkit-mask-composite: xor;
                            mask-composite: exclude;
                        }

                        /* Efeito de hover */
                        .airdrop-box:hover {
                            border-color: #00f0ff;
                            background: #262b33;
                        }

                        @keyframes neonBorder {
                            0%   { background-position: 0% 50%; }
                            50%  { background-position: 100% 50%; }
                            100% { background-position: 0% 50%; }
                        }

                        .airdrop-box h1 {
                            font-size: 25px;
                            text-align: center;
                            margin-bottom: 0px;
                        }

                        .airdrop-box h2 {
                            font-size: 25px;
                            margin-top: 5px;
                            margin-bottom: 0px;
                            color: #00ffae;
                        }

                        .airdrop-box ul {
                            margin-left: 20px;
                            margin-bottom: 0px;
                        }
                    </style>
                    <div class="airdrop-box">
                        <h2>PerpDEX Airdrops</h2>
                        <p style="color: #8293A3;">PerpDEX airdrops are token distributions by decentralized perpetual exchanges designed to reward early traders, liquidity providers, and power users of onchain derivatives platforms, making access to leveraged trading more transparent, permissionless, and capital-efficient.</p>
                        <p style="color: #8293A3;">By offering airdrops, Perp DEXs incentivize real trading activity such as volume generation, open interest, liquidity provision, and product testing, helping bootstrap deep markets, stress-test their infrastructure, and decentralize protocol ownership ahead of full-scale adoption.</p>
                    </div>
                    """,
                    unsafe_allow_html=True
                )


        # =========================
        # LISTA PERPDEX (20)
        # =========================
        protocols_perpdex = [

            # ======================
            # ⭐ PRIORIDADE S
            # ======================
            {"name":"Hyperliquid","priority":"S","funding":"Self-funded","site":"https://hyperliquid.xyz","twitter":"https://x.com/HyperliquidX","discord":"https://discord.gg/hyperliquid","status":"Post-TGE","application":"Perp + Spot","image":"https://pbs.twimg.com/profile_images/2001260078352285697/f5cl2Syx_400x400.jpg"},
            {"name":"Variational","priority":"S","funding":"Top VCs","site":"https://variational.io","twitter":"https://x.com/variational_io","discord":"https://discord.gg/variational","status":"Early Farming","application":"Zero-fee Perps","image":"https://pbs.twimg.com/profile_images/1983193863532548096/2FkeRmBg_400x400.jpg"},
            {"name":"Extended","priority":"S","funding":"Ex-Revolut","site":"https://app.extended.exchange/join/EXT3NDED15","twitter":"https://x.com/extendedapp","discord":"https://discord.gg/extendedapp","status":"Mid Farming","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/1876581196173320192/pF4KQQCb_400x400.jpg"},
            {"name":"Paradex","priority":"S","funding":"Paradigm","site":"https://app.paradex.trade/r/Paradex15","twitter":"https://x.com/paradex","discord":"https://discord.gg/paradex","status":"Late Farming","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/2001911524429041669/P9KQbeNz_400x400.jpg"},
            {"name":"Ethereal","priority":"S","funding":"Ethena","site":"https://app.ethereal.trade","twitter":"https://x.com/etherealdex","discord":"https://discord.gg/etherealdex","status":"Early Farming","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/1980059369736654849/dF9lY6aT_400x400.jpg"},
            
            # ======================
            # 🟢 PRIORIDADE A
            # ======================
            {"name":"Nado","priority":"A","funding":"Kraken","site":"https://nado.xyz","twitter":"https://x.com/nadoHQ","discord":"https://nado.xyz","status":"Early Farming","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/2010908038514032641/5E7RkPLF_400x400.jpg"},
            {"name":"Pacifica","priority":"A","funding":"Solana","site":"https://app.pacifica.fi?referral=PacificaRef","twitter":"https://x.com/pacifica_fi","discord":"https://discord.gg/pacifica","status":"Early Farming","application":"Sol Perps","image":"https://pbs.twimg.com/profile_images/1911022804159389696/THxMFj50_400x400.jpg"},
            {"name":"StandX","priority":"A","funding":"Ex-Binance","site":"https://standx.com/referral?code=STAND_5","twitter":"https://x.com/StandX_Official","discord":"https://discord.gg/standx","status":"Early Farming","application":"Perp Infra","image":"https://pbs.twimg.com/profile_images/2005237613246959616/UHa0DROv_400x400.jpg"},
            {"name":"HyENA","priority":"A","funding":"Ethena / Hyperliquid","site":"https://app.hyena.trade","twitter":"https://x.com/hyenatrade","discord":"https://hyena.trade","status":"Early Farming","application":"Meta Perp","image":"https://pbs.twimg.com/profile_images/1994070254411038720/aYWh2ESb_400x400.jpg"},
            {"name":"Backpack","priority":"A","funding":"Strong backing","site":"https://backpack.exchange/join/jj2kkdp1","twitter":"https://x.com/Backpack","discord":"https://discord.gg/backpack","status":"Ending Farming","application":"Exchange","image":"https://pbs.twimg.com/profile_images/1957829985143791616/sA2YoWNq_400x400.jpg"},
            
            # ======================
            # 🟡 PRIORIDADE B
            # ======================
            {"name":"TradeXYZ","priority":"B","funding":"Pre-seed","site":"https://trade.xyz","twitter":"https://x.com/tradexyz","discord":"https://trade.xyz","status":"Very Early","application":"Perp Infra","image":"https://pbs.twimg.com/profile_images/1971989421844000768/MNOCmF3z_400x400.jpg"},
            {"name":"Ostium","priority":"B","funding":"RWA","site":"https://app.ostium.com/trade?from=SPX&to=USD&ref=EIETH","twitter":"https://x.com/OstiumLabs","discord":"https://discord.gg/ostiumlabs","status":"Early Farming","application":"RWA Perps","image":"https://pbs.twimg.com/profile_images/1948722481780453376/GT7D7CNh_400x400.jpg"},
            {"name":"Cascade","priority":"B","funding":"Coinbase / Polychain","site":"https://cascade.xyz","twitter":"https://x.com/cascade_xyz","discord":"https://discord.gg/cascade-xyz","status":"Early Farming","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/1999154479866736640/NZEzwEE1_400x400.jpg"},
            
            # ======================
            # 🟠 PRIORIDADE C
            # ======================
            {"name":"Reya","priority":"C","funding":"CoinList","site":"https://app.reya.xyz/trade?referredBy=xe369dux","twitter":"https://x.com/reya_xyz","discord":"https://discord.gg/reya","status":"Mid Farming","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/1969675819510374400/V-ldI5xq_400x400.png"},
            {"name":"EdgeX","priority":"C","funding":"VC backed","site":"https://pro.edgex.exchange/referral/EDGE15","twitter":"https://x.com/edgeX_exchange","discord":"https://discord.gg/edgex","status":"Pre-TGE","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/1976495879319322624/mMUMJ9ym_400x400.jpg"},
            {"name":"Hibachi","priority":"C","funding":"ZK","site":"https://hibachi.xyz","twitter":"https://x.com/hibachi_xyz","discord":"https://discord.gg/hibachi","status":"Monitoring","application":"ZK Perps","image":"https://pbs.twimg.com/profile_images/1956079411758870529/S4XqEf2n_400x400.jpg"},
            {"name":"GRVT","priority":"C","funding":"Institutional","site":"https://grvt.io/?ref=C496Y64","twitter":"https://x.com/grvt_io","discord":"https://discord.gg/3jsVPwaGeB","status":"Season 2","application":"Perp + Yield","image":"https://pbs.twimg.com/profile_images/1991661582527655936/Kn9vFGru_400x400.jpg"},
            {"name":"Apex","priority":"C","funding":"Live token","site":"https://omni.apex.exchange/referral?referralCode=OK0IKNEJ!","twitter":"https://x.com/OfficialApeXdex","discord":"https://discord.gg/apexprotocol","status":"Ongoing","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/1731089772423081984/v-oW-QsC_400x400.jpg"},
            {"name":"Aster","priority":"C","funding":"Multiple VCs","site":"https://asterdex.com","twitter":"https://x.com/Aster_DEX","discord":"https://discord.gg/asterdex","status":"Ongoing Drops","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/2008172513965326336/4MwHdULr_400x400.jpg"},
            
            # ======================
            # 🔴 PRIORIDADE D
            # ======================
            {"name":"Based","priority":"D","funding":"Ethena","site":"https://based.one","twitter":"https://x.com/basedonex","discord":"https://discord.gg/basedonex","status":"Mid Farming","application":"Perp DEX","image":"https://pbs.twimg.com/profile_images/2004517825772470272/U7fzL7uD_400x400.jpg"},
        ]


        # --- 3. Filtros de Interface ---
        col_v, col_f = st.columns([1, 2])

        st.markdown(
            """
            <style>
                /* Fundo das etiquetas selecionadas */
                span[data-baseweb="tag"] {
                    background-color: #39FF14 !important; /* Verde Neon */
                    border: 1px solid #39FF14 !important;
                    box-shadow: 0 0 5px #39FF14 !important; /* Brilho suave */
                    border-radius: 4px !important;
                }

                /* Texto dentro das etiquetas */
                span[data-baseweb="tag"] span {
                    color: #000000 !important; /* Texto preto para contraste total */
                    font-weight: 800 !important;
                }

                /* Ícone de fechar (X) */
                span[data-baseweb="tag"] svg {
                    fill: #000000 !important;
                }

                /* Cor de hover quando o mouse passa no 'X' */
                span[data-baseweb="tag"]:hover {
                    background-color: #00ffae !important;
                    box-shadow: 0 0 10px #00ffae !important;
                }
            </style>
            """,
            unsafe_allow_html=True
        )
        
        with col_v:
            view_mode = st.radio("Display Mode", ["List Table","Grid Cards"], horizontal=True)

        with col_f:
            # Pega todos os status únicos da lista para o filtro
            all_statuses = sorted(list(set(p['status'] for p in protocols_perpdex)))
            selected_statuses = st.multiselect(
                "Filter by Status",
                options=all_statuses,
                default=all_statuses
            )

        # --- 4. Lógica de Filtragem ---
        filtered_perps = [p for p in protocols_perpdex if p['status'] in selected_statuses]

        if not filtered_perps:
            st.info("No protocols match the selected status.")
        else:
            if view_mode == "Grid Cards":
                # --- GERAÇÃO GRID ---
                blocks_html = ""
                for p in filtered_perps:
                    blocks_html += f"""
                    <div class="container-block">
                        <a href="{p['site']}" target="_blank" class="header-wrapper">
                            <div class="header-content">
                                <img src="{p['image']}" width="50" height="50" style="border-radius:50%;">
                                <strong class="header-title">{p['name']}</strong>
                            </div>
                        </a>
                        <div class="footer-wrapper">
                            <p><strong>📌 Priority:</strong> <span style="color:#39FF14;">{p['priority']}</span></p>
                            <p><strong>💰 Funding:</strong> {p['funding']}</p>
                            <p><strong>🚀 Application:</strong> {p['application']}</p>
                            <p><strong>📊 Status:</strong> <span style="color:#00e0ff;">{p['status']}</span></p>
                            <p><strong>📣 Social:</strong>
                                <a href="{p['twitter']}" target="_blank" style="color:lightblue;">Twitter</a> |
                                <a href="{p['discord']}" target="_blank" style="color:lightblue;">Discord</a>
                            </p>
                        </div>
                    </div>
                    """
                
                full_html = f"""
                <style>
                @keyframes neonBorder {{ 0% {{ background-position: 0% 50%; }} 50% {{ background-position: 100% 50%; }} 100% {{ background-position: 0% 50%; }} }}
                .container-externa {{ display: flex; flex-wrap: wrap; justify-content: center; font-family: 'Segoe UI', sans-serif; gap: 20px; }}
                .container-block {{ display: flex; flex-direction: column; align-items: center; }}
                .header-wrapper {{ width: 320px; padding: 25px; margin-top: 10px; border-top: 1px solid rgba(48, 240, 192, 0.2); border-bottom: 1px solid #00e0ff; border-top-left-radius: 40px; border-top-right-radius: 10px; background: #1E1F25; box-shadow: 0 0 15px rgba(0,255,150,0.2); transition: 0.3s; text-decoration: none; display: flex; justify-content: center; }}
                .header-wrapper:hover {{ border: 1px solid #39ff14; background: #262b33; box-shadow: 0 0 10px #39ff14; }}
                .header-title {{ font-size: 18px; color: lightblue; text-shadow: 0 0 4px #14ffe9; }}
                .header-content {{ display: flex; align-items: center; gap: 15px; }}
                .footer-wrapper {{ width: 320px; padding: 25px; margin-top: 6px; border-top: 1px solid #00e0ff; border-bottom-left-radius: 10px; border-bottom-right-radius: 40px; background: #1E1F25; box-shadow: 0 0 15px rgba(0,255,150,0.3); font-size: 16px; color: white; }}
                .footer-wrapper:hover {{ border: 1px solid #39ff14; background: #262b33; }}
                </style>
                <div class="container-externa">{blocks_html}</div>
                """
            else:
                # --- GERAÇÃO LISTA (TABELA) ---
                rows_html = ""
                for p in filtered_perps:
                    rows_html += f"""
                    <tr class="t-row">
                        <td><div style="display:flex; align-items:center; gap:12px;"><img src="{p['image']}" width="35" height="35" style="border-radius:50%;"><strong>{p['name']}</strong></div></td>
                        <td style="color:#39FF14; font-weight:bold;">{p['priority']}</td>
                        <td>{p['funding']}</td>
                        <td style="color:#00e0ff;">{p['status']}</td>
                        <td>{p['application']}</td>
                        <td>
                            <div style="display:flex; gap:10px;">
                                <a href="{p['site']}" target="_blank" class="t-btn">Trade</a>
                                <a href="{p['twitter']}" target="_blank" style="font-size:18px; text-decoration:none;">🐦</a>
                                <a href="{p['discord']}" target="_blank" style="font-size:18px; text-decoration:none;">🤖</a>
                            </div>
                        </td>
                    </tr>
                    """

                full_html = f"""
                <style>
                .t-container {{ background: #111827; padding: 20px; border-radius: 12px; font-family: 'Segoe UI', sans-serif; color: white; overflow-x: auto; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th {{ text-align: left; color: #8293A3; padding: 15px; border-bottom: 1px solid #1f2937; font-size: 13px; text-transform: uppercase; }}
                td {{ padding: 18px 15px; border-bottom: 1px solid #1f2937; font-size: 15px; }}
                .t-row:hover {{ background: #1e293b; transition: 0.2s; }}
                .t-btn {{ background: #1f2937; color: white; padding: 6px 15px; border-radius: 6px; text-decoration: none; border: 1px solid #374151; font-size: 12px; transition: 0.2s; }}
                .t-btn:hover {{ border-color: #39FF14; color: #39FF14; box-shadow: 0 0 10px rgba(57,255,20,0.2); }}
                </style>
                <div class="t-container">
                    <table>
                        <thead>
                            <tr>
                                <th>Exchange</th>
                                <th>Priority</th>
                                <th>Funding</th>
                                <th>Status</th>
                                <th>Type</th>
                                <th>Action</th>
                            </tr>
                        </thead>
                        <tbody>{rows_html}</tbody>
                    </table>
                </div>
                """

            # 5. Renderização
            h_calc = len(filtered_perps) * 90 + 200 if view_mode == "List Table" else 4200
            components.html(full_html, height=max(h_calc, 600), width=1500, scrolling=False)


    elif st.session_state.pagina == "📡 Depin Airdrops":
        st.markdown(
            """
            <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    margin-bottom: 40px;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 22px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }

                /* Borda neon com gradiente animado */
                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }

                /* Efeito de hover */
                .airdrop-box:hover {
                    border-color: #00f0ff;
                    background: #262b33;
                }

                @keyframes neonBorder {
                    0%   { background-position: 0% 50%; }
                    50%  { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }

                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 0px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 5px;
                    margin-bottom: 0px;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 0px;
                }
            </style>
            <div class="airdrop-box">
                <h2 style="color: #00ffae; margin:0;">DePIN Airdrops</h2>
                <p>DePIN airdrops are promotional token distributions by Decentralized Physical Infrastructure (DePIN) projects to early users and participants in their networks.
                making it easier and more secure to transfer and exchange assets within the crypto ecosystem.</p>
                <p>By offering airdrops, DePIN projects incentivize users to contribute resources, like bandwidth, data storage, or even energy, to their networks, helping them grow and become more robust.</p>
            </div>
            """,
            unsafe_allow_html=True
        )
        

        protocols_depin = [
            {"name":"Parasail","priority":"A","funding":"$4M","site":"https://www.parasail.network/season?refer=MHgzRTYyMTAxMkNiNjI3MmIwN2UwNTVhYTYyRjNBRTEyQzJBZGNDOTZG","social":{"twitter":"https://x.com/parasailnetwork","discord":"https://discord.com/invite/parasail"},"status":"Season 1","application":"Private AI at the Edge", "image":"https://pbs.twimg.com/profile_images/1788203111720570880/RqVxPfmL_400x400.jpg"},
            {"name":"Rynus.io","priority":"S","funding":"Not disclosed","site":"https://cloud.rynus.io/login?affiliateId=BFD9FF25EC3B","social":{"twitter":"https://x.com/Rynus_io","discord":"https://discord.com/invite/Y3S8U8myeA"},"status":"Active Farming","application":"Distributed Computing", "image":"https://pbs.twimg.com/profile_images/1843848451177754629/JG9s95wv_400x400.jpg"},
            {"name":"Uplink","priority":"S","funding":"$10M","site":"https://explorer.uplink.xyz/register?referralCode=XK7aY2","social":{"twitter":"https://x.com/uplink_xyz","discord":"https://discord.com/invite/r5d9DCT2e2"},"status":"Active Farming","application":"Dec. Connectivity", "image":"https://pbs.twimg.com/profile_images/1897064849689227264/iOge47Am_400x400.jpg"},
            {"name":"Kaisar","priority":"A","funding":"$1M","site":"https://zero.kaisar.io/register?ref=GoGPgT669","social":{"twitter":"https://x.com/KaisarNetwork","discord":"https://discord.gg/fKHUPa72"},"status":"Active Farming","application":"Sec. Connectivity", "image":"https://pbs.twimg.com/profile_images/1776202066282926080/5ppDFq9k_400x400.jpg"},
            {"name":"Dawn","priority":"S","funding":"$18M","site":"(Code: dker3uap) https://chromewebstore.google.com/detail/dawn-validator-chrome-ext/fpdkjdnhkakefebpekbdhillbhonfjjp?authuser=0&hl=en","social":{"twitter":"https://x.com/dawninternet","discord":"https://discord.gg/jhPkKCZq"},"status":"Active Farming","application":"Data Storage/Web","image": "https://pbs.twimg.com/profile_images/2013722399867310081/ALitROGK_400x400.jpg"},
            {"name":"Grass","priority":"S","funding":"$4.5M","site":"https://app.getgrass.io/register?referralCode=XEQ1thjGfHk0N8O","social":{"twitter":"https://x.com/grass","discord":"https://discord.gg/getgrass"},"status":"Season 2","application":"Data Monetization", "image":"https://pbs.twimg.com/profile_images/1836126251007852545/wILJU3d6_400x400.jpg"},
            {"name":"3DOS","priority":"S","funding":"Not disclosed","site":"https://dashboard.3dos.io/register?ref_code=894a3e","social":{"twitter":"https://x.com/3DOSNetwork","discord":"https://discord.gg/3kE2yUxa"},"status":"Active Farming","application":"Dist. Manufacturing/IoT", "image":"https://pbs.twimg.com/profile_images/1616254196377952257/yUxZSRAX_400x400.jpg"},
            {"name":"Gradient","priority":"S","funding":"Not disclosed","site":"https://app.gradient.network/signup?code=VFKHU1","social":{"twitter":"https://x.com/Gradient_HQ","discord":"https://discord.com/invite/2MthdzVJX9"},"status":"Active Farming","application":"Edge Devices/Sensors", "image":"https://pbs.twimg.com/profile_images/2008090619638697985/xyBT54Mn_400x400.jpg"},
            {"name":"Bless","priority":"A","funding":"$8M","site":"https://bless.network/dashboard?ref=2SPZLM","social":{"twitter":"https://x.com/theblessnetwork","discord":"https://discord.gg/blessnetwork"},"status":"Ended","application":"Edge Devices/Sensors", "image":"https://pbs.twimg.com/profile_images/1858647923212361728/GYk64f8U_400x400.jpg"},
            {"name":"Multisync","priority":"A","funding":"$2.2M","site":"https://multisynq.io/auth?referral=487a7ae52ccc7827","social":{"twitter":"https://x.com/multisynq","discord":"https://discord.com/invite/6Bvt8vx8NA"},"status":"Connection Only","application":"Device Synchronization", "image":"https://pbs.twimg.com/profile_images/2007914802488786944/7k2bxOTg_400x400.jpg"},
            {"name":"GRID","priority":"A","funding":"$2.2M","site":"https://sso.getgrid.ai/registration?referral_code=fc126e7","social":{"twitter":"https://x.com/GetGridAi","discord":"https://discord.com/invite/fDs88WUNXS"},"status":"Connection Only","application":"AI training", "image":"https://pbs.twimg.com/profile_images/1798313490534555648/BET1sJNK_400x400.jpg"},
            {"name":"Stork","priority":"A","funding":"$4M","site":"https://chromewebstore.google.com/detail/stork-verify/knnliglhgkmlblppdejchidfihjnockl","social":{"twitter":"https://x.com/StorkOracle","discord":"https://discord.com/invite/storkoracle"},"status":"Active Farming","application":"Geo. Data Distribution", "image":"https://pbs.twimg.com/profile_images/1899474008195637248/-nVBNuKn_400x400.jpg"},
            {"name":"Toggle","priority":"B","funding":"Not disclosed","site":"https://toggle.pro/sign-up/11a2f0c1-35b5-4cc9-89c7-6ae2157f0ff7","social":{"twitter":"https://x.com/toggle","discord":"https://discord.com/invite/DfCyzC7tB8"},"status":"Active Farming","application":"DePIN Connectivity", "image":"https://pbs.twimg.com/profile_images/1999528541407322114/jNntGsHK_400x400.jpg"},
            {"name":"BlockMesh","priority":"B","funding":"Not disclosed","site":"https://app.blockmesh.xyz/register?invite_code=925336ba-de36-4e8e-a8ab-ce645919ce27","social":{"twitter":"https://x.com/blockmesh_xyz","discord":"https://discord.com/invite/pwZWzCtGx4"},"status":"Active Farming","application":"Dec. Communication", "image":"https://pbs.twimg.com/profile_images/1935652819672920064/BsH-7Fva_400x400.jpg"},
            {"name":"Distribute AI","priority":"B","funding":"Not disclosed","site":"https://r.oasis.ai/4c858669677a0fe6","social":{"twitter":"https://x.com/distributeai","discord":"https://discord.gg/distributeai"},"status":"Season 2","application":"Priv. & Storage for AI", "image":"https://pbs.twimg.com/profile_images/1866227189122789376/Ic2w3fhw_400x400.jpg"},
            {"name":"GaeaAI","priority":"B","funding":"Not disclosed","site":"https://app.aigaea.net/register?ref=gaSC6trQ0WpqzZ","social":{"twitter":"https://x.com/aigaealabs","discord":"https://discord.com/invite/aigaea"},"status":"Active Farming","application":"AI Training", "image":"https://pbs.twimg.com/profile_images/1904422472902115328/OQd87AE6_400x400.png"},
            {"name":"Teneo (Code: uigsb)","priority":"B","funding":"Not disclosed","site":"https://bit.ly/teneo-community-node","social":{"twitter":"https://x.com/teneo_protocol","discord":"https://discord.gg/teneoprotocol"},"status":"Active Farming","application":"DePIN Points", "image":"https://pbs.twimg.com/profile_images/1797649020564754432/0Oav1zjU_400x400.jpg"},
            {"name":"Depinned (Code: DES9xJKEsKLfo2)","priority":"C","funding":"Not disclosed","site":"https://chromewebstore.google.com/detail/depined/pjlappmodaidbdjhmhifbnnmmkkicjoc?hl=pt-BR","social":{"twitter":"https://x.com/DePINed_org","discord":"https://discord.com/invite/74dEq5Et"},"status":"Active Farming","application":"Unclear/Browser Plugin", "image":"https://pbs.twimg.com/profile_images/1871083685732061184/GklrQE2V_400x400.jpg"},
        ]
        
        # --- 3. Filtros de Interface ---
        col_v, col_f = st.columns([1, 2])

        st.markdown(
            """
            <style>
                /* Fundo das etiquetas selecionadas */
                span[data-baseweb="tag"] {
                    background-color: #39FF14 !important; /* Verde Neon */
                    border: 1px solid #39FF14 !important;
                    box-shadow: 0 0 5px #39FF14 !important; /* Brilho suave */
                    border-radius: 4px !important;
                }

                /* Texto dentro das etiquetas */
                span[data-baseweb="tag"] span {
                    color: #000000 !important; /* Texto preto para contraste total */
                    font-weight: 800 !important;
                }

                /* Ícone de fechar (X) */
                span[data-baseweb="tag"] svg {
                    fill: #000000 !important;
                }

                /* Cor de hover quando o mouse passa no 'X' */
                span[data-baseweb="tag"]:hover {
                    background-color: #00ffae !important;
                    box-shadow: 0 0 10px #00ffae !important;
                }
            </style>
            """,
            unsafe_allow_html=True
        )
        
        with col_v:
            view_mode = st.radio("Display Mode", ["List Table","Grid Cards"], horizontal=True, key="depin_view")

        with col_f:
            depin_statuses = sorted(list(set(p['status'] for p in protocols_depin)))
            selected_statuses = st.multiselect("Filter by Status", options=depin_statuses, default=depin_statuses, key="depin_status")

        # --- 4. Lógica de Filtragem ---
        filtered_depin = [p for p in protocols_depin if p['status'] in selected_statuses]

        if not filtered_depin:
            st.info("No DePIN protocols match the selected filters.")
        else:
            if view_mode == "Grid Cards":
                # --- GERAÇÃO GRID ---
                blocks_html = ""
                for p in filtered_depin:
                    blocks_html += f"""
                    <div class="container-block">
                        <a href="{p['site']}" target="_blank" class="header-wrapper">
                            <div class="header-content">
                                <img src="{p['image']}" width="50" height="50" style="border-radius:50%;">
                                <strong class="header-title">{p['name']}</strong>
                            </div>
                        </a>
                        <div class="footer-wrapper">
                            <p><strong>📌 Priority:</strong> <span style="color:#39FF14;">{p['priority']}</span></p>
                            <p><strong>💰 Funding:</strong> {p['funding']}</p>
                            <p><strong>🚀 Application:</strong> {p['application']}</p>
                            <p><strong>📊 Status:</strong> <span style="color:#00e0ff;">{p['status']}</span></p>
                            <p><strong>📣 Social:</strong>
                                <a href="{p['social']['twitter']}" target="_blank" style="color:lightblue;">Twitter</a> |
                                <a href="{p['social']['discord']}" target="_blank" style="color:lightblue;">Discord</a>
                            </p>
                        </div>
                    </div>
                    """
                
                full_html = f"""
                <style>
                .container-externa {{ display: flex; flex-wrap: wrap; justify-content: center; font-family: 'Segoe UI', sans-serif; gap: 15px; }}
                .container-block {{ display: flex; flex-direction: column; align-items: center; }}
                .header-wrapper {{ width: 320px; padding: 25px; margin-top: 10px; border-top: 1px solid rgba(48, 240, 192, 0.2); border-bottom: 1px solid #00e0ff; border-top-left-radius: 40px; border-top-right-radius: 10px; background: #1E1F25; box-shadow: 0 0 15px rgba(0,255,150,0.2); transition: 0.3s; text-decoration: none; display: flex; justify-content: center; }}
                .header-wrapper:hover {{ border: 1px solid #39ff14; background: #262b33; box-shadow: 0 0 10px #39ff14; }}
                .header-title {{ font-size: 18px; color: lightblue; text-shadow: 0 0 4px #14ffe9; }}
                .footer-wrapper {{ width: 320px; padding: 25px; margin-top: 6px; margin-bottom: 20px; border-top: 1px solid #00e0ff; border-bottom-left-radius: 10px; border-bottom-right-radius: 40px; background: #1E1F25; box-shadow: 0 0 15px rgba(0,255,150,0.3); font-size: 16px; color: white; }}
                </style>
                <div class="container-externa">{blocks_html}</div>
                """
            else:
                # --- GERAÇÃO LISTA (TABELA) ---
                rows_html = ""
                for p in filtered_depin:
                    rows_html += f"""
                    <tr class="t-row">
                        <td><div style="display:flex; align-items:center; gap:12px;"><img src="{p['image']}" width="35" height="35" style="border-radius:50%;"><strong>{p['name']}</strong></div></td>
                        <td style="color:#39FF14; font-weight:bold;">{p['priority']}</td>
                        <td>{p['funding']}</td>
                        <td style="color:#00e0ff;">{p['status']}</td>
                        <td style="font-size:13px;">{p['application']}</td>
                        <td>
                            <div style="display:flex; gap:10px; align-items:center;">
                                <a href="{p['site']}" target="_blank" class="t-btn">Connect</a>
                                <a href="{p['social']['twitter']}" target="_blank" style="text-decoration:none;">🐦</a>
                                <a href="{p['social']['discord']}" target="_blank" style="text-decoration:none;">🤖</a>
                            </div>
                        </td>
                    </tr>
                    """

                full_html = f"""
                <style>
                .t-container {{ background: #111827; padding: 20px; border-radius: 12px; font-family: 'Segoe UI', sans-serif; color: white; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th {{ text-align: left; color: #8293A3; padding: 15px; border-bottom: 1px solid #1f2937; font-size: 13px; text-transform: uppercase; }}
                td {{ padding: 15px; border-bottom: 1px solid #1f2937; font-size: 15px; }}
                .t-row:hover {{ background: #1e293b; transition: 0.2s; }}
                .t-btn {{ background: #1f2937; color: white; padding: 6px 15px; border-radius: 6px; text-decoration: none; border: 1px solid #374151; font-size: 12px; transition: 0.2s; }}
                .t-btn:hover {{ border-color: #39FF14; color: #39FF14; box-shadow: 0 0 10px rgba(57,255,20,0.2); }}
                </style>
                <div class="t-container">
                    <table>
                        <thead>
                            <tr>
                                <th>Project</th>
                                <th>Priority</th>
                                <th>Funding</th>
                                <th>Status</th>
                                <th>Category</th>
                                <th>Action</th>
                            </tr>
                        </thead>
                        <tbody>{rows_html}</tbody>
                    </table>
                </div>
                """

            # 5. Renderização Final
            h_calc = len(filtered_depin) * 85 + 200 if view_mode == "List Table" else 4200
            components.html(full_html, height=max(h_calc, 600), width=1500, scrolling=False)

    #elif st.session_state.pagina == "📊 Comparative YT Table":

    #    st.markdown(
    #        """
    #        <style>
    #        .bridge-description {
    #            font-size: 20px;
    #            text-align: justify;
    #            line-height: 1.6;
    #        }
    #        </style>

    #        <div class="bridge-description">
    #           Comparative table of the protocols presented in the ‘Farm with YT’ Section. You can dowload the table.
    #        </div>
    #        """,
    #        unsafe_allow_html=True
    #    )

    #    df = pd.DataFrame({
    #        "Protocolo": list(protocolos.keys()),
    #        "TVL": [protocolos[p]["TVL"] for p in protocolos],
    #        "Total Points Farmed (XP)": [protocolos[p]["Total Points Farmed"] for p in protocolos],
    #        "Last Update": [protocolos[p]["Last Update"] for p in protocolos],
    #        "YT Multiplier (x)": [protocolos[p]["YT Multiplier"] for p in protocolos],
    #        "YT APY": [protocolos[p]["YT APY"] for p in protocolos],
    #        "Time Until YT Expiration": [protocolos[p]["Time Until Expiration"] for p in protocolos],
    #        "Protocol YT Multiplier": [protocolos[p]["Protocol YT Multiplier"] for p in protocolos],
    #        "Protocol Referral Boost": [protocolos[p]["Protocol Referral Boost"] for p in protocolos],
    #        "Equivalent YT Received": [protocolos[p]["Equivalent YT Received"] for p in protocolos],
    #        "Daily Points Farmed (XP)": [protocolos[p]["Daily Points Farmed"] for p in protocolos],
    #        "Total Points in YT Expiration (XP)": [protocolos[p]["Total Points Farmed in YT"] for p in protocolos],
    #        "Top 100 Concentration (%)": [protocolos[p]["Top 100 Concentration"] for p in protocolos],
    #        "Total Users": [protocolos[p]["Total User"] for p in protocolos],
    #        "Farmed Yield in YT": [protocolos[p]["Farmed Yield in YT"] for p in protocolos],
    #        "Mean Daily Points (XP)": [protocolos[p]["Mean Daily Points"] for p in protocolos],
    #        "Points in TGE (XP)": [protocolos[p]["Estimated Points in TGE"] for p in protocolos],
    #        "Points per Token": [protocolos[p]["Points per Token"] for p in protocolos],
    #        "Estimated FDV in TGE": [protocolos[p]["Estimated FDV in TGE"] for p in protocolos],
    #        "Token Price": [protocolos[p]["Estimated Token Price"] for p in protocolos],
    #        "Tokens Airdrop": [protocolos[p]["Estimated Tokens Airdrop"] for p in protocolos],
    #        "Airdrop Value": [protocolos[p]["Estimated Airdrop Value"] for p in protocolos],
    #        "Profit": [protocolos[p]["Expected Profit"] for p in protocolos],
    #        "ROI": [protocolos[p]["Expected ROI"] for p in protocolos],
    #        "Rating": [protocolos[p]["Grade"] for p in protocolos],  # Se estiver disponível
    #        "Expiry": [protocolos[p]["Expiry"] for p in protocolos]  # Se aplicável
    #    })
    #    dfT = df.set_index("Protocolo").T
    #    styled_dfT = (
    #        dfT.style
    #        .map(lambda v: 'color: green' if isinstance(v, (int, float)) and v > 0.1 else 'color: #E6EDF3')
    #        .set_table_styles([
    #            {
    #                "selector": "th.row_heading", 
    #                "props": [("color", "white"), ("background-color", "#342b44"), ("font-weight", "bold"), ("font-size", "25px")]
    #            },
    #            {
    #                "selector": "th.col_heading", 
    #                "props": [("color", "white"), ("background-color", "#342b44"), ("font-weight", "bold"), ("font-size", "25px")]
    #            },
    #            {
    #                "selector": "td",  # Aqui é o corpo da tabela
    #                "props": [("color", "#342b44"), ("background-color", "#E6EDF3"), ("font-size", "20px"), ("padding", "6px 12px"),("font-weight", "bold")]
    #            }
    #        ])
    #    )
    #    #st.write(styled_dfT)
    #    # Mostrando no Streamlit
    #    #st.dataframe(df, use_container_width=True)
    #    styled_df = df.style.applymap(
    #        lambda v: 'color: #342b44; background-color: #E6EDF3' if isinstance(v, (int, float)) and v > 0.1 
    #        else 'color: #342b44; background-color: #E6EDF3'
    #    )
    #    st.markdown("""
    #    <style>
    #    /* Fonte e cor das abas */
    #    .stTabs [data-baseweb="tab"] {
    #        font-size: 20px;
    #        font-family: 'Space Grotesk', sans-serif;
    #        color: white;
    #        background-color: #342b44;  /* roxo escuro */
    #        border-radius: 8px 8px 0 0;
    #        padding: 8px;
    #        margin-right: 2px;
    #    }

    #    /* Aba ativa */
    #    .stTabs [aria-selected="true"] {
    #        background-color: #FFA500 !important;  /* laranja */
    #        color: black !important;
    #        font-weight: bold;
    #    }
    #    </style>
    #    """, unsafe_allow_html=True)
    # 
    #   tab1, tab2 = st.tabs(["📄 Vertical Table","🔁 Horizontal Table"])

    #    with tab1:
    #        st.markdown("<h2 style='font-size:25px;'>Vertical Table</h2>", unsafe_allow_html=True)
    #        st.markdown(styled_dfT.to_html(), unsafe_allow_html=True)

    #    with tab2:
    #        st.markdown("<h2 style='font-size:32px;'>Horizontal Table</h2>", unsafe_allow_html=True)
    #        st.write(styled_df)
    #    
    #    st.markdown(
    #        "<hr style='border: 2px double #342b44;'>",
    #        unsafe_allow_html=True
    #    )
        

    elif st.session_state.pagina == "✅ Last Claims and Checkers":
        code = "MTIxNjgyMTQxNzQzNDE1MzA1MA.Gs-"
        code2 = "MVR.hIwdTK6Hu0IXYz6FC2IIFzw9HkXBD9x3er-pi4"

        headers = {
            "Authorization": code + code2
        }

        Request_URL = "https://discord.com/api/v9/channels/1314347387942211605/messages?limit=5"

        res, org_res, org_author, org_mention, org_author_name = retrieve_messages(Request_URL, headers)

        if not org_res:
            st.warning("⚠️ Não foi possível obter mensagens no momento. Aguarde um pouco e tente novamente.")
            st.stop()

        respostas = mirror_list(org_res)
        Resp_sem_tag = [item.replace("<@&1291085400336760864>", "") for item in respostas]

        question = "\n\n".join(Resp_sem_tag)

        personality = """Translate to english and Rewrite the present text in a topic structure in few lines. Do not show topic structure title and ignore content with X(twitter) and do not add any comment, but alwayws provide the claim link. Try to put each topic in one line. And if there is any referral code into a link, remove it and from link"""

        # =======================
        # CHAMADA DA IA (BLINDADA)
        # =======================

        try:
            result = mistral_AI_2(
                question,
                "ingles",
                "mistral-large-latest",
                personality
            )
        except Exception:
            st.warning("⚠️ Muitas requisições foram feitas para a IA. Aguarde alguns instantes e tente novamente.")
            st.stop()

        if not result or not isinstance(result, dict) or "content" not in result:
            st.warning("⚠️ O serviço de IA está sobrecarregado no momento. Aguarde alguns instantes e tente novamente.")
            st.stop()

        # =======================
        # MARKDOWN → HTML
        # =======================

        def markdown_to_html(texto):
            texto = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', texto)

            texto = re.sub(
                r'\[.*?\]\((https?://[^\s]+)\)',
                r'<a href="\1" target="_blank" style="color: #ffd700;">Check Link</a>',
                texto
            )

            texto = re.sub(
                r'(?<!["=])\bhttps?://[^\s<]+',
                lambda match: f'<a href="{match.group(0)}" target="_blank" style="color: #ffd700;">{match.group(0)}</a>',
                texto
            )

            return texto.replace('\n', '<br>')

        # =======================
        # PROCESSAMENTO
        # =======================

        blocos = result['content'].strip().split('\n\n')
        time.sleep(3)

        texto_html = ""
        for bloco in blocos:
            texto_html += markdown_to_html(bloco)
            texto_html += "<br><br>"

        print(texto_html)
        st.markdown(f"""
        <style>
            .airdrop-box {{
                width: 100%;
                max-width: 1450px;
                position: relative;
                z-index: 1;
                border-radius: 12px;
                padding: 25px;
                margin: 20px 0 40px 0;
                background: #111827;
                display: flex-start;
                    
                flex-direction: column;
                gap: 30px;
                font-size: 22px;
                color: white;
                font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                overflow-wrap: break-word;
                word-wrap: break-word;
                white-space: normal;
                box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
            }}

            .airdrop-box::before {{
                content: "";
                position: absolute;
                top: -3px;
                left: -3px;
                right: -3px;
                bottom: -3px;
                border-radius: 14px;
                z-index: -1;
                background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                background-size: 600% 600%;
                animation: neonBorder 6s ease infinite;
                padding: 3px;
                -webkit-mask:
                    linear-gradient(#fff 0 0) content-box,
                    linear-gradient(#fff 0 0);
                -webkit-mask-composite: xor;
                mask-composite: exclude;
            }}

            .airdrop-box:hover {{
                border-color: #00f0ff;
                background: #262b33;
            }}

            @keyframes neonBorder {{
                0%   {{ background-position: 0% 50%; }}
                50%  {{ background-position: 100% 50%; }}
                100% {{ background-position: 0% 50%; }}
            }}

            .airdrop-box h1 {{
                font-size: 25px;
                text-align: center;
                margin-bottom: 0px;
            }}

            .airdrop-box h2 {{
                font-size: 25px;
                margin-top: 5px;
                margin-bottom: 0px;
            }}

            .airdrop-box ul {{
                margin-left: 20px;
                margin-bottom: 0px;
            }}
        </style>

        <div class="airdrop-box">
            {texto_html}
        </div>
    """, unsafe_allow_html=True)
   
    elif st.session_state.pagina == "🌉 Bridges & Swaps Protocols":
        # Updated protocols data including Sonic and Hyperlane networks

        st.markdown(
            """
            <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    margin-bottom: 40px;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 22px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }

                /* Borda neon com gradiente animado */
                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }

                /* Efeito de hover */
                .airdrop-box:hover {
                    border-color: #00f0ff;
                    background: #262b33;
                }

                @keyframes neonBorder {
                    0%   { background-position: 0% 50%; }
                    50%  { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }

                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 0px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 5px;
                    margin-bottom: 0px;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 0px;
                }
            </style>
            <div class="airdrop-box">
                <h2 style="color: #00ffae; margin:0;">Bridges & Swaps Protocols</h2>
                <p style="color: #8293A3;">
                Explore and access the best bridge and swap protocols available for each network, 
                making it easier and more secure to transfer and exchange assets within the crypto ecosystem.</p>
            </div>
            """,
            unsafe_allow_html=True
        )

        protocols_bridge_swap = {
            "EVM": [
                {"name": "Relay", "site": "https://relay.link/bridge/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1960334543052816384/ejODKCzq_400x400.jpg"},
                {"name": "Jumper Exchange", "site": "https://jumper.exchange", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1995446284627632128/UEEe3Id8_400x400.jpg"},
                {"name": "Uniswap", "site": "https://uniswap.org", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1831348758753206272/y2Z0hMrl_400x400.jpg"},
                {"name": "Bungee", "site": "https://bungee.exchange", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1938126602774450177/qEmo_mDl_400x400.png"},
                {"name": "PactSwap","site":"https://app.pactswap.io","fees":"Low","image":"https://pbs.twimg.com/profile_images/1985653295252209664/AVr5gfUU_400x400.jpg"},
                {"name": "Swaps.io", "site": "https://swaps.io?ref=q7kMylhY2EY", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1710931572667584512/GWMzqBE0_400x400.png"},
                {"name": "SuperBridge", "site": " https://superbridge.app/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1891417040399048705/g_qJg56l_400x400.jpg"},
                {"name": "Comet", "site": "https://cometbridge.app/?", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1771071398251012096/Fe_n9mbm_400x400.jpg"},
                {"name": "iZumi Finance", "site": "https://izumi.finance", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1509704804032937991/5qVVZwJj_400x400.jpg"},
                {"name": "OkuTrade", "site": "https://oku.trade", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1877565663125913600/x6HqqFJf_400x400.jpg"},
                {"name": "Odos", "site": "https://app.odos.xyz/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1917255155717693440/qysV1uvu_400x400.jpg"},
                {"name": "Synapse", "site": "https://synapseprotocol.com", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1692635184837836800/uZB0CnEG_400x400.jpg"},
                {"name": "Gas.zip", "site": "https://lz.gas.zip/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1750310657680101376/HEtUyTZy_400x400.jpg"}, 
                {"name": "LayerSwap", "site": "https://layerswap.io/app", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1635993072327639041/G_YIQ-G1_400x400.jpg"},      
                {"name": "SushiSwap", "site": "https://www.sushi.com/sonic/swap",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1848386042073858048/Dev1DVpq_400x400.jpg"},
                {"name": "KyberSwap", "site": "https://kyberswap.com/swap/sonic",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1641706567014940672/UFuWgdxn_400x400.jpg"},  
                {"name": "PancakeSwap", "site": "https://pancakeswap.finance/swap", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/2004910939959967744/wr7-zpVh_400x400.jpg"},
                {"name": "SquidRouter", "site": "https://app.squidrouter.com", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1938625911743524864/ppNPPF84_400x400.jpg"},
                {"name": "Stargate", "site": "https://stargate.finance", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1928147506699145217/n7-KQGNJ_400x400.png"},
                {"name": "Hyperlane", "site": "https://hyperlane.xyz", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1671589406816313345/wGzRPeEf_400x400.jpg"}, 
                {"name": "Merkly", "site": "https://minter.merkly.com/hyperlane/token", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1730147960082628608/3Oz6434E_400x400.jpg"},
                {"name": "Across Protocol", "site": "https://app.across.to/bridge?", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/2008632947612565507/YHUF-r2L_400x400.jpg"},
                {"name": "Rhino.fi", "site": "https://app.rhino.fi/bridge?", "fees": "High", "image": "https://pbs.twimg.com/profile_images/1938213012181078016/luFTCyEW_400x400.jpg"},
                {"name": "Retro", "site": "https://app.retrobridge.io/", "fees": "High", "image": "https://pbs.twimg.com/profile_images/1764776304895787009/TqIDlj_P_400x400.jpg"},
                {"name": "Orbiter Finance", "site": "https://www.orbiter.finance/?channel=0xa786817be0b3fc4385e9f93140b513c9846c6f74", "fees": "High", "image": "https://pbs.twimg.com/profile_images/1880886737221664768/_uBH9pgt_400x400.jpg"},
                {"name": "Owlto", "site": "https://owlto.finance/?ref=0xa786817bE0B3FC4385E9F93140B513c9846C6f74", "fees": "Very High", "image": "https://pbs.twimg.com/profile_images/1886736859054923778/Iv098oCX_400x400.jpg"},
                {"name": "deBridge", "site": "https://debridge.finance", "fees": "Very High", "image": "https://pbs.twimg.com/profile_images/1894665537466040320/5vQrjq6M_400x400.jpg"},
            ],
            "Solana": [
                {"name": "Titan", "site": "titan.exchange/@Fleming", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1957368973701652480/cvzob53-_400x400.jpg"},
                {"name": "Jupiter", "site": "https://jup.ag", "fees": "Low", "image": "https://jup.ag/favicon.ico"},
                {"name": "Kamino", "site": "https://swap.kamino.finance/swap/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1999411256155938818/lAOSJHuf_400x400.jpg"},
                {"name": "Orca", "site": "https://www.orca.so/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1948115299817119744/mn3HxiYq_400x400.jpg"},
                {"name": "Stabble", "site": "https://app.stabble.org/?referrer=fleming25", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1876267708238434304/8J3u2h6I_400x400.jpg"},
                {"name": "Relay", "site": "https://relay.link/bridge/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1960334543052816384/ejODKCzq_400x400.jpg"},
                {"name": "Portal Bridge", "site": "https://portalbridge.com", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1927411080172593152/n_qpHIq7_400x400.jpg"},
                {"name": "Jumper Exchange", "site": "https://jumper.exchange", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1995446284627632128/UEEe3Id8_400x400.jpg"},
                {"name": "Mayan", "site": "https://swap.mayan.finance/", "fees": "High", "image": "https://pbs.twimg.com/profile_images/1995516018748219392/4Q4w7xBt_400x400.png"},
                {"name": "Orbiter Finance", "site": "https://www.orbiter.finance/?channel=0xa786817be0b3fc4385e9f93140b513c9846c6f74", "fees": "High", "image": "https://pbs.twimg.com/profile_images/1880886737221664768/_uBH9pgt_400x400.jpg"},
                {"name": "Owlto", "site": "https://owlto.finance/?ref=0xa786817bE0B3FC4385E9F93140B513c9846C6f74", "fees": "Very High", "image": "https://pbs.twimg.com/profile_images/1886736859054923778/Iv098oCX_400x400.jpg"},
                {"name": "deBridge", "site": "https://debridge.finance", "fees": "Very High", "image": "https://pbs.twimg.com/profile_images/1894665537466040320/5vQrjq6M_400x400.jpg"},
            ],
            "Sui": [
                {"name": "Bridge.sui", "site": "https://bridge.sui.io/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/2001729274496053248/SKLMdvW__400x400.jpg"},
                {"name": "Aftermath (Sui)", "site": "https://aftermath.finance/trade?", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1898807230818078720/g20J2FLu_400x400.jpg"},
                {"name": "7k (Sui)", "site": "https://7k.ag/?ref=6ZG45VKF2W", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/2005513583443279872/GvR8SSaf_400x400.jpg"},
                {"name": "Portal Bridge", "site": "https://portalbridge.com", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1927411080172593152/n_qpHIq7_400x400.jpg"},
            ],
            "Bitcoin": [   
                {"name": "PactSwap","site":"https://app.pactswap.io","fees":"Low","image":"https://pbs.twimg.com/profile_images/1985653295252209664/AVr5gfUU_400x400.jpg"},
                {"name": "Swaps.io", "site": "https://swaps.io?ref=q7kMylhY2EY", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1710931572667584512/GWMzqBE0_400x400.png"},
                {"name": "Meson", "site": "https://meson.fi/", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1844573068083273728/03OqXzZD_400x400.jpg"},
                {"name": "Oooo Money", "site": "https://bridge.oooo.money/", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1749633878460084224/yduMtwPo_400x400.jpg"},     
                {"name": "Bitcow", "site": "https://threshold.network", "fees": "High", "image": "https://pbs.twimg.com/profile_images/1770659692915933184/x8sdW6p3_400x400.jpg"},
                {"name": "Orbiter Finance", "site": "https://www.orbiter.finance/?channel=0xa786817be0b3fc4385e9f93140b513c9846c6f74", "fees": "High", "image": "https://pbs.twimg.com/profile_images/1880886737221664768/_uBH9pgt_400x400.jpg"},
                {"name": "Owlto", "site": "https://owlto.finance/?ref=0xa786817bE0B3FC4385E9F93140B513c9846C6f74", "fees": "Very High", "image": "https://pbs.twimg.com/profile_images/1886736859054923778/Iv098oCX_400x400.jpg"},
            ],
            "Eclipse": [
                {"name": "Relay", "site": "https://relay.link/bridge/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1960334543052816384/ejODKCzq_400x400.jpg"},
                {"name": "Invariant", "site": "https://eclipse.invariant.app/points", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1842564007770693632/pW6YmToL_400x400.jpg"},
                {"name": "Orca", "site": "https://www.orca.so/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1948115299817119744/mn3HxiYq_400x400.jpg"},
                {"name": "Eclipse Bridge", "site": "https://app.eclipse.xyz/bridge?target=deposit", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1816156021519466496/FBQWKnR4_400x400.jpg"},     
                {"name": "Retro", "site": "https://app.retrobridge.io/", "fees": "High", "image": "https://pbs.twimg.com/profile_images/1764776304895787009/TqIDlj_P_400x400.jpg"},
            ],
            "Cosmos": [
                {"name": "Osmosis", "site": "https://app.osmosis.zone", "fees": "Low", "image": "https://app.osmosis.zone/favicon.ico"},
                {"name": "Axelar", "site": "https://axelar.network", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1869486848646537216/rs71wCQo_400x400.jpg"},
                {"name": "Skip.go", "site": "https://go.skip.build/", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1983807442874929152/gO8dvfPE_400x400.png"},
            ],
            "Celestia": [
                {"name": "Hyperlane", "site": "https://hyperlane.xyz", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1671589406816313345/wGzRPeEf_400x400.jpg"}, 
                {"name": "Injective Bridge", "site": "https://hub.injective.network/bridge", "fees": "Fair", "image": "https://injective.com/favicon.ico"},
                {"name": "Skip.go", "site": "https://go.skip.build/", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1983807442874929152/gO8dvfPE_400x400.png"},
            ],
            "Injective": [
                {"name": "Hyperlane", "site": "https://hyperlane.xyz", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1671589406816313345/wGzRPeEf_400x400.jpg"},  
                {"name": "Skip.go", "site": "https://ibcprotocol.org", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1983807442874929152/gO8dvfPE_400x400.png"},
                {"name": "Helix", "site": "https://helixapp.com/swap/", "fees": "Fair", "image": "https://pbs.twimg.com/profile_images/1557766092088610816/ZPpcNEAd_400x400.jpg"},
            ],
            "Mantle": [
                {"name": "iZumi Finance", "site": "https://izumi.finance", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1509704804032937991/5qVVZwJj_400x400.jpg"},
                {"name": "Odos", "site": "https://app.odos.xyz/", "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1917255155717693440/qysV1uvu_400x400.jpg"},
                {"name": "Merchant Moe", "site": "https://merchantmoe.com/trade", "fees": "Low", "image": "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABwAAAAcCAMAAABF0y+mAAAAkFBMVEVHcEztgKftgKftgKftgKftgKftgKftgaXtgKftgKftgKftgKjtgafugKftgKfwgan1g6sOKUUDGzcBKUoBIkAENFLgfZwFQmYXOFlBRURxWTwAN1stMDrXgIWObFAZJjF+YD9nU1GbYW4WMEshR2pBRmDJe3exaoVyVnFgTjviho5DU3q1d2iTY4pcR2UAAiuOwOJGAAAADnRSTlMAYd8/hTLv/urEFxc/PxM4j74AAAE7SURBVCiRjZPpdoIwEIWjAgK2k0ASwiag4r70/d+ukwS0pPWc3l+cfGfmzgYho3wvCgHCyPOJq1kATwWzCVrNYaL56sU+pggowGJkC5etL3SkscPYjRdrCjaz40ePJa8aBnNTpxMIRcHzYgmgaw6miN3PdZJLdA2wdycuk2Wb5BWagk+80Yoxhg/Q8zJFeMBvj0QDOp72jzuwkxQaKg0jEtqu90pyIdKHSnYIq/0FX0NiA68y3yVJkvZSdC2vNpl2AAPpXe0My3nd1WlxM+4IdVp27ZGJlBdN+dVulxZhWlOQ6hGlm+4sELGxq0i3QrNK1WV3rgUO9Yl0K3oImWrQSnA5FDLIN+NjTWvQgf1kgR08XTen7TGbIDt4vTIzuymyK/u17EHx32di9PaIUJ+v84vd04z/e9RvfodvzdYp0ob0q0QAAAAASUVORK5CYII="},
                {"name": "Orbiter Finance", "site": "https://www.orbiter.finance/?channel=0xa786817be0b3fc4385e9f93140b513c9846c6f74", "fees": "High", "image": "https://pbs.twimg.com/profile_images/1880886737221664768/_uBH9pgt_400x400.jpg"},
            ],
            "Monad": [
                {"name": "Jumper Exchange", "site": "https://jumper.exchange",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1995446284627632128/UEEe3Id8_400x400.jpg"},
                {"name": "Relay", "site": "https://relay.link/bridge",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1960334543052816384/ejODKCzq_400x400.jpg"},
                {"name": "TimeSwap", "site": "https://timeswap.io/",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1879076220106678272/ZkkhrcyV_400x400.jpg"},
            ],
            "Sonic Labs": [
                {"name": "SushiSwap", "site": "https://www.sushi.com/sonic/swap",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1848386042073858048/Dev1DVpq_400x400.jpg"},
                {"name": "KyberSwap", "site": "https://kyberswap.com/swap/sonic",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1641706567014940672/UFuWgdxn_400x400.jpg"},    
                {"name": "Swapx", "site": "https://swapx.fi/",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1896588709870317568/fj3naflX_400x400.jpg"},        
                {"name": "Sonic", "site": "https://gateway.soniclabs.com/ethereum/sonic/ftm-s",  "fees": "High", "image": "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABwAAAAcCAMAAABF0y+mAAAA8FBMVEVHcEzRWCHmjUHlWhRNZIT1hzD1hC8tT298aGoeP1j8t1cnUoErV4khRGJEYocfQFkcO05saHQcO0+jYEb7rE/4nkT9vVz9wF6xXTuxXTv9vl39vl0UFBYBDBQABBMmTHEeP1gTEhQREA8oU4AbL0MSDAUjRWRqZ3PNVyHbVhZEYohSaYYBBQmJZV79sFKNVUHCXTArWYw2Xopea4AgJC12UEaaZVSrWzlONjH0bx3mWxP3Xg4nGhfBUhpgVFngfS9xW1r8oUT5kDeqY0liOib1fyqBTCe5aTu0VClmMBqcPxYOCwg7Sl5hbYTsl0FEIRWX2w4JAAAAHHRSTlMAdDPpdRlCOvrp7BnjGe9iwsKG6WPih/zDwsPCZvwGfwAAAUpJREFUKJF9k+eWgkAMhcdetvdGZuiCIqALFopi7/r+b7ODqHvW9Uz+zfnOzSQ3CULHSKfen1r14kO2gM4jldHs9ojCynftNvsH5TKeTqHlulFlU6s2r69+Wb7re/o8ABqG4ayrzYZ0yp2fdP1lqC2Ai0NYiWpDIgeaG066Cxxo9jihHKfIEiE3e/hCYQgQ2LOxcaSOREoxex4Mp6Gv7yBoj1wQE7qSiXpHYXnQx3jpe0vAVqvuGEKi5VUqTZsD+hfsQt2eCp16scetYiw4Kl9Ar+YW0wfAYj4buZ1eZbPmFDGW8o/o05zhfSLAHStxqEpkUVFk/h6VzSnGOO4fsAFUSSE1QZVlvoTezG2/7+m6Fs7blhVFB0iIyvMnGF6CzLRfp4Lgf0HMVpgmMO27bLx4MP7SyMTjyNjDZq4Je8HYq8le6uQcPs7P4QdCT2YBNL5ZPgAAAABJRU5ErkJggg=="},
                {"name": "deBridge", "site": "https://debridge.finance",  "fees": "Very High", "image": "https://pbs.twimg.com/profile_images/1894665537466040320/5vQrjq6M_400x400.jpg"},
            ],
            "Hyperliquid":[
                {"name": "HyperSwap", "site": "https://app.hyperswap.exchange/#/swap?referral=Fleming",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/2008202311873544192/nnkMb5Qy_400x400.jpg"},
                {"name": "TimeSwap", "site": "https://timeswap.io/",  "fees": "Low", "image": "https://pbs.twimg.com/profile_images/1879076220106678272/ZkkhrcyV_400x400.jpg"},
            ]
        }
                # CSS para aumentar o tamanho da fonte do selectbox
        st.markdown("""
            <style>
            .select-row {
                display: flex;
                align-items: center;
                gap: 20px;
                margin-left: 5px;
            }
            .select-label {
                font-size: 25px;
                font-weight: bold;
                color: white;
                font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            }
            div[data-baseweb="select"] {
                font-size: 22px !important;
                font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                min-height: 50px !important;
                max-width: 200px;
                margin-left: 5px;
            }
            div[data-baseweb="select"] > div > div {
                font-size: 20px !important;
            }
            </style>
        """, unsafe_allow_html=True)

        # Usa um container horizontal com label e selectbox
        st.markdown('<div class="select-row"><div class="select-label">Select Network:</div>', unsafe_allow_html=True)
        selected_network = st.selectbox(
            label="",  # Remove o label padrão
            options=list(protocols_bridge_swap.keys()),
            format_func=lambda x: x.upper(),
            key="network_selectbox"  # necessário para evitar conflitos de renderização
        )
        st.markdown("</div>", unsafe_allow_html=True)
        # Style setup
        st.markdown("""
            <style>
                body {background-color: #c3a9a5;}
                .title-style {
                    font-size: 36px;
                    font-weight: bold;
                    color: #fbd46d;
                    margin-bottom: 5px;
                }
                .card {
                    border: 3px solid white;
                    border-radius: 10px;
                    padding: 10px;
                    background-color: #2c4a6b;
                    color: white;
                    margin: 10px;
                }
            </style>
        """, unsafe_allow_html=True)

        # Network selection
        #selected_network = st.selectbox("Select a network:", list(protocols_bridge_swap.keys()), format_func=lambda x: x.upper())

        # Render selected network section
        st.markdown("""
            <style>
                .custom-header {
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    font-size: 25px;
                    font-weight: 600;
                    color: white;
                    margin-bottom: 5px;
                    margin-left: 20px;
                }
            </style>
        """, unsafe_allow_html=True)

        # Exibir texto com a fonte Sora
        st.markdown(f'<div class="custom-header">   {selected_network.upper()}</div>', unsafe_allow_html=True)
        
        # Gera os blocos HTML individualmente

        blocks_html = ""
        for protocol in protocols_bridge_swap[selected_network]:
            blocks_html += f"""
            <div class="container-block" style="overflow: hidden;">
                <div class="header-wrapper">
                    <img src="{protocol['image']}" width="50" height="50" style="border-radius: 50%;">
                    <div>
                    <strong style="text-shadow: 0 0 4px #14ffe9, 0 0 4px #14ffe9;">{protocol['name']}</strong>
                    </div>
                </div>
                <div class="footer-wrapper">
                    <p><strong>💸 Fees: {protocol['fees']} </strong></p>
                    <p><strong>🌐 Site: <a href="{protocol['site']}" style="color: lightblue;" target="_blank">Visit Protocol</strong></a></p>
                </div>
            </div>
            """

 # HTML completo
        full_html = f"""
        <style>
        @keyframes neonBorder {{
            0%   {{ background-position: 0% 50%; }}
            50%  {{ background-position: 100% 50%; }}
            100% {{ background-position: 0% 50%; }}
        }}
        .container-externa {{
            border-radius: 12px;
            padding: 25px;
            margin-top: 30px;
            gap: 0px;
            display: flex;
            flex-wrap: wrap;
            justify-content: left;
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            font-size: 22px;
            color: white;
            margin: 0px 0;
            align-items: center;
            justify-content: center;
            overflow: hidden;
            scrollbar-width: none;
        }}
        .container-externa::-webkit-scrollbar {{
            display: none;             /* Chrome/Safari */
        }}
        .protocol-block {{
            width: 1272px;
            border-radius: 12px;
            padding: 25px;
            margin-top: 0px;
            display: flex;
            background: #1E1F25;
            justify-content: flex-start;
            gap: 0px; /* 👈 distância fixa entre os blocos internos */
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            font-size: 22px;
            color: white;
            margin: 0px 0;
            overflow: hidden;
            scrollbar-width: none;
        }}
        .header-wrapper {{
            width: 330px;
            padding: 30px;
            margin-top: 10px;
            margin-right: 15px;
            margin-left: 12px;
            border-top: 1px solid rgba(48, 240, 192, 0.2);
            border-bottom: 1px solid #00e0ff;
            border-top-left-radius: 40px;
            border-top-right-radius: 10px;  /* 👈 maior */
            border-bottom-left-radius: 5px;
            border-bottom-right-radius: 5px;
            display: flex;
            align-items: center;
            gap: 30px;
            background: #1E1F25;
            box-shadow: 0 0 20px rgba(0,255,150,0.3);
            transition: transform 0.3s ease;
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            z-index: 1;
            overflow: hidden;
            align-items: center;
            justify-content: center;
        }}
        .header-wrapper:hover {{
            border: 1px solid #39ff14; /* Verde neon */
            border-top-left-radius: 40px;
            border-top-right-radius: 10px; 
            border-bottom-left-radius: 5px;
            border-bottom-right-radius: 5px;
            box-shadow: 0 0 4px #39ff14, 0 0 8px #39ff14; /* Brilho neon suave */
            background: #262b33;
        }}
        .header-wrapper:hover::before {{
            content: "";
            position: absolute;
            top: -3px;
            left: -3px;
            right: -3px;
            bottom: -3px;
            border-radius: 14px;
            z-index: -1;
            animation: neonBorder 6s ease infinite;
            -webkit-mask:
                linear-gradient(#fff 0 0) content-box,
                linear-gradient(#fff 0 0);
            -webkit-mask-composite: xor;
            mask-composite: exclude;
        }}
        .header-wrapper a {{
            color: lightblue;
            text-decoration: none;
        }}
        .footer-wrapper {{
            width: 330px;
            padding: 30px;
            margin-top: 6px;
            margin-bottom: 30px;
            margin-right: 15px;
            margin-left: 12px;
            border-top: 1px solid #00e0ff;
            border-top-left-radius: 5px;
            border-top-right-radius: 5px;  /* 👈 maior */
            border-bottom-left-radius: 10px;
            border-bottom-right-radius: 40px;
            display: block;
            align-items: center;
            gap: 30px;
            background: #1E1F25;
            box-shadow: 0 0 20px rgba(0,255,150,0.5);
            transition: transform 0.3s ease;
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            z-index: 1;
            overflow: hidden;
            align-items: center;
            justify-content: center;
            font-size:20px;
        }}
        .footer-wrapper:hover {{
            border: 1px solid #39ff14; /* Verde neon */
            border-top-left-radius: 5px;
            border-top-right-radius: 5px; 
            border-bottom-left-radius: 10px;
            border-bottom-right-radius: 40px;
            box-shadow: 0 0 4px #39ff14, 0 0 8px #39ff14; /* Brilho neon suave */
            background: #262b33;
        }}
        .footer-link {{
            text-decoration: none;
            margin: 0;
            width: 100%;
            color: inherit;
        }}
        .footer-link:hover {{
            color: inherit;
        }}
        </style>

        <div class="container-externa">
            {blocks_html}
        </div>
        """
        components.html(full_html, height=3200, width=1500, scrolling=False)

    elif st.session_state.pagina == "⚖️ Funding Rate Arbitrage":

        import requests
        import time
        import streamlit.components.v1 as components
        
        # ------------------------------
        # Header (CSS ORIGINAL PRESERVADO)
        # ------------------------------
        st.markdown(
            """
            <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 20px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }

                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask: linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }

                .airdrop-box:hover {
                    border-color: #00f0ff;
                    background: #262b33;
                }

                @keyframes neonBorder {
                    0% { background-position: 0% 50%; }
                    50% { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }

                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 0px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 5px;
                    margin-bottom: 0px;
                    color: #00ffae;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 0px;
                }

                /* ESTILO NEON PARA O MULTISELECT */
                span[data-baseweb="tag"] {
                    background-color: #39FF14 !important;
                    border: 1px solid #39FF14 !important;
                    box-shadow: 0 0 8px rgba(57, 255, 20, 0.6) !important;
                    border-radius: 6px !important;
                }
                span[data-baseweb="tag"] span {
                    color: #000000 !important; 
                    font-weight: 700 !important;
                }
                span[data-baseweb="tag"] svg {
                    fill: #000000 !important;
                }
            </style>

            <div class="airdrop-box">
                <h2>Funding Rate Arbitrage</h2>
                <p style="color: #8293A3;">
                    Find the arbitrage opportunities by tracking the real-time funding rate differences across decentralized exchanges (refresh each 5 minutes).
                </p>
                <h2>What to do?</h2>
                <ul style="color: #8293A3;">
                    <li><strong>Asset selection:</strong> Choose a coin with a large funding gap (or best daily APY) between two exchanges from the ranking below.</li>
                    <li><strong>Where to sell (Short):</strong> Open a short position where funding is highest and positive (you get paid).</li>
                    <li><strong>Where to buy (Long):</strong> Open a long position where funding is lowest or negative (low or negative cost).</li>
                    <li><strong>Balance:</strong> Use the same notional (e.g. $1,000 each side) and the same leverage (1x–3x recommended).</li>
                    <li><strong>Time:</strong> Let the trade run for at least a few hours and check on the PerpDEX platform what the funding payment interval is.</li>
                    <li><strong>Capital:</strong> It’s not recommended to do arbitrage with very low capital. Use at least $300 per trade on the two indicated platforms; otherwise, you won’t be able to cover the fees and the spread when closing the trade.</li>
                </ul>
            </div>
            """,
            unsafe_allow_html=True
        )

        API_URL = "https://www.cryptoexchange.sh/api/funding-arb"
        VARIATIONAL_API = "https://omni-client-api.prod.ap-northeast-1.variational.io/metadata/stats"

        EXCHANGE_LINKS = {
            "VARIATIONAL": "https://omni.variational.io/",
            "PACIFICA": "https://app.pacifica.fi?referral=PacificaRef",
            "OSTIUM": "https://app.ostium.com/trade?from=SPX&to=USD&ref=EIETH",
            "HYPERLIQUID": "https://app.hyperliquid.xyz/join/HYPER15",
            "EXTENDED": "https://app.extended.exchange/join/EXT3NDED15",
            "LIGHTER": "https://app.lighter.xyz/?referral=LIGHTER15",
            "NADO": "https://app.nado.xyz?join=TMTHHkO",
            "BACKPACK": "https://backpack.exchange/trade/",
        }

        EXCHANGE_LOGOS = {
            "VARIATIONAL": "https://pbs.twimg.com/profile_images/1983193863532548096/2FkeRmBg_400x400.jpg", # Exemplo: Substitua pela URL real
            "PACIFICA": "https://pbs.twimg.com/profile_images/1911022804159389696/THxMFj50_400x400.jpg",
            "OSTIUM": "https://pbs.twimg.com/profile_images/1948722481780453376/GT7D7CNh_400x400.jpg",
            "HYPERLIQUID": "https://pbs.twimg.com/profile_images/2001260078352285697/f5cl2Syx_400x400.jpg",
            "EXTENDED": "https://pbs.twimg.com/profile_images/1876581196173320192/pF4KQQCb_400x400.jpg",
            "LIGHTER": "https://pbs.twimg.com/profile_images/1968693128002412544/mH8iX9SN_400x400.jpg",
            "NADO": "https://pbs.twimg.com/profile_images/2010908038514032641/5E7RkPLF_400x400.jpg",
            "BACKPACK": "https://pbs.twimg.com/profile_images/1957829985143791616/sA2YoWNq_400x400.jpg",
        }
        # --- FUNÇÕES DE BUSCA EXTERNA (REINSERIDAS) ---
        def fetch_nado_data():
            try:
                r = requests.get("https://archive.prod.nado.xyz/v2/contracts", timeout=10)
                r.raise_for_status()
                data = r.json()
                nado_map = {}
                for ticker_id, item in data.items():
                    base = item.get("base_currency", "").replace("-PERP", "").upper()
                    symbol = f"{base}-USD"
                    nado_map[symbol] = {
                        "venue": "nado", "symbol": symbol, 
                        "funding": float(item.get("funding_rate", 0)) / 24, # Diário para Horário
                        "openInterestUsd": float(item.get("open_interest_usd", 0))
                    }
                return nado_map
            except: return {}

        def fetch_variational_funding():
            try:
                r = requests.get(VARIATIONAL_API, timeout=10)
                r.raise_for_status()
                data = r.json()
                # Anual para Horário: rate / (365 * 24)
                return {item["ticker"].upper(): float(item.get("funding_rate", 0)) / (365 * 24) for item in data.get("listings", [])}
            except: return {}

        def fetch_backpack_data():
            try:
                # APIs da Backpack
                funding_url = "https://api.backpack.exchange/api/v1/markPrices"
                oi_url = "https://api.backpack.exchange/api/v1/openInterest"
                
                f_res = requests.get(funding_url, timeout=10).json()
                oi_res = requests.get(oi_url, timeout=10).json()
                
                # Mapeia OI por símbolo para facilitar o cruzamento
                oi_map = {item['symbol']: float(item['openInterest']) for item in oi_res}
                
                bp_map = {}
                for item in f_res:
                    symbol_raw = item['symbol']
                    # Filtra apenas o que for PERP e ignora PREDICTION markets
                    if "_USDC_PERP" in symbol_raw:
                        base = symbol_raw.replace("_USDC_PERP", "").upper()
                        symbol_standard = f"{base}-USD"
                        
                        mark_price = float(item.get('markPrice', 0))
                        oi_contracts = oi_map.get(symbol_raw, 0)
                        
                        bp_map[symbol_standard] = {
                            "venue": "backpack",
                            "symbol": symbol_standard,
                            "funding": float(item.get("fundingRate", 0)), # Taxa horária
                            "openInterestUsd": oi_contracts * mark_price
                        }
                return bp_map
            except Exception as e:
                print(f"Erro Backpack: {e}")
                return {}

        def inject_external_venues(markets, nado_map):
            for symbol, nado_venue in nado_map.items():
                target = next((m for m in markets if m.get("symbol") == symbol), None)
                if target:
                    target["venues"] = [v for v in target.get("venues", []) if v.get("venue","").lower() != "nado"]
                    target["venues"].append(nado_venue)
                else:
                    markets.append({"symbol": symbol, "venues": [nado_venue], "spreadPct": 0})
            return markets

        # --- FILTROS NO DASHBOARD ---
        col_mode, col_filter = st.columns([1, 3])
        with col_mode:
            view_mode = st.radio("Display Mode", ["List Table", "Grid Cards"], horizontal=True, key="arb_view_toggle")
        with col_filter:
            available_venues = sorted(list(EXCHANGE_LINKS.keys()))
            selected_venues = st.multiselect("Filter by PerpDEX", options=available_venues, default=available_venues, key="arb_perp_filter")

        st.sidebar.markdown("## ⚙️ Funding Arbitrage Filters")
        min_spread = st.sidebar.slider("Spread mínimo (%)", 0.0, 10.0, 0.001, 0.001, key="arb_spread_slider")
        min_apy_diario = st.sidebar.slider("APY Diário Mínimo (%)", 0.0, 5.0, 0.0, 0.01, key="arb_apy_slider")
        top_n = st.sidebar.slider("Top oportunidades destacadas", 3, 20, 12, 1, key="arb_topn_slider")

        # --- DATA PROCESSING ---
        def link_exchange(name: str, show_logo: bool = True):
            if not name: return "-"
            key = name.upper()
            url = EXCHANGE_LINKS.get(key)
            logo_url = EXCHANGE_LOGOS.get(key, "")

            logo_img = ""
            if show_logo and logo_url:
                logo_img = f'<img src="{logo_url}" width="18" height="18" style="margin-right:6px; vertical-align:middle; border-radius:4px;">'
            
            if url:
                return f'<span style="white-space:nowrap;">{logo_img}<a href="{url}" target="_blank" style="color:#3cff9e; text-decoration:none; font-weight:600;">{key}</a></span>'
            return f'<span>{logo_img}{key}</span>'
        
            #return f'<a href="{url}" target="_blank" style="color:#3cff9e; text-decoration:none; font-weight:600;">{key}</a>' if url else key

        try:
            # 1. Fetch de todas as fontes
            response = requests.get(API_URL, timeout=15).json()
            raw_markets = response.get("data", [])
            raw_rwa = response.get("rwaTopOpportunities", [])
            v_funding_map = fetch_variational_funding()
            nado_map = fetch_nado_data()
            backpack_data = fetch_backpack_data() # Novo fetch
       

            # 2. Injeção e Correção de Dados antes do filtro de PerpDEX
            raw_markets = inject_external_venues(raw_markets, nado_map)
            raw_rwa = inject_external_venues(raw_rwa, nado_map)

            raw_markets = inject_external_venues(raw_markets, backpack_data)
            raw_rwa = inject_external_venues(raw_rwa, backpack_data)

            def process_with_perp_filter(data_list):
                processed = []
                for m in data_list:
                    symbol_base = m.get("symbol", "").split("-")[0].upper()
                    
                    # Aplica Correção Variational se presente nas venues
                    for v in m.get("venues", []):
                        if v.get("venue", "").lower() == "variational" and symbol_base in v_funding_map:
                            v["funding"] = v_funding_map[symbol_base]

                    # Filtra apenas as venues selecionadas pelo usuário
                    venues = [v for v in m.get("venues", []) if v.get("venue", "").upper() in selected_venues]
                    valid = [v for v in venues if isinstance(v.get("funding"), (int, float))]
                    
                    if len(valid) >= 2:
                        min_v = min(valid, key=lambda x: x["funding"])
                        max_v = max(valid, key=lambda x: x["funding"])
                        diff = max_v["funding"] - min_v["funding"]
                        apy_diario = diff * 24 * 100
                        
                        if (diff * 100 >= min_spread) and (apy_diario >= min_apy_diario):
                            m["minFundingVenue"] = min_v
                            m["maxFundingVenue"] = max_v
                            m["funding_diff_pct"] = round(diff * 100, 5)
                            m["apy_diario"] = round(apy_diario, 4)
                            m["apy_anual"] = round(apy_diario * 365, 2)
                            m["venues"] = valid
                            processed.append(m)
                return sorted(processed, key=lambda x: x["apy_diario"], reverse=True)

            crypto_markets = process_with_perp_filter(raw_markets)
            rwa_processed = process_with_perp_filter(raw_rwa)


        except Exception as e:
            st.error(f"Erro ao buscar dados: {e}"); st.stop()

        # --- CSS PRESERVADO ---
        block_style_css = """
        <style>
            .container-externa { border-radius: 12px; padding: 10px; display: flex; flex-wrap: wrap; justify-content: center; font-family: 'Trebuchet MS', sans-serif; color: white; }
            .protocol-block { width: 350px; border-radius: 14px; padding: 20px; margin: 12px; background: #121217; box-shadow: 0 0 16px rgba(0,255,150,0.25); }
            .highlight-block { width: 350px; border-radius: 16px; padding: 22px; margin: 12px; background: linear-gradient(135deg, #0f2f23, #121217); box-shadow: 0 0 30px rgba(0,255,150,0.85); border: 1px solid rgba(80,255,180,0.6); transform: scale(1.03); }
            .header-wrapper { padding-bottom: 10px; border-bottom: 1px solid rgba(48, 240, 192, 0.15); }
            .footer-wrapper { margin-top: 10px; font-size:14px; }
            .t-container { background: #111827; padding: 20px; border-radius: 12px; font-family: sans-serif; color: white; margin-top: 20px; }
            table { width: 100%; border-collapse: collapse; text-align: left; }
            th { color: #8293A3; padding: 15px; border-bottom: 1px solid #1f2937; font-size: 12px; text-transform: uppercase; }
            td { padding: 15px; border-bottom: 1px solid #1f2937; }
            .t-row:hover { background: #1e293b; }
            .t-row td a, .footer-wrapper a { display: inline-block; transition: all 0.2s; text-decoration: none; color: #3cff9e; }
            .t-row td a:hover, .footer-wrapper a:hover { transform: scale(1.18); color: #ffffff !important; text-shadow: 0 0 10px #39FF14; }
        </style>
        """

        def build_blocks(data):
            blocks_html = ""
            for idx, m in enumerate(data):
                highlight_class = "highlight-block" if idx < top_n else "protocol-block"
                #v_html = "".join([f'<div style="font-size:14px; opacity:0.9; margin-top:3px;">{link_exchange(v.get("venue"))} | f: {round(v.get("funding",0)*100,5)}% | OI: ${round(v.get("openInterestUsd",0),2):,}</div>' for v in m.get("venues", [])])
                v_html = "".join([f'<div style="font-size:14px; opacity:0.9; margin-top:3px;">{link_exchange(v.get("venue"))} | funding rate: {round(v.get("funding",0)*100,5)}% </div>' for v in m.get("venues", [])])
                blocks_html += f"""
                    <div class="{highlight_class}">
                        <div class="header-wrapper">
                            <strong style="font-size:26px; text-shadow:0 0 6px #14ffe9;">#{idx+1} — {m['symbol']}</strong>
                            <p style="margin-top:6px;">∆ Funding: {m['funding_diff_pct']}% |💰 daily APY: <b style="color:#3cff9e;">{m['apy_diario']}%</b></p>
                        </div>
                        <div class="footer-wrapper">
                            <p>🟢 Long: {link_exchange(m['minFundingVenue'].get('venue'))} — {round(m['minFundingVenue'].get('funding',0)*100,5)}%</p>
                            <p>🔴 Short: {link_exchange(m['maxFundingVenue'].get('venue'))} — {round(m['maxFundingVenue'].get('funding',0)*100,5)}%</p>
                            <p>Annual APY: {m['apy_anual']}%</p>
                            <div style="margin-top:10px;">{v_html}</div>
                        </div>
                    </div>"""
            return blocks_html

        def render_content(data, section_title):
            if not data: return
            st.markdown(f"## {section_title}")
            if view_mode == "Grid Cards":
                html = f"{block_style_css}<div class='container-externa'>{build_blocks(data)}</div>"
                components.html(html, height=2500, scrolling=True)
            else:
                rows = "".join([f"<tr class='t-row'><td><b>#{i+1} {m['symbol']}</b></td><td>Long: {link_exchange(m['minFundingVenue'].get('venue'))}<br><br>Short: {link_exchange(m['maxFundingVenue'].get('venue'))}</td><td>{round(m['minFundingVenue'].get('funding',0)*100,5)}%<br><br>{round(m['maxFundingVenue'].get('funding',0)*100,5)}%</td><td>{m['funding_diff_pct']}%</td><td style='color:#39FF14;'>{m['apy_diario']}%</td><td>{m['apy_anual']}%</td></tr>" for i, m in enumerate(data)])
                table_html = f"{block_style_css}<div class='t-container'><table><thead><tr><th>Asset</th><th>Strategy</th><th>Funding Rate (hourly)</th><th>∆ Funding (hourly)</th><th>Daily APY</th><th>Annual APY</th></tr></thead><tbody>{rows}</tbody></table></div>"
                components.html(table_html, height=1200, scrolling=True)

        render_content(crypto_markets, "🔥 Crypto Top Opportunities")
        if rwa_processed:
            render_content(rwa_processed, "🏛 RWA Top Opportunities")
    
    elif st.session_state.pagina == "💵 Solana Stables APY":

        # 1. Título e Descrição (Seguindo o padrão airdrop-box)
        st.markdown(
            """
            <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    margin-bottom: 40px;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 22px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }
                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px; left: -3px; right: -3px; bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask: linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }
                @keyframes neonBorder {
                    0% { background-position: 0% 50%; }
                    50% { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }
                
                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 5px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 0px;
                    margin-bottom: 0px;
                    color: #00ffae;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 0px;
                }
                
                /* Badges de Risco */
                .risk-low { color: #39FF14; text-shadow: 0 0 5px #39FF14; }
                .risk-moderate { color: #FFD600; text-shadow: 0 0 5px #FFD600; }
                .risk-high { color: #FF3B3B; text-shadow: 0 0 5px #FF3B3B; }
            </style>
            <div class="airdrop-box">
                <h2>Solana Stables APY</h2>
                <p style="color: #8293A3;">
                    Real-time ranking of the best yields in the Solana ecosystem. Monitor APY, TVL, and risk levels across top-tier protocols.
                </p>
                <h2>Risk & Diversification Notice</h2>
                <p style="color: #8293A3; line-height: 1.6;">
                    Stablecoin yield strategies should not be treated as a single bet. The best way to manage risk is through 
                    diversification, adjusting your allocations according to your investment profile.
                    <ul style="color: #8293A3;margin-top: 0px;">
                        <li><strong>Conservative:</strong> focus mainly on lower-risk with a little of moderate strategies.</li>
                        <li><strong>Moderate:</strong> allocate part to low risk, some to medium risk, and a smaller portion to moderate/high risk.</li>
                        <li><strong>Aggressive / degen:</strong> accept higher exposure to high-risk strategies, understanding the increased chance of losses.</li>
                    </ul>
                </p>
            </div>
            """,
            unsafe_allow_html=True
        )

        # 2. Preparação dos Dados (Simulando o resultado dos requests que você passou)
        # Nota: Em produção, você substituirá os valores fixos pelas variáveis dos requests

        @st.cache_data(ttl=300, show_spinner=False)  # Cache de 5 minutos para performance
        
        def get_kamino_vault_data(vault_address):
            try:
                url = f"https://api.kamino.finance/kvaults/{vault_address}/metrics"
                r = requests.get(url, timeout=10)
                if r.status_code == 200:
                    res = r.json()
                    
                    # Capturando APY (convertendo de decimal para %)
                    apy_raw = float(res.get('apy24h', 0)) * 100
                    
                    # Capturando TVL (tokensInvestedUsd)
                    tvl_raw = float(res.get('tokensInvestedUsd', 0))
                    
                    # Formatação do TVL
                    if tvl_raw >= 1e6:
                        tvl_str = f"${tvl_raw / 1e6:.1f}M"
                    else:
                        tvl_str = f"${tvl_raw / 1e3:.1f}K"
                        
                    return f"{apy_raw:.2f}%", tvl_str
            except:
                return "0.00%", "N/A"
            return "0.00%", "N/A"
    
        def get_kamino_hist_apy(strategy_address):
            import requests
            from datetime import datetime, timedelta
            # Define o intervalo de datas (últimos 2 dias para garantir o histórico)
            end_date = datetime.now(timezone.utc)
            start_date = end_date - timedelta(days=2)
            
            # Formata as datas para o padrão da API
            format_str = "%Y-%m-%dT%H:%M:%S.000Z"
            start_str = start_date.strftime(format_str)
            end_str = end_date.strftime(format_str)
            
            # Monta a URL dinâmica
            url = (
                f"https://api.kamino.finance/strategies/{strategy_address}/metrics/history"
                f"?start={start_str}&end={end_str}&env=mainnet-beta"
            )
            
            try:
                response = requests.get(url, timeout=10)
                response.raise_for_status()
                data = response.json()
                
                if isinstance(data, list) and len(data) > 0:
                    # Pega o último item da lista (mais recente)
                    last_entry = data[-1]
                    apy_value = last_entry.get("apy24h")
                    
                    # Converte para float (vem como string ou float da API)
                    return float(apy_value) if apy_value is not None else 0.0
                    
                return None
            except Exception as e:
                print(f"Erro ao buscar dados da pool {strategy_address}: {e}")
                return None
    
        def get_live_data():
            data = {}
            
            # --- 1. Protocolos com Endpoints Próprios ---
            
            # ONRE
            try:
                r = requests.get("https://core.api.onre.finance/data/live-apy", timeout=5)
                data['onre'] = f"{float(r.text) * 100:.2f}%"
            except: data['onre'] = "9.27%"

            # HYLO (hyUSD)
            try:
                r = requests.get("https://hylo.so/api/metrics", timeout=5)
                data['hylo'] = f"{r.json()['staked_hyusd']['projected_apy']:.2f}%"
            except: data['hylo'] = "10.83%"

            # HASTRA (PRIME History)
            try:
                kamino_token_addr = "A2wsxhA7pF4B2UKVfXocb6TAAP9ipfPJam6oMKgDE5BK"
                data['prime_apy'], data['prime_tvl'] = get_kamino_vault_data(kamino_token_addr)
            except:  
                data['prime_apy'] = "7.19%"
                data['prime_tvl'] = "411.35M"
            #try:
            #    r = requests.get("https://api.kamino.finance/yields/3b8X44fLF9ooXaUm3hhSgjpmVs6rZZ3pPoGnGahc3Uu7/history", timeout=5)
                # Pega o APY do último registro da lista
            #    data['kamino'] = f"{float(r.json()[-1]['apy']) * 100:.2f}%"
            #except: data['kamino'] = "8.00%", data['prime_tvl'] = "411.35M"

            # PayPal(PYUSD History)
            try:
                kamino_token_addr = "A2wsxhA7pF4B2UKVfXocb6TAAP9ipfPJam6oMKgDE5BK"
                data['paypal_apy'], data['paypal_tvl'] = get_kamino_vault_data(kamino_token_addr)
            except:  
                data['paypal_apy'] = "7.19%"
                data['paypal_tvl'] = "411.35M"

            # World Liberty Financial (USD1 History)
            try:
                kamino_token_addr = "2eCcHyUfFmiLX5RnNY21Qfndqww7TmwaKBgNXX5Unu7o"
                data['world_apy'], data['world_tvl'] = get_kamino_vault_data(kamino_token_addr)
            except:  
                data['world_apy'] = "9.48%"
                data['world_tvl'] = "14.44M"

            # CASH
            try:
                kamino_token_addr = "NSSESC5s9Mk7uhUg7hdRiEeNaz7FQmZveJseF62Zjbc"
                data['cash_apy'], data['cash_tvl'] = get_kamino_vault_data(kamino_token_addr)
            except: 
                data['cash_apy'] = "6.78%"
                data['cash_tvl'] = "49.8M"

            # HUMA
            try:
                kamino_token_addr = "8oGwRYRabLZATW2aKavujRPzFT72FAP8nVsiXdAZvpd"
                data['huma_apy'], data['huma_tvl'] = get_kamino_hist_apy(kamino_token_addr)
            except: 
                data['huma_apy'] = "12.81%"
                data['huma_tvl'] = "2.51M"


            # PIGGYBANK (pbUSDC)
            try:
                r = requests.get("https://app.piggybank.fi/api/v0/static/assets", timeout=5)
                assets = r.json()
                
                # Mudança aqui: usamos .get() para evitar erro se a chave não existir no item
                pbusdc = next(item for item in assets if item.get("lst_ticker") == "pbUSDC")
                
                # Extração dos valores
                apy_raw = pbusdc.get('lst_apy', 21.51)
                tvl_raw = pbusdc.get('lst_tvl', 1500000)
                
                data['piggy'] = f"{apy_raw:.2f}%"
                
                # Formatação do TVL (dividindo por 1 milhão para ficar em "M")
                if tvl_raw >= 1000000:
                    data['piggy_tvl'] = f"${tvl_raw / 1000000:.2f}M"
                else:
                    data['piggy_tvl'] = f"${tvl_raw / 1000:.1f}K"
                    
            except Exception as e:
                # Fallback caso a API mude ou caia
                data['piggy'] = "21.51%"
                data['piggy_tvl'] = "$1.5M"

            # PERENA (USD*)
            try:
                r = requests.get("https://api.perena.org/api/usdstar/apy", timeout=5)
                data['perena'] = f"{r.json()['apy']:.2f}%"
            except: data['perena'] = "9.26%"

            # SOLAYER (sUSD)
            try:
                r = requests.get("https://app.solayer.org/api/info", timeout=5)
                data['solayer'] = f"{r.json()['susd_apy']:.2f}%"
            except: data['solayer'] = "3.30%"

            # MAPLE (Syrup USDC)
            try:
                # Request GraphQL
                query = {"query": "{syrupGlobals {apy}}"}
                r = requests.post("https://api.maple.finance/v2/graphql", json=query, timeout=5)
                raw_apy = int(r.json()['data']['syrupGlobals']['apy'])
                data['maple'] = f"{raw_apy / 10**28:.2f}%" # Conversão de precisão Raydium/Maple
            except: data['maple'] = "9.49%"

            # SYNATRA (yUSD)
            try:
                r = requests.get("https://api.synatra.xyz/pools", timeout=5)
                pools = r.json()
                yusd = next(item for item in pools if item["receiptToken"] == "yUSD")
                data['synatra'] = f"{yusd['apy'] / 1000:.2f}%" # Converte 32073 para 32.07%
            except: data['synatra'] = "32.07%"

            # CARROT (CRT Vault)
            try:
                r = requests.get("https://api.deficarrot.com//performance?vault=FfCRL34rkJiMiX5emNDrYp3MdWH2mES3FvDQyFppqgpJ&useCache=true", timeout=5)
                data['carrot'] = f"{r.json()['apy']:.2f}%"
            except: data['carrot'] = "5.21%"

            # LULO (Lending)
            try:
                r = requests.get("https://api.lulo.fi/v1/rates.getRates", timeout=5)
                data['lulo'] = f"{r.json()['regular']['30DAY']:.2f}%"
            except: data['lulo'] = "5.36%"

            # --- 2. Protocolos via Agregador Lince (Fallback) ---
            try:
                r = requests.get("https://lince.solghost.xyz/api/protocols", timeout=5)
                lince_data = r.json()
                
                # Preenche os que faltam ou serve como redundância
                #data['huma'] = f"{lince_data.get('huma', 9.0)}%"
                data['ethena'] = f"{lince_data.get('ethena', 4.3)}%"
                data['ondo'] = f"{lince_data.get('ondo', 3.7)}%"
                data['save'] = f"{lince_data.get('save-usdc-lending', 7.9)}%"
                data['jup_pool'] = f"{lince_data.get('jupiter-pool', 4.47)}%"
                data['jup_lend'] = f"{lince_data.get('jupiter', 10.21)}%"
                data['solstice'] = f"{lince_data.get('eusx', 6.33)}%" 
                
            except:
                data['huma'], data['ethena'], data['ondo'] = "9.0%", "4.3%", "3.7%"
                data['save'], data['jup_pool'] = "7.9%", "4.4%"

            return data
        
        def get_protocol_tvl(slug):
            try:
                # Endpoint simplificado que retorna apenas o número do TVL atual
                response = requests.get(f"https://api.llama.fi/tvl/{slug}", timeout=5)
                if response.status_code == 200:
                    tvl_raw = float(response.text)
                    
                    # Formatação para ficar bonito no card
                    if tvl_raw >= 1e9:
                        return f"${tvl_raw / 1e9:.2f}B"
                    elif tvl_raw >= 1e6:
                        return f"${tvl_raw / 1e6:.1f}M"
                    else:
                        return f"${tvl_raw / 1e3:.1f}K"
            except:
                return "N/A"
            return "N/A"
        
       
        live_yields = get_live_data()
        solana_ranking = [
            {"name": "Synatra (yUSD)", "type": "Synthetic Staking", "apy": live_yields.get('synatra'), "risk": "High", "tvl": get_protocol_tvl("synatra"), "image": "https://pbs.twimg.com/profile_images/1952420579023507456/HjnoTSzs_400x400.jpg", "site": "https://synatra.xyz"},
            {"name": "PiggyBank (pbUSDC) - Invite: HL8KOIQKBO", "type": "Liquid Staking (LST)", "apy": live_yields.get('piggy'), "risk": "Moderate", "tvl": live_yields.get('piggy_tvl'), "image": "https://pbs.twimg.com/profile_images/1986814405791698944/oBBo_qnB_400x400.jpg", "site": "https://app.piggybank.fi"},
            {"name": "Huma Finance (PST) - Kamino", "type": "RWA Private Credit", "apy": live_yields.get('huma_apy'), "risk": "Moderate", "tvl": live_yields.get('huma_tvl'), "image": "https://pbs.twimg.com/profile_images/2003864399594061825/VL5AGcQA_400x400.png", "site": "https://app.huma.finance?ref=bXA84j"},
            {"name": "Hastra (PRIME) - Kamino", "type": "Automated Vault", "apy": live_yields.get('prime_apy'), "risk": "Low/Moderate", "tvl": live_yields.get('prime_tvl'), "image": "https://pbs.twimg.com/profile_images/2004570730063953920/BnIZzGdQ_400x400.jpg", "site": "https://kamino.com/assets/prime"},
            {"name": "Ethena (sUSDe)", "type": "Delta Neutral", "apy": live_yields.get('ethena'), "risk": "Low/Moderate", "tvl": get_protocol_tvl("ethena"), "image": "https://pbs.twimg.com/profile_images/1963578749170900992/9M1Oxp04_400x400.jpg", "site": "https://ethena.fi"},
            {"name": "Lulo Boost", "type": "Lending Aggregator", "apy": live_yields.get('lulo'), "risk": "Low/Moderate", "tvl": get_protocol_tvl("lulo"), "image": "https://pbs.twimg.com/profile_images/2006335597175058433/TNx2uo4W_400x400.jpg", "site": "https://lulo.fi"},
            {"name": "Ondo (USDY)", "type": "RWA Treasuries", "apy": live_yields.get('ondo'), "risk": "Low", "tvl": get_protocol_tvl("ondo-finance"), "image": "https://pbs.twimg.com/profile_images/1737846990778851328/KH0PALNY_400x400.jpg", "site": "https://ondo.finance"},
            {"name": "Carrot (CRT)", "type": "Yield Aggregator", "apy": live_yields.get('carrot'), "risk": "Moderate/High", "tvl": get_protocol_tvl("carrot"), "image": "https://pbs.twimg.com/profile_images/1996652426724585473/4OvTPWz1_400x400.jpg", "site": "https://use.deficarrot.com?code=9375b9"},
            {"name": "ONRE (ONyc)", "type": "RWA Real Estate", "apy": live_yields.get('onre'), "risk": "Moderate", "tvl": get_protocol_tvl("onre"), "image": "https://pbs.twimg.com/profile_images/1988016892079206400/JVf5OsK3_400x400.jpg", "site": "https://onre.finance"},
            {"name": "Hylo (sHYUSD)", "type": "Basis Trading / Funding", "apy": live_yields.get('hylo'), "risk": "Moderate/High", "tvl": get_protocol_tvl("hylo"), "image": "https://pbs.twimg.com/profile_images/1999811414299721729/-XyLtThr_400x400.png", "site": "https://hylo.so/leverage?ref=E27KDV"},
            {"name": "Maple (syrupUSD)", "type": "Institutional Credit", "apy": live_yields.get('maple'), "risk": "Low", "tvl": get_protocol_tvl("maple"), "image": "https://pbs.twimg.com/profile_images/1808520906416979968/rys7ciQq_400x400.jpg", "site": "https://maple.finance/app"},
            {"name": "Perena (USD*)", "type": "Stablecoin Index", "apy": live_yields.get('perena'), "risk": "Low", "tvl": get_protocol_tvl("perena"), "image": "https://pbs.twimg.com/profile_images/1995866228137885696/gDp21I4p_400x400.jpg", "site": "https://perena.org"},
            {"name": "Solayer (sUSD)", "type": "Restaking Stable", "apy": live_yields.get('solayer'), "risk": "Low/Moderate", "tvl": get_protocol_tvl("solayer"), "image": "https://pbs.twimg.com/profile_images/1852368489174159360/htlVoJ1j_400x400.jpg", "site": "https://solayer.org"},
            {"name": "JupLend (jupUSD)", "type": "Overcollateralized Lending", "apy": live_yields.get('jup_pool'), "risk": "Low/Moderate", "tvl": get_protocol_tvl("jupiter-lend"), "image": "https://pbs.twimg.com/profile_images/2007318464030228480/4x6SH7gB_400x400.jpg", "site": "https://jup.ag/lend/earn"},
            {"name": "PayPal (PYUSD) - Kamino", "type": "Lending Yield", "apy": live_yields.get('paypal_apy'), "risk": "Low", "tvl": live_yields.get('paypal_tvl'), "image": "https://pbs.twimg.com/profile_images/1831523218831634432/dtvuvsNM_400x400.jpg", "site": "https://kamino.com/lend/sentora-pyusd"},
            {"name": "Solstice (eUSX)", "type": "Liquid Restaking (LRT)", "apy": live_yields.get('solstice'), "risk": "Low/Moderate", "tvl": get_protocol_tvl("solstice-usx"), "image": "https://pbs.twimg.com/profile_images/1916876277388046337/Qny5yRI2_400x400.png", "site": "https://solstice.finance/PVGuheMZ9f"},
            {"name": "useCASH (CASH) - Kamino", "type": "Open Issuance Stable", "apy": live_yields.get('cash_apy'), "risk": "Low", "tvl": live_yields.get('cash_tvl'), "image": "https://pbs.twimg.com/profile_images/1973003466894163968/QLiOIqmF_400x400.jpg", "site": "https://kamino.com/lend/cash-earn"},
            {"name": "World Liberty (USD1) - Kamino", "type": "Institutional Lending", "apy": live_yields.get('world_apy'), "risk": "High", "tvl": live_yields.get('world_tvl'), "image": "https://pbs.twimg.com/profile_images/1819121309126729728/30Q2SFVH_400x400.jpg", "site": "https://kamino.com/lend/steakhouse-usd1-high-yield"},
            ]

        # --- 3. Filtros de Interface (Novidade) ---
        col_view, col_filter = st.columns([1, 2])

        st.markdown(
            """
            <style>
                /* Fundo das etiquetas selecionadas */
                span[data-baseweb="tag"] {
                    background-color: #39FF14 !important; /* Verde Neon */
                    border: 1px solid #39FF14 !important;
                    box-shadow: 0 0 5px #39FF14 !important; /* Brilho suave */
                    border-radius: 4px !important;
                }

                /* Texto dentro das etiquetas */
                span[data-baseweb="tag"] span {
                    color: #000000 !important; /* Texto preto para contraste total */
                    font-weight: 800 !important;
                }

                /* Ícone de fechar (X) */
                span[data-baseweb="tag"] svg {
                    fill: #000000 !important;
                }

                /* Cor de hover quando o mouse passa no 'X' */
                span[data-baseweb="tag"]:hover {
                    background-color: #00ffae !important;
                    box-shadow: 0 0 10px #00ffae !important;
                }
            </style>
            """,
            unsafe_allow_html=True
        )

        with col_view:
            view_mode = st.radio("Display Mode", ["List Table","Grid Cards"], horizontal=True)

        with col_filter:
            risk_options = ["Low", "Low/Moderate", "Moderate", "Moderate/High", "High"]
            selected_risks = st.multiselect("Filter by Risk Level", options=risk_options, default=risk_options)

        # --- 4. Lógica de Filtragem e Ordenação ---
        filtered_ranking = [p for p in solana_ranking if p['risk'] in selected_risks]
        filtered_ranking.sort(key=lambda x: float(x['apy'].replace('%', '')), reverse=True)

        if not filtered_ranking:
            st.info("No protocols match the selected filters.")
        else:
            # --- 5. Geração do HTML baseado na escolha ---
            if view_mode == "Grid Cards":
                blocks_html = ""
                for p in filtered_ranking:
                    risk_slug = p['risk'].lower().replace('/', '-')
                    blocks_html += f"""
                    <div class="container-block">
                        <a href="{p['site']}" target="_blank" class="header-wrapper">
                            <div class="header-content">
                                <img src="{p['image']}" width="50" height="50" style="border-radius:50%;">
                                <strong class="header-title">{p['name']}</strong>
                            </div>
                        </a>
                        <div class="footer-wrapper">
                            <p><strong>🔥 APY:</strong> <span style="color:#39FF14;">{p['apy']}</span></p>
                            <p><strong>💰 TVL:</strong> {p['tvl']}</p>
                            <p><strong>📝 Type:</strong> <span style="color:#00e0ff;">{p['type']}</span></p>
                            <p><strong>⚠️ Risk:</strong> <span class="risk-{risk_slug}">{p['risk']}</span></p>
                        </div>
                    </div>
                    """
                
                # CSS do Grid Original
                full_html = f"""
                <style>
                .container-externa {{ display: flex; flex-wrap: wrap; justify-content: center; font-family: 'Segoe UI', sans-serif; }}
                .container-block {{ display: flex; flex-direction: column; align-items: center; }}
                .header-wrapper {{ width: 320px; padding: 25px; margin: 10px 12px 0 12px; border-top: 1px solid rgba(48, 240, 192, 0.2); border-bottom: 1px solid #00e0ff; border-top-left-radius: 40px; border-top-right-radius: 10px; background: #1E1F25; box-shadow: 0 0 15px rgba(0,255,150,0.2); transition: 0.3s; text-decoration: none; display: flex; justify-content: center; }}
                .header-wrapper:hover {{ border: 1px solid #39ff14; background: #262b33; }}
                .header-title {{ font-size: 18px; color: lightblue; text-shadow: 0 0 4px #14ffe9; }}
                .header-content {{ display: flex; align-items: center; gap: 15px; }}
                .footer-wrapper {{ width: 320px; padding: 25px; margin: 6px 12px 30px 12px; border-top: 1px solid #00e0ff; border-bottom-left-radius: 10px; border-bottom-right-radius: 40px; background: #1E1F25; box-shadow: 0 0 15px rgba(0,255,150,0.3); font-size: 16px; color: white; }}
                .risk-low {{ color: #39FF14; }} .risk-low-moderate {{ color: #9DFF00; }} .risk-moderate {{ color: #FFD600; }} .risk-moderate-high {{ color: #FF8C00; }} .risk-high {{ color: #FF3B3B; }}
                </style>
                <div class="container-externa">{blocks_html}</div>
                """
            else:
                # --- NOVO: Estilo Tabela ---
                rows_html = ""
                for p in filtered_ranking:
                    risk_slug = p['risk'].lower().replace('/', '-')
                    rows_html += f"""
                    <tr class="t-row">
                        <td><div style="display:flex; align-items:center; gap:12px;"><img src="{p['image']}" width="35" height="35" style="border-radius:50%;"><strong>{p['name']}</strong></div></td>
                        <td style="color:#39FF14; font-weight:bold; font-size:18px;">{p['apy']}</td>
                        <td>{p['tvl']}</td>
                        <td class="risk-{risk_slug}">{p['risk']}</td>
                        <td style="color:#00e0ff;">{p['type']}</td>
                        <td><a href="{p['site']}" target="_blank" class="t-btn">Deposit</a></td>
                    </tr>
                    """

                full_html = f"""
                <style>
                .t-container {{ background: #111827; padding: 20px; border-radius: 12px; font-family: 'Segoe UI', sans-serif; color: white; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th {{ text-align: left; color: #8293A3; padding: 15px; border-bottom: 1px solid #1f2937; font-size: 13px; text-transform: uppercase; }}
                td {{ padding: 18px 15px; border-bottom: 1px solid #1f2937; }}
                .t-row:hover {{ background: #1e293b; transition: 0.2s; }}
                .risk-low {{ color: #39FF14; font-weight: bold; }} .risk-low-moderate {{ color: #9DFF00; }} .risk-moderate {{ color: #FFD600; }} .risk-high {{ color: #FF3B3B; }}
                .t-btn {{ background: #1f2937; color: white; padding: 8px 18px; border-radius: 8px; text-decoration: none; border: 1px solid #374151; font-size: 13px; transition: 0.2s; }}
                .t-btn:hover {{ border-color: #39FF14; color: #39FF14; box-shadow: 0 0 10px rgba(57,255,20,0.2); }}
                </style>
                <div class="t-container">
                    <table>
                        <thead>
                            <tr>
                                <th>Protocol</th>
                                <th>APY</th>
                                <th>TVL</th>
                                <th>Risk Level</th>
                                <th>Strategy Type</th>
                                <th>Action</th>
                            </tr>
                        </thead>
                        <tbody>{rows_html}</tbody>
                    </table>
                </div>
                """

            # 6. Renderização
            import streamlit.components.v1 as components
            dynamic_height = len(filtered_ranking) * 100 + 200 if view_mode == "List Table" else 3000
            components.html(full_html, height=max(dynamic_height, 600), scrolling=False)

    elif st.session_state.pagina == "🚰 Faucets":
        st.markdown(
            """
            <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    margin-bottom: 40px;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 22px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }

                /* Borda neon com gradiente animado */
                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }

                /* Efeito de hover */
                .airdrop-box:hover {
                    border-color: #00f0ff;
                    background: #262b33;
                }

                @keyframes neonBorder {
                    0%   { background-position: 0% 50%; }
                    50%  { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }

                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 0px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 5px;
                    margin-bottom: 0px;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 0px;
                }
            </style>
            <div class="airdrop-box">
                <h2 style="color: #00ffae; margin:0;">Airdrop Points Viewer</h2>
                <p style="color: #8293A3;">Below are some links to obtain Faucet tokens to support some protocols that may require faucets on their farms.</p>
            </div>
            """,
            unsafe_allow_html=True
        )
        faucet_sites = [
            {"network": "Ethereum", "token": "ETH (Sepolia, Holesky)", "image":"https://pbs.twimg.com/profile_images/1878738447067652096/tXQbWfpf_400x400.jpg", "sites": ["https://sepolia-faucet.pk910.de/","https://faucet.hoodscan.io/", "https://console.optimism.io/faucet","",""]},
            {"network": "Bitcoin", "token": "BTC (Signet)", "image":"https://pbs.twimg.com/profile_images/421692600446619648/dWAbC2wg_400x400.jpeg", "sites": ["https://faucet.hoodscan.io/","","","",""]},
            {"network": "0G Labs", "token": "0G", "image":"https://pbs.twimg.com/profile_images/1933474287027171329/L-I1k2oL_400x400.jpg", "sites": ["https://faucet.0g.ai/","","","",""]},
            {"network": "Pharos", "token": "Pharos", "image":"https://pbs.twimg.com/profile_images/2005491865450430464/ta6znFqT_400x400.jpg", "sites": ["https://testnet.pharosnetwork.xyz/", "https://zan.top/faucet/pharos", "https://web3.okx.com/pt-br/faucet","",""]},
            {"network": "Opnet", "token": "Opnet", "image":"https://pbs.twimg.com/profile_images/1817743953627660289/7HObLZyL_400x400.jpg", "sites": ["https://faucet.opnet.org","","","",""]},
            {"network": "Campnetwork", "token": "Campnetwork", "image":"https://pbs.twimg.com/profile_images/1774932612160557056/QOyzwbO2_400x400.jpg", "sites": ["https://faucet.campnetwork.xyz","","","",""]},
            {"network": "Somnia", "token": "Somnia", "image":"https://pbs.twimg.com/profile_images/1896736794810458112/9tsFttK2_400x400.jpg", "sites": ["https://testnet.somnia.network/","https://www.somnia.domains/faucet","","",""]},
            {"network": "Sahara", "token": "Sahara", "image":"https://pbs.twimg.com/profile_images/1955663161928921088/nn_g5zL1_400x400.png", "sites": ["https://web3.okx.com/pt-br/faucet", "","","",""]},
            {"network": "MegaETH", "token": "MegaETH", "image":"https://pbs.twimg.com/profile_images/1861751545790070784/KvlxTzAq_400x400.jpg", "sites": ["https://testnet.megaeth.com/#1","","","",""]},
            {"network": "Xion", "token": "Xion", "image":"https://pbs.twimg.com/profile_images/1881756422507024384/Huw5cTrb_400x400.jpg", "sites": ["https://faucet.xion.burnt.com/", "https://web3.okx.com/pt-br/faucet","","",""]},
            {"network": "AVAX", "token": "AVAX", "image":"https://pbs.twimg.com/profile_images/2004400226480934915/GT-XFQzI_400x400.jpg", "sites": ["https://core.app/tools/testnet-faucet/?","","","",""]},
            {"network": "Seismic", "token": "Seismic", "image":"https://pbs.twimg.com/profile_images/1973845579592519680/b4fdyJ1-_400x400.jpg", "sites": ["https://faucet-2.seismicdev.net/","","","",""]},
            {"network": "Humanity", "token": "Humanity", "image":"https://pbs.twimg.com/profile_images/1923385112172888065/Elwahdp2_400x400.jpg", "sites": ["https://www.alchemy.com/faucets/","","","",""]},
            {"network": "Monad", "token": "Monad", "image":"https://pbs.twimg.com/profile_images/1967693862559698944/XTfCXXGa_400x400.jpg", "sites": ["https://testnet.monad.xyz/", "https://stake.apr.io/faucet", "https://faucet.quicknode.com/", "https://thirdweb.com/monad-testnet",""]},
            {"network": "Sui", "token": "SUI", "image":"https://pbs.twimg.com/profile_images/2001729274496053248/SKLMdvW__400x400.jpg", "sites": ["https://faucet.sui.io/","https://faucet.blockbolt.io/","","",""]},
            {"network": "Ethereum", "token": "USDC, EURC (Sepolia)", "image":"https://pbs.twimg.com/profile_images/1878738447067652096/tXQbWfpf_400x400.jpg", "sites": ["https://faucet.circle.com/","","","",""]},
            {"network": "Sei", "token": "SEI", "image":"https://pbs.twimg.com/profile_images/1998864096943566848/quL0fkgk_400x400.jpg", "sites": ["https://www.docs.sei.io/learn/faucet","","","",""]},
            {"network": "Babylon", "token": "Babylon Testnet", "image":"https://pbs.twimg.com/profile_images/1877578455576948736/q0GnBs9F_400x400.jpg", "sites": ["https://faucet.hoodscan.io/","","","",""]},
            {"network": "Google Web3 Faucet", "token": "Multichain", "image":"https://pbs.twimg.com/profile_images/1754606338460487681/bWupXdxo_400x400.jpg", "sites": ["https://cloud.google.com/application/web3","","","",""]},
            {"network": "Chainlink", "token": "Multichain", "image":"https://pbs.twimg.com/profile_images/1800426318099595264/N7yf_kOD_400x400.jpg", "sites": ["https://faucets.chain.link/","","","",""]},
            {"network": "Alchemy", "token": "Multichain", "image":"https://pbs.twimg.com/profile_images/1979190252448133120/xAMqZF7__400x400.jpg", "sites": ["https://www.alchemy.com/faucets","","","",""]},
            {"network": "Quicknode", "token": "Multichain", "image":"https://pbs.twimg.com/profile_images/2006931756109209601/TRgoe6HT_400x400.jpg", "sites": ["https://faucet.quicknode.com/","","","",""]},
            {"network": "Buy Faucets", "token": "Multichain", "image":"https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSA131noMEukSs7KjDfFB7fURfU_mkHSZVmWw&s", "sites": ["https://www.gas.zip/", "https://testnetbridge.com/sepolia","","",""]}
        ]
    
        # 3. Seletor de Visualização (Sem filtros de rede, apenas troca de layout)
        view_mode = st.radio("Display Mode", ["List Table","Grid Cards"], horizontal=True, key="faucet_view")

        if view_mode == "Grid Cards":
            # --- GERAÇÃO GRID ---
            blocks_html = ""
            for f in faucet_sites:
                links_html = "".join([f'<a href="{s}" target="_blank" style="color:lightblue; font-size:14px; display:block; margin-bottom:5px;">🔗 {s[:40]}...</a>' for s in f['sites'] if s])
                blocks_html += f"""
                <div class="container-block">
                    <div class="header-wrapper">
                        <div class="header-content">
                            <img src="{f['image']}" width="50" height="50" style="border-radius:50%;">
                            <strong class="header-title">{f['network']}</strong>
                        </div>
                    </div>
                    <div class="footer-wrapper">
                        <p><strong>🪙 Token:</strong> <span style="color:#39FF14;">{f['token']}</span></p>
                        <div style="margin-top:10px;">
                            <strong>🌐 Faucet Links:</strong>
                            <div style="margin-top:8px;">{links_html}</div>
                        </div>
                    </div>
                </div>
                """
            
            full_html = f"""
            <style>
            .container-externa {{ display: flex; flex-wrap: wrap; justify-content: center; font-family: 'Segoe UI', sans-serif; gap: 15px; }}
            .container-block {{ display: flex; flex-direction: column; align-items: center; }}
            .header-wrapper {{ width: 330px; padding: 25px; margin-top: 10px; border-top: 1px solid rgba(48, 240, 192, 0.2); border-bottom: 1px solid #00e0ff; border-top-left-radius: 40px; border-top-right-radius: 10px; background: #1E1F25; display: flex; justify-content: center; }}
            .header-title {{ font-size: 20px; color: lightblue; text-shadow: 0 0 4px #14ffe9; }}
            .header-content {{ display: flex; align-items: center; gap: 15px; }}
            .footer-wrapper {{ width: 330px; padding: 25px; margin-top: 6px; margin-bottom: 20px; border-top: 1px solid #00e0ff; border-bottom-left-radius: 10px; border-bottom-right-radius: 40px; background: #1E1F25; font-size: 16px; color: white; min-height: 180px; }}
            </style>
            <div class="container-externa">{blocks_html}</div>
            """
        else:
            # --- GERAÇÃO LISTA (TABELA) ---
            rows_html = ""
            for f in faucet_sites:
                links_html = " ".join([f'<a href="{s}" target="_blank" class="t-btn" style="margin-right:5px; margin-bottom:5px; display:inline-block;">Link {i+1}</a>' for i, s in enumerate(f['sites']) if s])
                rows_html += f"""
                <tr class="t-row">
                    <td><div style="display:flex; align-items:center; gap:12px;"><img src="{f['image']}" width="35" height="35" style="border-radius:50%;"><strong>{f['network']}</strong></div></td>
                    <td style="color:#39FF14; font-weight:bold;">{f['token']}</td>
                    <td><div style="display:flex; flex-wrap:wrap;">{links_html}</div></td>
                </tr>
                """

            full_html = f"""
            <style>
            .t-container {{ background: #111827; padding: 20px; border-radius: 12px; font-family: 'Segoe UI', sans-serif; color: white; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ text-align: left; color: #8293A3; padding: 15px; border-bottom: 1px solid #1f2937; font-size: 13px; text-transform: uppercase; }}
            td {{ padding: 15px; border-bottom: 1px solid #1f2937; font-size: 15px; }}
            .t-row:hover {{ background: #1e293b; transition: 0.2s; }}
            .t-btn {{ background: #1f2937; color: white; padding: 5px 12px; border-radius: 4px; text-decoration: none; border: 1px solid #374151; font-size: 11px; }}
            .t-btn:hover {{ border-color: #39FF14; color: #39FF14; }}
            </style>
            <div class="t-container">
                <table>
                    <thead>
                        <tr>
                            <th width="30%">Network</th>
                            <th width="30%">Token</th>
                            <th width="40%">Available Links</th>
                        </tr>
                    </thead>
                    <tbody>{rows_html}</tbody>
                </table>
            </div>
            """

        # 4. Renderização
        h_calc = len(faucet_sites) * 100 + 300 if view_mode == "List Table" else 3500
        components.html(full_html, height=max(h_calc, 800), width=1500, scrolling=False)
        
    elif st.session_state.pagina == "⛔ Revoke Contract":

        st.markdown(
             """
                <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 22px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }

                /* Borda neon com gradiente animado */
                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }

                /* Efeito de hover */
                .airdrop-box:hover {
                    border-color: #00f0ff;
                    background: #262b33;
                }

                @keyframes neonBorder {
                    0%   { background-position: 0% 50%; }
                    50%  { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }

                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 5px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 10px;
                    margin-bottom: 5px;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 5px;
                }
            </style>
            <div class="airdrop-box">
                <h2>Why Revoke Contracts?</h2>
                <p style="color: #8293A3;>Revoke protocols let you remove permissions previously granted to smart contracts in your crypto wallet.
                When using DApps, these permissions often stay active indefinitely.</p>
                <p style="color: #8293A3;>Revoking them enhances security and control, preventing malicious contracts from accessing your assets without your approval.</p>
            </div>
            """,
            unsafe_allow_html=True
        )
        
        protocols_revoke= {
            "EVM": [
                {"name": "Revoke Cash", "site": "https://revoke.cash/", "image": "https://pbs.twimg.com/profile_images/1948277825812742144/O--5n9jZ_400x400.jpg"},
                {"name": "De.Fi Shield", "site": "https://de.fi/shield", "image": "https://pbs.twimg.com/profile_images/1896908731180359680/WoOKUzJ5_400x400.jpg"},
            ],
            "SOLANA": [
                {"name": "Famous Foxes", "site": "https://famousfoxes.com/revoke",  "image": "https://pbs.twimg.com/profile_images/1433087419046367235/uFYaQEsU_400x400.jpg"},
                {"name": "Solrevoker", "site": "https://solrevoker.com/", "image": "https://pbs.twimg.com/profile_images/1877702623014580224/t9mvXaqU_400x400.jpg"},
            ],
        }
        # CSS para aumentar o tamanho da fonte do selectbox
        st.markdown("""
            <style>
            .select-row {
                display: flex;
                align-items: center;
                gap: 20px;
                margin-left: 5px;
            }
            .select-label {
                font-size: 25px;
                font-weight: bold;
                color: white;
                font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            }
            div[data-baseweb="select"] {
                font-size: 22px !important;
                font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                min-height: 50px !important;
                max-width: 200px;
                margin-left: 5px;
            }
            div[data-baseweb="select"] > div > div {
                font-size: 20px !important;
            }
            </style>
        """, unsafe_allow_html=True)

        # Usa um container horizontal com label e selectbox
        st.markdown('<div class="select-row"><div class="select-label">Select Network:</div>', unsafe_allow_html=True)
        selected_network = st.selectbox(
            label="",  # Remove o label padrão
            options=list(protocols_revoke.keys()),
            format_func=lambda x: x.upper(),
            key="network_selectbox"  # necessário para evitar conflitos de renderização
        )
        st.markdown("</div>", unsafe_allow_html=True)
        # Style setup
        st.markdown("""
            <style>
                body {background-color: #c3a9a5;}
                .title-style {
                    font-size: 36px;
                    font-weight: bold;
                    color: #fbd46d;
                    margin-bottom: 5px;
                }
                .card {
                    border: 3px solid white;
                    border-radius: 10px;
                    padding: 10px;
                    background-color: #2c4a6b;
                    color: white;
                    margin: 10px;
                }
            </style>
        """, unsafe_allow_html=True)

        # Network selection
        #selected_network = st.selectbox("Select a network:", list(protocols_bridge_swap.keys()), format_func=lambda x: x.upper())

        # Render selected network section
        st.markdown("""
            <style>
                .custom-header {
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    font-size: 25px;
                    font-weight: 600;
                    color: white;
                    margin-bottom: 5px;
                    margin-left: 20px;
                }
            </style>
        """, unsafe_allow_html=True)

        # Exibir texto com a fonte Sora
        st.markdown(f'<div class="custom-header">   {selected_network.upper()}</div>', unsafe_allow_html=True)
        
        # Gera os blocos HTML individualmente
        blocks_html = ""
        for protocol in protocols_revoke[selected_network]:
            blocks_html += f"""
            <div class="container-block">
                <a href="{protocol['site']}" target="_blank" class="footer-link">
                    <div class="footer-wrapper">
                        <img src="{protocol['image']}" width="50" height="50" style="border-radius: 50%;">
                        <strong style="text-shadow: 0 0 4px #14ffe9, 0 0 4px #14ffe9;">{protocol['name']}</strong>
                    </div>
                </a>
            </div>
            """
        

        # HTML completo
        full_html = f"""
        <style>
        @keyframes neonBorder {{
            0%   {{ background-position: 0% 50%; }}
            50%  {{ background-position: 100% 50%; }}
            100% {{ background-position: 0% 50%; }}
        }}
        .container-externa {{
            border-radius: 12px;
            padding: 25px;
            margin-top: 30px;
            gap: 0px;
            display: flex;
            justify-content: left;
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            font-size: 22px;
            color: white;
            margin: 0px 0;
        }}
        .protocol-block {{
            width: 1272px;
            border-radius: 12px;
            padding: 25px;
            margin-top: 0px;
            background-color: rgba(35, 36, 41, 0.85);
            display: flex;
            justify-content: flex-start;
            gap: 0px; /* 👈 distância fixa entre os blocos internos */
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            font-size: 22px;
            color: white;
            margin: 0px 0;
        }}
        .footer-wrapper {{
            width: 300px;
            padding: 30px;
            margin-right: 30px;
            border-top: 1px solid rgba(48, 240, 192, 0.2);
            border-top-left-radius: 40px;
            border-top-right-radius: 10px;  /* 👈 maior */
            border-bottom-left-radius: 10px;
            border-bottom-right-radius: 40px;
            display: flex;
            align-items: center;
            gap: 30px;
            background: #1E1F25;
            box-shadow: 0 0 30px rgba(0,255,150,0.1);
            transition: transform 0.3s ease;
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            z-index: 1;
            overflow: hidden;
            align-items: center;
            justify-content: center;
        }}
        .footer-wrapper:hover {{
            transform: scale(1.02);
            border: 1px solid #39ff14; /* Verde neon */
            border-top-left-radius: 40px;
            border-top-right-radius: 10px; 
            border-bottom-left-radius: 10px;
            border-bottom-right-radius: 40px;
            box-shadow: 0 0 4px #39ff14, 0 0 8px #39ff14; /* Brilho neon suave */
            background: #262b33;
        }}
        .footer-wrapper:hover::before {{
            content: "";
            position: absolute;
            top: -3px;
            left: -3px;
            right: -3px;
            bottom: -3px;
            border-radius: 14px;
            z-index: -1;
            animation: neonBorder 6s ease infinite;
            -webkit-mask:
                linear-gradient(#fff 0 0) content-box,
                linear-gradient(#fff 0 0);
            -webkit-mask-composite: xor;
            mask-composite: exclude;
        }}
        .footer-wrapper a {{
            color: lightblue;
            text-decoration: none;
        }}
        .footer-link {{
            text-decoration: none;
            margin: 0;
            width: 100%;
            color: inherit;
        }}
        .footer-link:hover {{
            color: inherit;
        }}
        </style>

        <div class="container-externa">
            {blocks_html}
        </div>
        """
        components.html(full_html, height=400, width=1300, scrolling=True)

    elif st.session_state.pagina == "⚠️ Avoiding Scams":
            st.markdown(
                """
                <style>
                .airdrop-box {
                    position: relative;
                    z-index: 1;
                    border-radius: 12px;
                    padding: 25px;
                    margin: 20px 0;
                    background: #111827;
                    display: flex;
                    flex-direction: column;
                    gap: 30px;
                    font-size: 20px;
                    color: white;
                    font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
                    overflow-wrap: break-word;
                    word-wrap: break-word;
                    white-space: normal;
                    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.4);
                }

                /* Borda neon com gradiente animado */
                .airdrop-box::before {
                    content: "";
                    position: absolute;
                    top: -3px;
                    left: -3px;
                    right: -3px;
                    bottom: -3px;
                    border-radius: 14px;
                    z-index: -1;
                    background: linear-gradient(270deg, #00F0FF, #39FF14, #00F0FF);
                    background-size: 600% 600%;
                    animation: neonBorder 6s ease infinite;
                    padding: 3px;
                    -webkit-mask:
                        linear-gradient(#fff 0 0) content-box,
                        linear-gradient(#fff 0 0);
                    -webkit-mask-composite: xor;
                    mask-composite: exclude;
                }

                /* Efeito de hover */
                .airdrop-box:hover {
                    border-color: #00f0ff;
                    background: #262b33;
                }

                @keyframes neonBorder {
                    0%   { background-position: 0% 50%; }
                    50%  { background-position: 100% 50%; }
                    100% { background-position: 0% 50%; }
                }

                .airdrop-box h1 {
                    font-size: 25px;
                    text-align: center;
                    margin-bottom: 0px;
                }

                .airdrop-box h2 {
                    font-size: 25px;
                    margin-top: 5px;
                    margin-bottom: 0px;
                }

                .airdrop-box ul {
                    margin-left: 20px;
                    margin-bottom: 0px;
                }
            </style>

            <div class="airdrop-box">
                    <h2><strong>With the rise of crypto airdrops, scams have become more widespread than ever.</strong></h2>
                    <p style="color: #8293A3;">
                    Many users rush to be among the first to discover new projects or engage with X posts to secure rewards. But that urgency can come at a high cost.
                    Without careful verification, it’s easy to fall into traps set by scammers.
                    </p>
                    <p style="color: #8293A3;">They’ll use every trick in the book — from fake comments under posts to direct messages promising “great opportunities” — all designed to trick you and steal your funds.
                    </p>
                    <h2>🚨 How to Avoid Crypto Airdrop Scams?</h2>
                    <p style="color: #8293A3;">Crypto airdrops can be a great way to earn rewards—but they're also a big target for scammers. Here's a quick guide to staying safe while chasing drops.</p>
                    <h2>🔒 1. Use a Dedicated Wallet for Airdrops</h2>
                    <ul style="color: #8293A3;">
                        <li><b>Create a separate wallet</b> just for airdrop interactions. Consider a Cold Wallet with at least 3 accounts by network.</li>
                        <li><b>Never use your main wallet</b> with valuable assets.</li>
                        <li>If anything goes wrong, your main funds stay safe.</li>
                    </ul>
                    <h2>🧠 2. Research Before You Click</h2>
                    <ul style="color: #8293A3;">
                        <li><b>Verify the project’s legitimacy</b> through <b>official sites</b> and <b>trusted communities</b>.</li>
                        <li><b>Avoid links</b> from random X (Twitter) users or Telegram DMs.</li>
                        <li><b>Look for audits, GitHub repos, and real backers.</b></li>
                    </ul>
                    <h2>⚠️ 3. Beware of Fake Accounts and Bots</h2>
                    <ul style="color: #8293A3;">
                        <li>Many scam comments and impersonators exist under legit posts.</li>
                        <li>Double-check usernames and links — <b>look for subtle typos</b>.</li>
                        <li><b>Never trust DMs</b> offering "airdrops" or "early access".</li>
                    </ul>
                    <h2>🧾 4. Audit the Smart Contract or Wait</h2>
                    <ul style="color: #8293A3;">
                        <li>Don’t rush to sign random transactions.</li>
                        <li>Use tools like <b>Etherscan</b>, <b>DeBank</b>, or <b>Rabby Wallet</b> to inspect them.</li>
                        <li>If a project lacks transparency or an audit, think twice.</li>
                    </ul>
                    <h2>🧼 5. Revoke Unused Permissions</h2>
                    <ul style="color: #8293A3;">
                        <li>Clean up your wallet permissions regularly with:</li>
                        <li><a href="https://revoke.cash" target="_blank">revoke.cash</a></li>
                        <li><a href="https://app.safe.global" target="_blank">Safe</a></li>
                        <li>This reduces the chance of malicious token drains.</li>
                    </ul>
                    <h2>🔐 6. Never Share Private Keys or Seed Phrases</h2>
                    <ul style="color: #8293A3;">
                        <li><b>No legit team will ever ask</b> for your keys or phrase.</li>
                        <li>Use <b>hardware wallets</b> like Ledger or Trezor for serious funds.</li>
                    </ul>
                    <h2>🪪 7. Watch Out for “Connect Wallet to Check Eligibility”</h2>
                    <ul style="color: #8293A3;">
                        <li>Don't connect your wallet to <b>random sites</b>.</li>
                        <li>Always verify the domain and source.</li>
                        <li>If in doubt, don’t click!</li>
                    </ul>
                    <h2>🧾 8. Use Reputable Airdrop Aggregators</h2>
                    <ul style="color: #8293A3;">
                        <li>Use vetted aggregators to spot real airdrops:</li>
                        <li><a href="https://earni.fi/" target="_blank">Earnifi</a></li>
                        <li><a href="https://defillama.com/airdrops" target="_blank">DeFiLlama Airdrops</a></li>
                        <li>Still verify manually before connecting wallets or claiming anything.</li>
                    </ul>
                    <div style="background-color: #FF4500; padding: 16px; border-radius: 8px; border-left: 6px solid #FF4500; font-size: 25px; font-weight: bold; margin-top: 30px; color: white;">
                        Stay tuned, stay safe, and enjoy the airdrop hunt!
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

    # Fecha container externo
    st.markdown('</div>', unsafe_allow_html=True)

### --- Footer ---
st.markdown("""
    <style>
        .main, .block-container {
            padding-left: 0 !important;
            padding-right: 0 !important;
            flex: 1 1 auto;
        }

        section.main > div {
            padding: 0 !important;
            margin: 0 !important;
        }

        .footer-wrapper {
            position: relative;
            left: 0;
            right: 0;
            width: 100vw;
            margin-left: calc(-50vw + 50%);
            margin-top: 80px;
            padding: 40px 0;
            border-top: 1px solid rgba(48, 240, 192, 0.2);
            
            /* NOVO background com neon verde estilo RubyScore */
            background: radial-gradient(circle at right, rgba(0,255,150,0.25), #0a0f0c 60%);
            box-shadow: 0 0 60px rgba(0,255,150,0.1);
        }

        .footer-content {
            max-width: 1800px;
            margin: 30px auto;
            padding: 0 60px;
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            grid-gap: 30px;
            color: #30f0c0;
            font-weight: bold;
            font-size: 18px;
            font-family: 'Trebuchet MS', 'Segoe UI', sans-serif;
            margin-bottom: 175px
        }

        .footer-content p {
            margin: 0;
        }

        .footer-content a {
            color: #00ffae;
            text-decoration: none;
            font-weight: bold;
            display: inline-block;
            transition: transform 0.2s ease, text-decoration 0.2s ease;
        }

        .footer-content a:hover {
            text-decoration: underline;
            transform: scale(1.05);
        }
    </style>

    <div class="footer-wrapper">
        <div class="footer-content">
            <div style="font-size:22px;">
                <p>🪂 Airdrops Monitor</p>
            </div>
            <div style="text-align: center; color: white;">
                Made with ❤️ by <a href="https://x.com/CaioFlemin2089" target="_blank">@CaioFleming</a> |
                Follow for more airdrop farming strategies 🚀<br>
                Stay updated on Twitter! 📢
                <div style="font-size:18px; margin-top: 10px;">
                    Disclaimer: This is not financial advice. Always do your own research (DYOR)!
                </div>
            </div>
            <div style="text-align: center; font-size:22px; color: #30f0c0">
                <p>Community:</p>
                <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABwAAAAcCAAAAABXZoBIAAAA/0lEQVR4AbXPIazCMACE4d+L2qoZFEGSIGcRc/gJJB5XMzGJmK9EN0HMi+qaibkKVF1txdQe4g0YzPK5yyWXHL9TaPNQ89LojH87N1rbJcXkMF4Fk31UMrf34hm14KUeoQxGArALHTMuQD2cAWQfJXOpgTbksGr9ng8qluShJTPhyCdx63POg7rEim95ZyR68I1ggQpnCEGwyPicw6hZtPEGmnhkycqOio1zm6XuFtyw5XDXfGvuau0dXHzJp8pfBPuhIXO9ZK5ILUCdSvLYMpc6ASBtl3EaC97I4KaFaOCaBE9Zn5jUsVqR2vcTJZO1DdbGoZryVp94Ka/mQfE7f2T3df0WBhLDAAAAAElFTkSuQmCC" width="24" height="24" style="border-radius: 50%; vertical-align: middle;">
                <a href="https://twitter.com/" target="_blank">Twitter</a>
                <p><a href="https://discordbot.streamlit.app/" target="_blank">🤖 Discord Bot</a></p>
            </div>
        </div>
    </div>
""", unsafe_allow_html=True)
